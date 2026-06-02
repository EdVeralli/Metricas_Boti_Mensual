# -*- coding: utf-8 -*-
"""
Temas Consultados - Metricas mensuales de Boti

Reemplaza el flujo manual del notebook 'mensajes.ipynb' del repo
Tablero-de-mensajes-y-disparadores:
  - Antes: descargar CSV de Athena a mano + correr notebook celda por celda.
  - Ahora: un solo script que lee fechas del config, consulta Athena,
           clasifica mensajes por categoria y genera todos los outputs.

Genera:
  - output/temas_consultados_<mes>_<año>.csv        Dataset crudo limpio
  - output/top_200_mensajes_<mes>_<año>.csv         Top 200 mensajes mas repetidos
  - output/comparison_mensajes_<mes>_<año>.xlsx     Comparacion mes vs mes anterior
  - output/reporte_mensajes_<mes>_<año>.xlsx        Comparativo + Top100 Otros
  - output/top_variaciones_<mes>_<año>.xlsx         Top 3 subidas/bajadas por categoria
  - output/ranking_temas_<mes>_<año>.txt            Texto formateado para el TABLERO
                                                    (celda 'Temas mas consultados')

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBADataScientist
"""

import boto3
import awswrangler as wr
import pandas as pd
import numpy as np
import os
import sys
import time
import re
import string
import unicodedata
from datetime import datetime
from calendar import monthrange
from collections import Counter
from contextlib import contextmanager
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# NLTK - stopwords en español (puede tardar la primera vez al bajar el pack)
try:
    from nltk.corpus import stopwords
    _ = stopwords.words('spanish')  # forzar descarga si falta
except LookupError:
    import nltk
    nltk.download('stopwords', quiet=True)
    from nltk.corpus import stopwords

# ============================================================================
# HELPERS DE LOGGING
# ============================================================================

def log(msg=""):
    """Print con timestamp y flush inmediato (clave en consolas Windows)."""
    if msg:
        print("[{:%H:%M:%S}] {}".format(datetime.now(), msg), flush=True)
    else:
        print("", flush=True)

@contextmanager
def step(nombre):
    """Context manager: imprime inicio/fin de un paso pesado con tiempo elapsed."""
    t0 = time.time()
    log(">>> INICIO: {}".format(nombre))
    try:
        yield
    finally:
        log("<<< FIN:    {} ({:.1f}s)".format(nombre, time.time() - t0))

# ============================================================================
# CONFIGURACION
# ============================================================================

CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': '../config_fechas.txt',
}

# ============================================================================
# CATEGORIAS Y CLASIFICACION (replica exacta del notebook mensajes.ipynb)
# ============================================================================

CATEGORIAS = {
    "Salud (Médicos/Salud)": [
        "salud", "medico", "medicos", "hospital", "clinica", "doctor", "turno medico",
        "guardia", "salu", "medic", "cedetac", "oftalmologia", "resonancia",
        "mamografia", "odontologia", "dermatologia", "tomografia", "traumatologia",
        "cesac", "cardiologia", "gastroenterologia", "neurologia", "ginecologia",
        "endocrinologia", "ecografia", "pediatria", "laboratorio", "dentista",
        "oftalmologo", "alergia"
    ],
    "Turnos (Generales)": [
        "turno", "cita", "reserva", "agenda", "sacar turno", "pedir turno", "turn",
        "turnos"
    ],
    "Infracciones y Multas": [
        "multa", "multas", "infraccion", "infracsion", "infraciones", "contravencion",
        "sancion", "controlador", "audiencia", "libre deuda", "audiencia virtual",
        "descargo", "pago voluntario", "plan pago vencido", "puntos", "scoring",
        "charla virtual"
    ],
    "Movilidad + Estacionamiento": [
        "estacionamiento", "estaciona", "parquimetro", "parquimetros", "garage",
        "garaje", "donde estacionar", "estaciono", "transito", "subte"
    ],
    "Licencia de Conducir": [
        "licencia", "licensia", "carnet", "registro", "renovar licencia",
        "licencia conducir", "lisencia", "lisensia", "otorgamiento", "renovacion",
        "pista manejo", "cenat", "pista aprendizaje", "vencida"
    ],
    "Interacción Humana (Chatear)": [
        "chat", "chatear", "hablar", "conversar", "atencion humana",
        "atencion persona", "hablar humano", "chatear alguien", "ayuda",
        "operador", "asesor", "147", "humano", "linea 102", "108"
    ],
    "Educación (Inscripciones/Becas)": [
        "educacion", "educar", "beca", "becas", "inscripcion", "inscrip",
        "colegio", "universidad", "secundario", "primario", "secundaria",
        "primaria", "terciario", "terciaria", "curso", "cursos", "capacitaciones",
        "boleto estudiantil", "consultar vacante escolar", "vacante", "adultos 2000",
        "buepp", "boleto universitario", "escuela", "boleto", "talento tech",
        "ciudad bilingue"
    ],
    "MIBA y Validación": [
        "miba", "validacion", "tramite digital", "trámite digital", "valida",
        "identidad", "mi ba", "actualizar datos cuil", "blanqueo cuenta",
        "recuperacion cuenta", "cuenta", "crear cuenta", "ingresar",
        "olvide contrasena", "recuperar contrasena", "contrasena", "nivel",
        "blanqueo", "perfil", "clave"
    ],
    "Impuestos (AGIP / Patentes / ABL)": [
        "impuesto", "impuestos", "agip", "patente", "patentes", "abl", "tributo",
        "tasas", "tax", "boleta", "pagar", "comprobante plan pago vencido",
        "ingresos brutos", "moratoria", "deuda"
    ],
    "Telepase": [
        "telepase", "telepas", "tele pase", "peaje", "pase electronico",
        "pase digital", "ausa", "telepeaje"
    ],
    "DNI y Partidas": [
        "dni", "documento", "partida", "partidas", "certificado nacimiento",
        "certif nac", "pasaporte", "casamiento", "matrimonio", "acta nacimiento"
    ],
    "Trámites y Reclamos": [
        "tramite", "tramites", "tramit", "tram", "gestión", "gestiones",
        "tramitacion", "tramitando", "tad", "ciudadania portena", "cuidadania portena",
        "ciudadania", "rendir recibo", "escombros", "certificado domicilio",
        "vtv", "denuncia vial", "retiro escombros", "consultar detalles inspeccion local",
        "consultar solicitud", "denuncia", "habitacional",
        "hacer reclamo cierre pista practica manejo", "reclamo", "basura", "auto",
        "castracion", "sube", "subsidio habitacional", "vehiculo",
        "verificacion policial", "poda", "recoleccion programada", "cambio domicilio",
        "ruidos molestos", "cud", "defensa consumidor", "voluminosos", "trabajo",
        "residuos", "escombro", "bui", "vereda", "contenedor", "recoleccion",
        "solicitud", "retiro escombro", "acarreo", "residuos voluminosos", "moto",
        "constancia domicilio", "discapacidad", "mascotas", "higiene urbana",
        "limpieza", "reportar error", "ivc", "empleo", "manipulacion alimentos",
        "declaracion jurada", "docente", "desratizacion", "antecedentes penales",
        "retiro programado", "grua"
    ],
    "Turismo + Cultura + Deporte": [
        "turismo ciudad", "recorrer bam", "deportes", "natacion", "pileta",
        "polideportivo", "deporte", "turismo", "eventos"
    ],
    "Vacunación": [
        "vacuna antigripal", "vacunas", "vacuna", "vacunacion", "antigripal",
        "vacuna gripe", "vacunacion antigripal", "dengue", "covid", "gripe",
        "vacuna covid", "vacuna dengue", "fiebre amarilla"
    ]
}

ORDEN_PRIORIDAD = [
    "Salud (Médicos/Salud)",
    "Vacunación",
    "Infracciones y Multas",
    "Licencia de Conducir",
    "Telepase",
    "Movilidad + Estacionamiento",
    "MIBA y Validación",
    "Educación (Inscripciones/Becas)",
    "Impuestos (AGIP / Patentes / ABL)",
    "DNI y Partidas",
    "Turismo + Cultura + Deporte",
    "Interacción Humana (Chatear)",
    "Turnos (Generales)",
    "Trámites y Reclamos",
]

# Categorias que se excluyen del ranking top 10 final del TABLERO
# - 'Otros' es el cajon de sastre
# - 'Vacunacion' suele ser estacional; el equipo decide si la incluye o no
CATEGORIAS_EXCLUIR_DEL_RANKING = [
    'Otros',
    'Vacunación',
]
TOP_N_RANKING = 10

# ============================================================================
# LIMPIEZA DE TEXTO
# ============================================================================

STOPWORDS_CUSTOM = [
    'boti', 'hola', 'si', 'no', 'aca', 'ahi', 'buen', 'dia', 'buenos', 'dias',
    'buenas', 'tardes', 'que', '¡hola', 'quiero', 'saber', 'puedo', 'necesito',
    'cual', 'muchas', 'voy', 'vos', 'gracias'
]

def _build_stopwords_set():
    base = stopwords.words('spanish')
    return set(base + STOPWORDS_CUSTOM)

_STOP = _build_stopwords_set()

def limpiar_mensaje(message):
    """Lowercase + sin puntuacion + sin acentos + sin stopwords + sin tokens <=2 letras."""
    if not isinstance(message, str):
        return ''
    msg = message.lower()
    msg = msg.translate(str.maketrans('', '', string.punctuation))
    msg = unicodedata.normalize('NFKD', msg).encode('ascii', 'ignore').decode('utf-8', 'ignore')
    palabras = [w for w in msg.split() if w not in _STOP and len(w) > 2]
    return ' '.join(palabras)

def es_mensaje_basura(message):
    """True si el mensaje es basura (test/clean/reset/@/longitud <=2)."""
    if not isinstance(message, str):
        return True
    m = message.strip()
    if len(m) <= 2:
        return True
    if m.startswith('test') or m.startswith('@') or m.startswith('clean') or m.startswith('reset'):
        return True
    return False

def clasificar(texto):
    """Asigna la categoria de un mensaje siguiendo el orden de prioridad."""
    if not isinstance(texto, str):
        return "Otros"
    t = texto.lower()
    for cat in ORDEN_PRIORIDAD:
        for palabra in CATEGORIAS[cat]:
            if palabra in t:
                return cat
    return "Otros"

# ============================================================================
# CONFIG DE FECHAS
# ============================================================================

NOMBRES_MESES = ['', 'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
                 'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre']
ABREV_MESES = {1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr', 5: 'may', 6: 'jun',
               7: 'jul', 8: 'ago', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'}

def leer_config_fechas():
    """Lee config_fechas.txt y devuelve (mes, anio). Solo soporta MES+AÑO."""
    config_file = CONFIG['config_file']
    if not os.path.exists(config_file):
        raise FileNotFoundError("No se encuentra {}".format(config_file))

    mes = None
    anio = None
    with open(config_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#'):
                continue
            if line.startswith('MES='):
                mes = int(line.split('=')[1].strip())
            elif line.startswith('AÑO=') or line.startswith('ANO=') or line.startswith('ANIO='):
                anio = int(line.split('=')[1].strip())

    if mes is None or anio is None:
        raise ValueError("config_fechas.txt debe tener MES y AÑO definidos")
    if mes < 1 or mes > 12:
        raise ValueError("MES invalido: {}".format(mes))
    return mes, anio

def calcular_mes_anterior(mes, anio):
    if mes == 1:
        return 12, anio - 1
    return mes - 1, anio

def fechas_rango(mes, anio):
    """Devuelve (fecha_inicio_mes_anterior, fecha_fin_ultimo_mes) en 'YYYY-MM-DD'."""
    mes_ant, anio_ant = calcular_mes_anterior(mes, anio)
    fecha_inicio = "{:04d}-{:02d}-01".format(anio_ant, mes_ant)
    ultimo_dia = monthrange(anio, mes)[1]
    fecha_fin = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
    return fecha_inicio, fecha_fin

# ============================================================================
# ATHENA
# ============================================================================

def build_query(fecha_inicio, fecha_fin):
    """Replica la query del repo Tablero-de-mensajes-y-disparadores pero con
    rango de fechas explicito (2 meses: mes anterior + ultimo mes)."""
    return """SELECT DISTINCT session_id,
       DATE(creation_time) as creation_time,
       id,
       message
FROM "caba-piba-consume-zone-db"."boti_message_metrics_2"
WHERE CAST(session_creation_time AS DATE) BETWEEN date '{fi}' AND date '{ff}'
  AND msg_from = 'user'
  AND message IS NOT NULL
  AND message_type = 'Text'
  AND LENGTH(message) > 2
  AND COMPREHENSION_TYPE != 'Answer to question'""".format(fi=fecha_inicio, ff=fecha_fin)

def check_aws_credentials():
    """Verifica que las credenciales AWS esten configuradas y el rol sea correcto."""
    try:
        sts = boto3.client('sts', region_name=CONFIG['region'])
        identity = sts.get_caller_identity()
        user_arn = identity.get('Arn', '')
        log("[OK] Credenciales AWS validas")
        log("     ARN: {}".format(user_arn))
        if 'PIBADataScientist' not in user_arn:
            log("[ADVERTENCIA] No estas usando el rol correcto")
            log("     Se requiere: PIBADataScientist")
            log("     Ejecutar: aws-azure-login --profile default --mode=gui")
            log("     y SELECCIONAR el rol PIBADataScientist")
            return False
        return True
    except Exception as e:
        log("[ERROR] Error verificando credenciales: {}".format(e))
        if 'ExpiredToken' in str(e):
            log("     Tu sesion AWS expiro. Ejecuta:")
            log("     aws-azure-login --profile default --mode=gui")
        return False

def ejecutar_query(query):
    """Corre la query en Athena con manejo robusto de workgroup."""
    session = boto3.Session(region_name=CONFIG['region'])
    try:
        return wr.athena.read_sql_query(
            sql=query,
            database=CONFIG['database'],
            workgroup=CONFIG['workgroup'],
            boto3_session=session,
            ctas_approach=False,
            unload_approach=False,
        )
    except Exception as e:
        if 'workgroup' in str(e).lower():
            log("[ADVERTENCIA] Reintentando sin workgroup...")
            return wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False,
            )
        raise

# ============================================================================
# PROCESAMIENTO
# ============================================================================

def procesar_df(df):
    """Limpia + clasifica + agrega 'link' al DataFrame en bruto."""
    df = df.copy()

    # Eliminar mensajes basura
    log("    Filas iniciales: {:,}".format(len(df)))
    mask_basura = df['message'].apply(es_mensaje_basura)
    df = df[~mask_basura].copy()
    log("    Filas tras filtro de basura (test/@/clean/reset/<=2): {:,}".format(len(df)))

    # Limpiar texto
    df['mensaje_ok'] = df['message'].apply(limpiar_mensaje)

    # Eliminar los que quedaron vacios tras limpieza
    df = df[df['mensaje_ok'].str.strip().astype(bool)].copy()
    log("    Filas tras limpieza de texto: {:,}".format(len(df)))

    # Link a botmaker (los primeros 20 chars del session_id)
    df['link'] = 'https://go.botmaker.com/#/chats/' + df['session_id'].astype(str).str[:20]

    # Categoria
    df['categoria'] = df['mensaje_ok'].apply(clasificar)

    # Periodo
    df['creation_time'] = pd.to_datetime(df['creation_time'])
    df['mes'] = df['creation_time'].dt.to_period('M')
    return df

def calcular_comparison(df, mes_ant_str, ultimo_mes_str):
    """Comparacion mes vs mes anterior por mensaje individual (no por categoria)."""
    df_last = df[df['mes'].astype(str) == ultimo_mes_str]
    df_prev = df[df['mes'].astype(str) == mes_ant_str]

    cnt_last = Counter(df_last['mensaje_ok'])
    cnt_prev = Counter(df_prev['mensaje_ok'])
    tot_last = sum(cnt_last.values()) or 1
    tot_prev = sum(cnt_prev.values()) or 1

    rows = []
    for msg in set(cnt_last.keys()).union(cnt_prev.keys()):
        c_last = cnt_last.get(msg, 0)
        c_prev = cnt_prev.get(msg, 0)
        if c_prev > 0:
            var = ((c_last - c_prev) / c_prev) * 100
            status = 'existed'
            if c_last == 0:
                status = 'disappeared'
        else:
            var = float('inf')
            status = 'new'
        rows.append({
            'message': msg,
            'count_last_30_days': c_last,
            'percentage_last_30_days': round(c_last / tot_last * 100, 5),
            'count_before_last_30_days': c_prev,
            'percentage_before_last_30_days': round(c_prev / tot_prev * 100, 5),
            'variation_percentage': round(var, 5) if var != float('inf') else float('inf'),
            'status': status,
        })

    df_cmp = pd.DataFrame(rows)
    # Filtrar igual que el notebook: solo mensajes con >=10 en ambos periodos
    df_cmp = df_cmp[(df_cmp['count_last_30_days'] >= 10) &
                    (df_cmp['count_before_last_30_days'] >= 10)]
    return df_cmp

def calcular_conteo_categorias(df, mes_ant_period, ultimo_mes_period):
    """Conteo por categoria y mes. Devuelve (conteo_largo, pivot_resultado)."""
    df_mes = df[df['mes'].isin([mes_ant_period, ultimo_mes_period])].copy()
    df_mes['mes_str'] = df_mes['mes'].astype(str)
    conteo = df_mes.groupby(['mes_str', 'categoria']).size().reset_index(name='Q')

    mes_ant_str = str(mes_ant_period)
    ultimo_mes_str = str(ultimo_mes_period)

    pivot = conteo.pivot(index='categoria', columns='mes_str', values='Q').fillna(0)

    # Asegurar que las 2 columnas existan
    if mes_ant_str not in pivot.columns:
        pivot[mes_ant_str] = 0
    if ultimo_mes_str not in pivot.columns:
        pivot[ultimo_mes_str] = 0

    pivot['Var %'] = np.where(
        pivot[mes_ant_str] > 0,
        (pivot[ultimo_mes_str] - pivot[mes_ant_str]) / pivot[mes_ant_str] * 100,
        np.nan
    )

    pivot = pivot.sort_values(by=ultimo_mes_str, ascending=False).reset_index()
    pivot.index = pivot.index + 1
    pivot['Q Último Mes'] = pivot[ultimo_mes_str].apply(lambda x: "{:.1f}K".format(x / 1000))
    pivot['Q Mes Anterior'] = pivot[mes_ant_str].apply(lambda x: "{:.1f}K".format(x / 1000))
    pivot['Var %'] = pivot['Var %'].apply(lambda x: "NUEVO" if pd.isna(x) else "{:+.1f}%".format(x))

    resultado = pivot[['categoria', 'Q Último Mes', 'Var %', 'Q Mes Anterior']].copy()
    resultado.columns = ['Tema Último Mes', 'Q Mensajes', 'Var %', 'Q Mes Ant.']
    return conteo, resultado

def calcular_top_variaciones(df, mes_ant_period, ultimo_mes_period):
    """Top 3 al alza / a la baja por categoria a nivel mensaje."""
    df_mes = df[df['mes'].isin([mes_ant_period, ultimo_mes_period])]
    conteo_msg = (
        df_mes.groupby(['categoria', 'mensaje_ok', 'mes']).size().reset_index(name='Q')
    )
    pivot_msg = conteo_msg.pivot_table(
        index=['categoria', 'mensaje_ok'],
        columns='mes',
        values='Q',
        fill_value=0,
    ).reset_index()

    if mes_ant_period not in pivot_msg.columns:
        pivot_msg[mes_ant_period] = 0
    if ultimo_mes_period not in pivot_msg.columns:
        pivot_msg[ultimo_mes_period] = 0

    pivot_msg['Var %'] = np.where(
        pivot_msg[mes_ant_period] > 0,
        (pivot_msg[ultimo_mes_period] - pivot_msg[mes_ant_period]) / pivot_msg[mes_ant_period] * 100,
        np.where(pivot_msg[ultimo_mes_period] > 0, np.inf, np.nan),
    )
    pivot_msg = pivot_msg[pivot_msg[ultimo_mes_period] > 100]

    salida = []
    for cat in pivot_msg['categoria'].unique():
        subset = pivot_msg[pivot_msg['categoria'] == cat].copy()
        top_up = subset.sort_values(by='Var %', ascending=False).head(3).copy()
        top_up['Tendencia'] = '↑ Subida'
        top_down = subset.sort_values(by='Var %', ascending=True).head(3).copy()
        top_down['Tendencia'] = '↓ Bajada'
        salida.append(pd.concat([top_up, top_down]))

    if not salida:
        return pd.DataFrame()
    df_out = pd.concat(salida)
    df_out['Var %'] = df_out['Var %'].apply(
        lambda x: 'NUEVO' if x == np.inf else ("{:+.1f}%".format(x) if not pd.isna(x) else '-')
    )
    return df_out

def construir_ranking_texto(resultado_categorias):
    """Toma el DataFrame `resultado` y arma el texto '1-Cat\\n2-Cat\\n...' del TABLERO."""
    df_rank = resultado_categorias[
        ~resultado_categorias['Tema Último Mes'].isin(CATEGORIAS_EXCLUIR_DEL_RANKING)
    ].head(TOP_N_RANKING).reset_index(drop=True)
    lineas = ["{}-{}".format(i + 1, cat) for i, cat in enumerate(df_rank['Tema Último Mes'])]
    return '\n'.join(lineas)

def generar_excel_ranking(filepath, ranking_texto, periodo_label):
    """
    Genera un Excel con la fila 'Temas mas consultados' en formato template:
        Col A: 'Temas mas consultados'  (label)
        Col B: ranking de 10 lineas con wrap_text=True
    Replica el look del template del tablero.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Temas mas consultados'

    # Estilos
    label_font = Font(name='Arial', size=11, bold=True)
    value_font = Font(name='Arial', size=10)
    header_font = Font(name='Arial', size=11, bold=True)
    header_fill = PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid")
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    align_label = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_value = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Header
    ws['A1'] = 'Indicador'
    ws['B1'] = periodo_label
    for c in ('A1', 'B1'):
        ws[c].font = header_font
        ws[c].fill = header_fill
        ws[c].alignment = Alignment(horizontal='center', vertical='center')
        ws[c].border = thin

    # Fila del ranking
    ws['A2'] = 'Temas más consultados'
    ws['A2'].font = label_font
    ws['A2'].alignment = align_label
    ws['A2'].border = thin

    # El valor: texto multilinea (con \n) y wrap_text para que se vean los saltos
    ws['B2'] = ranking_texto
    ws['B2'].font = value_font
    ws['B2'].alignment = align_value
    ws['B2'].border = thin

    # Ajustar tamaños para que se vea como el screenshot del usuario
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 50
    # Alto de la fila 2: 10 lineas x ~15 puntos
    ws.row_dimensions[2].height = 175

    wb.save(filepath)

# ============================================================================
# MAIN
# ============================================================================

def main():
    log("=" * 60)
    log("TEMAS CONSULTADOS - METRICAS MENSUALES DE BOTI")
    log("=" * 60)

    # 1) Leer fechas
    try:
        mes, anio = leer_config_fechas()
    except Exception as e:
        log("[ERROR] {}".format(e))
        return

    mes_ant, anio_ant = calcular_mes_anterior(mes, anio)
    fecha_inicio, fecha_fin = fechas_rango(mes, anio)
    periodo = "{} {}".format(NOMBRES_MESES[mes], anio)
    log("Periodo (ultimo mes): {}".format(periodo))
    log("Mes anterior:         {} {}".format(NOMBRES_MESES[mes_ant], anio_ant))
    log("Rango de Athena:      {} a {}".format(fecha_inicio, fecha_fin))

    # 2) Verificar credenciales AWS
    log("")
    log("Verificando credenciales AWS...")
    if not check_aws_credentials():
        return

    # 3) Construir query
    query = build_query(fecha_inicio, fecha_fin)
    log("")
    log("Query a ejecutar:")
    log(query)

    # 4) Ejecutar query
    log("")
    log("ATENCION: la query escanea boti_message_metrics_2 por 2 meses.")
    log("          Puede tardar VARIOS MINUTOS - es normal.")
    with step("Query Athena (boti_message_metrics_2)"):
        df = ejecutar_query(query)
    log("[OK] {:,} filas descargadas".format(len(df)))

    if len(df) == 0:
        log("[ADVERTENCIA] La query no devolvio filas. Abortando.")
        return

    # 5) Procesar
    log("")
    with step("Procesando mensajes (filtros + limpieza + categorizacion)"):
        df = procesar_df(df)

    # 6) Periodos como Period
    mes_ant_period = pd.Period("{:04d}-{:02d}".format(anio_ant, mes_ant), freq='M')
    ultimo_mes_period = pd.Period("{:04d}-{:02d}".format(anio, mes), freq='M')

    # 7) Setup nombres de archivo
    mes_nombre = NOMBRES_MESES[mes]
    sufijo = "{}_{}".format(mes_nombre, anio)
    out_dir = CONFIG['output_folder']
    os.makedirs(out_dir, exist_ok=True)

    path_csv = os.path.join(out_dir, "temas_consultados_{}.csv".format(sufijo))
    path_top200 = os.path.join(out_dir, "top_200_mensajes_{}.csv".format(sufijo))
    path_cmp = os.path.join(out_dir, "comparison_mensajes_{}.xlsx".format(sufijo))
    path_reporte = os.path.join(out_dir, "reporte_mensajes_{}.xlsx".format(sufijo))
    path_topvar = os.path.join(out_dir, "top_variaciones_{}.xlsx".format(sufijo))
    path_ranking_txt = os.path.join(out_dir, "ranking_temas_consultados_{}.txt".format(sufijo))
    path_ranking_xlsx = os.path.join(out_dir, "ranking_temas_consultados_{}.xlsx".format(sufijo))

    # 8) CSV crudo limpio
    log("")
    with step("Guardando CSV crudo procesado"):
        df_out = df.drop(columns=['mes'])  # mes es un Period que no serializa lindo
        df_out.to_csv(path_csv, index=False, encoding='utf-8-sig')
    log("    [OK] {}".format(path_csv))

    # 9) Top 200 mensajes mas repetidos
    with step("Top 200 mensajes mas repetidos"):
        message_counts = Counter(df['mensaje_ok'])
        top_200 = pd.DataFrame(message_counts.most_common(10000),
                               columns=['message_ok', 'count'])
        top_200.to_csv(path_top200, index=False, encoding='utf-8-sig')
    log("    [OK] {}".format(path_top200))

    # 10) Comparison mes vs mes anterior (a nivel mensaje individual)
    mes_ant_str = str(mes_ant_period)
    ultimo_mes_str = str(ultimo_mes_period)
    with step("Comparison mensajes mes vs mes anterior"):
        df_cmp = calcular_comparison(df, mes_ant_str, ultimo_mes_str)
        df_cmp.to_excel(path_cmp, index=False)
    log("    [OK] {} ({:,} mensajes)".format(path_cmp, len(df_cmp)))

    # 11) Conteo y comparativo por categoria
    with step("Conteo por categoria + comparativo"):
        conteo, resultado = calcular_conteo_categorias(df, mes_ant_period, ultimo_mes_period)
    log("    Conteo por mes/categoria:")
    for _, row in conteo.sort_values(['mes_str', 'categoria']).iterrows():
        log("        {} | {:40s} | {:>10,}".format(row['mes_str'], row['categoria'], int(row['Q'])))

    # 12) reporte_mensajes.xlsx (Comparativo + Top100 Otros)
    with step("Excel reporte (Comparativo + Top100_Otros)"):
        otros_top100 = (
            df[df['categoria'] == 'Otros']
            .groupby('mensaje_ok').size()
            .reset_index(name='Cantidad')
            .sort_values(by='Cantidad', ascending=False)
            .head(100)
        )
        with pd.ExcelWriter(path_reporte) as writer:
            resultado.to_excel(writer, sheet_name='Comparativo', index=True)
            otros_top100.to_excel(writer, sheet_name='Top100_Otros', index=False)
    log("    [OK] {}".format(path_reporte))

    # 13) Top variaciones
    with step("Top 3 variaciones (subida/bajada) por categoria"):
        df_topvar = calcular_top_variaciones(df, mes_ant_period, ultimo_mes_period)
        if len(df_topvar) > 0:
            with pd.ExcelWriter(path_topvar) as writer:
                df_topvar.to_excel(writer, sheet_name='TopVariaciones', index=False)
    log("    [OK] {}".format(path_topvar))

    # 14) Ranking texto y Excel para el tablero
    with step("Ranking texto + Excel para el TABLERO"):
        ranking_texto = construir_ranking_texto(resultado)
        with open(path_ranking_txt, 'w', encoding='utf-8') as f:
            f.write(ranking_texto)
        # Periodo label estilo template: 'mar-26'
        periodo_label = "{}-{}".format(ABREV_MESES[mes], str(anio)[-2:])
        generar_excel_ranking(path_ranking_xlsx, ranking_texto, periodo_label)

    # (La actualizacion automatica de la planilla maestra fue descartada.
    #  El ranking queda solo en output/ranking_temas_<mes>_<anio>.xlsx)

    log("")
    log("=" * 60)
    log("RANKING TOP {} - {}".format(TOP_N_RANKING, periodo.upper()))
    log("=" * 60)
    for linea in ranking_texto.split('\n'):
        log("    " + linea)
    log("=" * 60)
    log("")
    log("Texto guardado en: {}".format(path_ranking_txt))
    log("Excel guardado en: {}".format(path_ranking_xlsx))
    log("Pegalo / copialo en la celda 'Temas mas consultados' del tablero.")

    log("")
    log("=" * 60)
    log("ARCHIVOS GENERADOS en {}/".format(out_dir))
    log("=" * 60)
    for p in [path_csv, path_top200, path_cmp, path_reporte, path_topvar, path_ranking_txt, path_ranking_xlsx]:
        log("    [OK] {}".format(os.path.basename(p)))

    log("")
    log("=" * 60)
    log("PROCESO COMPLETADO EXITOSAMENTE")
    log("=" * 60)

# ============================================================================
# EJECUCION
# ============================================================================

if __name__ == "__main__":
    try:
        sys.stdout.reconfigure(line_buffering=True)
        sys.stderr.reconfigure(line_buffering=True)
    except AttributeError:
        pass

    print("")
    print("=" * 60)
    print("  IMPORTANTE: LOGIN AWS REQUERIDO")
    print("=" * 60)
    print("  Antes de continuar, asegurate de estar logueado en AWS.")
    print("  Si tu token expiró, ejecutá en otra terminal:")
    print("")
    print("      aws-azure-login --profile default --mode=gui")
    print("")
    print("  Rol requerido: PIBADataScientist")
    print("=" * 60)
    try:
        input("  Presioná ENTER para continuar (o Ctrl+C para cancelar)...")
    except KeyboardInterrupt:
        print("\n  Cancelado por el usuario.")
        raise SystemExit(0)
    print("")

    main()
