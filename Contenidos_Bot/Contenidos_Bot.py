# -*- coding: utf-8 -*-
"""
Script para contar y comparar contenidos del Bot entre dos períodos
Lee archivos TSV exportados de Botmaker (rules-YYYY.MM.DD-HH.MM.tsv)
Genera un Excel con métricas de contenidos, intenciones y cambios

Uso:
    1. Colocar los 2 archivos TSV en esta carpeta (mes actual y mes anterior)
    2. Ejecutar: python Contenidos_Bot.py
    3. El script toma automáticamente los 2 TSV más recientes

Lee configuracion de fechas desde config_fechas.txt (para nombrar el output)
"""
import pandas as pd
import numpy as np
import re
import os
import glob
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from calendar import monthrange

# ==================== CONFIGURACION ====================
CONFIG = {
    'output_folder': 'output',
    'config_file': '../config_fechas.txt'
}

# Topics excluidos del conteo de contenidos relevantes
EXCLUDED_TOPICS = [
    'Core de bot', 'Revisar', '2. Onboardings', 'Derivaciones', 'Push', 'Notificaciones',
    'Z. Apagados, detonables', 'Z. Apagados, reciclables', 'Test', 'Utilidades para intenciones',
    'AWS Polly', 'Prueba', ' '
]

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    """Lee el archivo de configuracion de fechas. Retorna (modo, mes, anio, descripcion)"""
    rutas_posibles = [
        config_file,
        'config_fechas.txt',
        '../config_fechas.txt',
        '../../config_fechas.txt',
    ]

    archivo_encontrado = None
    for ruta in rutas_posibles:
        if os.path.exists(ruta):
            archivo_encontrado = ruta
            break

    if not archivo_encontrado:
        print("[ERROR] No se encuentra config_fechas.txt en ninguna ubicación")
        return None, None, None, None

    try:
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None

        with open(archivo_encontrado, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if line.startswith('MES='):
                    mes = int(line.split('=')[1].strip())
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio = int(line.split('=')[1].strip())
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()

        if fecha_inicio_str and fecha_fin_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            descripcion = "{} al {}".format(
                fecha_inicio.strftime('%d/%m/%Y'),
                fecha_fin.strftime('%d/%m/%Y')
            )
            return 'rango', None, None, descripcion

        if mes is not None and anio is not None:
            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)
            return 'mes', mes, anio, descripcion

        print("[ERROR] config_fechas.txt no contiene configuracion valida")
        return None, None, None, None

    except Exception as e:
        print("[ERROR] Error leyendo config_fechas.txt: {}".format(str(e)))
        return None, None, None, None


def get_month_name(mes):
    """Retorna el nombre del mes en español"""
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')


def get_month_abbr(mes):
    """Retorna la abreviatura del mes en español"""
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')


def buscar_tsv_mas_recientes():
    """
    Busca los 2 archivos TSV más recientes en la carpeta actual.
    El más reciente se considera "mes actual" y el anterior "mes anterior".
    Retorna: (path_current, path_previous) o (None, None) si no hay suficientes.
    """
    archivos_tsv = glob.glob('rules-*.tsv')

    if len(archivos_tsv) < 2:
        print("[ERROR] Se necesitan al menos 2 archivos TSV (mes actual y mes anterior)")
        print("    Archivos encontrados: {}".format(len(archivos_tsv)))
        if archivos_tsv:
            for f in archivos_tsv:
                print("      - {}".format(f))
        return None, None

    # Ordenar por nombre (que contiene la fecha YYYY.MM.DD-HH.MM)
    archivos_tsv.sort()

    path_previous = archivos_tsv[-2]
    path_current = archivos_tsv[-1]

    print("[OK] Archivos TSV encontrados:")
    print("    Mes anterior: {}".format(path_previous))
    print("    Mes actual:   {}".format(path_current))

    return path_current, path_previous


def cargar_tsv(filepath):
    """Carga un archivo TSV de Botmaker"""
    print("    Cargando {}...".format(filepath))

    # Probar diferentes codificaciones (utf-8-sig maneja BOM)
    encodings = ['utf-8-sig', 'utf-8', 'latin-1', 'iso-8859-1', 'cp1252']
    df = None

    for encoding in encodings:
        try:
            df = pd.read_csv(filepath, sep='\t', encoding=encoding, on_bad_lines='skip')
            print("    [OK] Codificación detectada: {}".format(encoding))
            break
        except Exception as e:
            continue

    if df is None:
        raise Exception("No se pudo leer el archivo con ninguna codificación")

    # LIMPIAR NOMBRES DE COLUMNAS
    # 1. Eliminar espacios al inicio y final
    # 2. Eliminar BOM (\ufeff) si existe
    # 3. Eliminar comillas si existen
    df.columns = df.columns.str.strip().str.replace('\ufeff', '').str.replace('"', '').str.replace("'", "")

    print("    [OK] {} registros cargados".format(len(df)))
    print("    Columnas detectadas: {}".format(list(df.columns)))

    # Verificar que existen las columnas necesarias
    columnas_requeridas = ['Active', 'Topic path', 'Topic', 'ID', 'Name', 'Bot Says']
    faltantes = [col for col in columnas_requeridas if col not in df.columns]

    if faltantes:
        print("    [ADVERTENCIA] Columnas faltantes: {}".format(faltantes))
        print("    [INFO] Columnas disponibles: {}".format(list(df.columns)))
    else:
        print("    [OK] Todas las columnas requeridas encontradas")

    return df


def normalizar_active(valor):
    """
    Normaliza el valor de Active a booleano.
    Maneja: 'true', 'True', 'TRUE', '1', 'yes', 'si', True, 1, etc.
    """
    if pd.isna(valor):
        return False
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return bool(valor)
    # Es string
    return str(valor).strip().lower() in ['true', '1', 'yes', 'si', 'verdadero', 'activo', 'activa']


def filtrar_contenidos_relevantes(df):
    """
    Filtra contenidos relevantes:
    - Excluye apagados (Active == False)
    - Excluye Topic paths que comienzan con topics excluidos
    - Excluye Topic path vacío
    - Excluye intenciones con Login o Push en el nombre
    """
    # Crear copia para no modificar el original
    df = df.copy()

    # Normalizar la columna Active a booleano
    if 'Active' in df.columns:
        df['Active'] = df['Active'].apply(normalizar_active)
        print("    [INFO] Active normalizado: {} activos, {} inactivos".format(
            df['Active'].sum(), 
            (~df['Active']).sum()
        ))
    else:
        print("    [ADVERTENCIA] No se encontró columna 'Active', asumiendo todos activos")
        df['Active'] = True

    # Filtrar: excluir inactivos
    filtered = df[df['Active'] == True].copy()

    # Filtrar por Topic path
    if 'Topic path' in filtered.columns:
        filtered = filtered[~filtered['Topic path'].str.startswith(tuple(EXCLUDED_TOPICS), na=False)]
        filtered = filtered[filtered['Topic path'].notna()]
        filtered = filtered[~filtered['Topic path'].str.contains('Login', na=False)]

    # Filtrar por Name (Push)
    if 'Name' in filtered.columns:
        filtered = filtered[~filtered['Name'].str.contains('push|PUSH|Push', na=False, regex=True)]

    return filtered


def clasificar_contenidos(df):
    """Clasifica intenciones en 'action' o 'wording' según Bot Says"""
    df = df.copy()

    if 'Bot Says' not in df.columns:
        print("    [ADVERTENCIA] No se encontró columna 'Bot Says', clasificando todo como 'action'")
        df['Classification'] = 'action'
        return df

    df['Classification'] = df['Bot Says'].apply(
        lambda x: 'action' if pd.isna(x) or len(str(x).strip()) <= 1 else 'wording'
    )
    return df


def calcular_metricas(df_current, df_previous):
    """Calcula todas las métricas de contenidos"""

    # Métricas originales (sin filtrar)
    original_topics_current = df_current['Topic'].nunique() if 'Topic' in df_current.columns else 0
    original_topics_previous = df_previous['Topic'].nunique() if 'Topic' in df_previous.columns else 0

    current_total_ids = df_current['ID'].nunique() if 'ID' in df_current.columns else 0
    previous_total_ids = df_previous['ID'].nunique() if 'ID' in df_previous.columns else 0

    # Normalizar Active para contar activos/inactivos
    if 'Active' in df_current.columns:
        df_current['Active'] = df_current['Active'].apply(normalizar_active)
        current_active_intents = df_current[df_current['Active'] == True]['ID'].nunique() if 'ID' in df_current.columns else 0
        current_inactive_intents = df_current[df_current['Active'] == False]['ID'].nunique() if 'ID' in df_current.columns else 0
        current_active_topics = df_current[df_current['Active'] == True]['Topic'].nunique() if 'Topic' in df_current.columns else 0
        current_inactive_topics = original_topics_current - current_active_topics
    else:
        current_active_intents = current_total_ids
        current_inactive_intents = 0
        current_active_topics = original_topics_current
        current_inactive_topics = 0

    if 'Active' in df_previous.columns:
        df_previous['Active'] = df_previous['Active'].apply(normalizar_active)
        previous_active_intents = df_previous[df_previous['Active'] == True]['ID'].nunique() if 'ID' in df_previous.columns else 0
        previous_inactive_intents = df_previous[df_previous['Active'] == False]['ID'].nunique() if 'ID' in df_previous.columns else 0
        previous_active_topics = df_previous[df_previous['Active'] == True]['Topic'].nunique() if 'Topic' in df_previous.columns else 0
        previous_inactive_topics = original_topics_previous - previous_active_topics
    else:
        previous_active_intents = previous_total_ids
        previous_inactive_intents = 0
        previous_active_topics = original_topics_previous
        previous_inactive_topics = 0

    # Filtrar contenidos relevantes
    print("\n  Filtrando contenidos relevantes (mes actual)...")
    filtered_current = filtrar_contenidos_relevantes(df_current)
    print("  Filtrando contenidos relevantes (mes anterior)...")
    filtered_previous = filtrar_contenidos_relevantes(df_previous)

    relevant_intentions_current = filtered_current['ID'].nunique() if 'ID' in filtered_current.columns else 0
    relevant_intentions_previous = filtered_previous['ID'].nunique() if 'ID' in filtered_previous.columns else 0
    relevant_topic_current = filtered_current['Topic'].nunique() if 'Topic' in filtered_current.columns else 0
    relevant_topic_previous = filtered_previous['Topic'].nunique() if 'Topic' in filtered_previous.columns else 0

    # Clasificar por wording/action
    filtered_current = clasificar_contenidos(filtered_current)
    filtered_previous = clasificar_contenidos(filtered_previous)

    wording_int_current = filtered_current[filtered_current['Classification'] == 'wording']['ID'].nunique() if 'ID' in filtered_current.columns else 0
    actions_int_current = filtered_current[filtered_current['Classification'] == 'action']['ID'].nunique() if 'ID' in filtered_current.columns else 0
    wording_topic_current = filtered_current[filtered_current['Classification'] == 'wording']['Topic'].nunique() if 'Topic' in filtered_current.columns else 0
    actions_topic_current = filtered_current[filtered_current['Classification'] == 'action']['Topic'].nunique() if 'Topic' in filtered_current.columns else 0

    wording_int_previous = filtered_previous[filtered_previous['Classification'] == 'wording']['ID'].nunique() if 'ID' in filtered_previous.columns else 0
    actions_int_previous = filtered_previous[filtered_previous['Classification'] == 'action']['ID'].nunique() if 'ID' in filtered_previous.columns else 0
    wording_topic_previous = filtered_previous[filtered_previous['Classification'] == 'wording']['Topic'].nunique() if 'Topic' in filtered_previous.columns else 0
    actions_topic_previous = filtered_previous[filtered_previous['Classification'] == 'action']['Topic'].nunique() if 'Topic' in filtered_previous.columns else 0

    # Comparar: nuevas, eliminadas, modificadas
    if 'ID' in filtered_current.columns and 'ID' in filtered_previous.columns:
        current_ids = set(filtered_current['ID'])
        previous_ids = set(filtered_previous['ID'])

        new_ids = current_ids - previous_ids
        removed_ids = previous_ids - current_ids
        common_ids = current_ids & previous_ids
    else:
        new_ids = set()
        removed_ids = set()
        common_ids = set()

    # Detectar modificaciones
    modified_ids = []
    if 'Bot Says' in filtered_current.columns and 'Bot Says' in filtered_previous.columns and 'ID' in filtered_current.columns:
        for id_ in common_ids:
            curr_rows = filtered_current[filtered_current['ID'] == id_]
            prev_rows = filtered_previous[filtered_previous['ID'] == id_]

            if not curr_rows.empty and not prev_rows.empty:
                curr_bot = " ".join(curr_rows['Bot Says'].dropna().astype(str).tolist())
                prev_bot = " ".join(prev_rows['Bot Says'].dropna().astype(str).tolist())

                if abs(len(curr_bot) - len(prev_bot)) > 5:
                    modified_ids.append(id_)

    if 'Topic' in filtered_current.columns and 'Topic' in filtered_previous.columns:
        current_topics_set = set(filtered_current['Topic'])
        previous_topics_set = set(filtered_previous['Topic'])

        new_topics = current_topics_set - previous_topics_set
        removed_topics = previous_topics_set - current_topics_set

        modified_topics = []
        if 'Bot Says' in filtered_current.columns and 'Bot Says' in filtered_previous.columns:
            for topic in current_topics_set & previous_topics_set:
                curr_rows = filtered_current[filtered_current['Topic'] == topic]
                prev_rows = filtered_previous[filtered_previous['Topic'] == topic]

                curr_bot = " ".join(str(v) for v in curr_rows['Bot Says'].dropna().tolist())
                prev_bot = " ".join(str(v) for v in prev_rows['Bot Says'].dropna().tolist())

                if abs(len(curr_bot) - len(prev_bot)) > 5:
                    modified_topics.append(topic)
    else:
        new_topics = set()
        removed_topics = set()
        modified_topics = []

    return {
        'current_total_ids': current_total_ids,
        'original_topics_current': original_topics_current,
        'current_active_intents': current_active_intents,
        'current_inactive_intents': current_inactive_intents,
        'current_active_topics': current_active_topics,
        'current_inactive_topics': current_inactive_topics,
        'previous_total_ids': previous_total_ids,
        'original_topics_previous': original_topics_previous,
        'previous_active_intents': previous_active_intents,
        'previous_inactive_intents': previous_inactive_intents,
        'previous_active_topics': previous_active_topics,
        'previous_inactive_topics': previous_inactive_topics,
        'relevant_intentions_current': relevant_intentions_current,
        'relevant_intentions_previous': relevant_intentions_previous,
        'relevant_topic_current': relevant_topic_current,
        'relevant_topic_previous': relevant_topic_previous,
        'wording_int_current': wording_int_current,
        'actions_int_current': actions_int_current,
        'wording_topic_current': wording_topic_current,
        'actions_topic_current': actions_topic_current,
        'wording_int_previous': wording_int_previous,
        'actions_int_previous': actions_int_previous,
        'wording_topic_previous': wording_topic_previous,
        'actions_topic_previous': actions_topic_previous,
        'new_ids': len(new_ids),
        'removed_ids': len(removed_ids),
        'modified_ids': len(modified_ids),
        'new_topics': len(new_topics),
        'removed_topics': len(removed_topics),
        'modified_topics': len(modified_topics),
    }


def crear_excel(filepath, metricas, modo, mes, anio, descripcion):
    """Crea el Excel de output con las métricas de contenidos"""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contenidos Bot'

    header_font = Font(bold=True)

    # Header de fecha
    if modo == 'mes':
        header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
    else:
        header_fecha = descripcion

    # Encabezados
    ws['A1'] = 'Dato'
    ws['B1'] = 'Mes Actual'
    ws['C1'] = 'Mes Anterior'
    ws['A1'].font = header_font
    ws['B1'].font = header_font
    ws['C1'].font = header_font

    # Datos
    filas = [
        ('Total Intenciones en el Bot', metricas['current_total_ids'], metricas['previous_total_ids']),
        ('Total Contenidos en el Bot', metricas['original_topics_current'], metricas['original_topics_previous']),
        ('Intenciones Prendidas', metricas['current_active_intents'], metricas['previous_active_intents']),
        ('Intenciones Apagadas', metricas['current_inactive_intents'], metricas['previous_inactive_intents']),
        ('Contenidos Prendidos', metricas['current_active_topics'], metricas['previous_active_topics']),
        ('Contenidos Apagados', metricas['current_inactive_topics'], metricas['previous_inactive_topics']),
        ('Total Intenciones Relevantes', metricas['relevant_intentions_current'], metricas['relevant_intentions_previous']),
        ('Total Contenidos Relevantes', metricas['relevant_topic_current'], metricas['relevant_topic_previous']),
        ('Intenciones Relevantes de Wording', metricas['wording_int_current'], metricas['wording_int_previous']),
        ('Intenciones Relevantes de Acciones', metricas['actions_int_current'], metricas['actions_int_previous']),
        ('Contenidos Relevantes de Wording', metricas['wording_topic_current'], metricas['wording_topic_previous']),
        ('Contenidos Relevantes de Acciones', metricas['actions_topic_current'], metricas['actions_topic_previous']),
    ]

    for idx, (dato, actual, anterior) in enumerate(filas, start=2):
        ws['A{}'.format(idx)] = dato
        ws['B{}'.format(idx)] = actual
        ws['C{}'.format(idx)] = anterior

    # Sección de cambios
    row = len(filas) + 3
    ws['A{}'.format(row)] = 'Cambios entre períodos'
    ws['A{}'.format(row)].font = Font(bold=True, size=12)
    row += 1

    ws['A{}'.format(row)] = 'Dato'
    ws['B{}'.format(row)] = 'Cantidad'
    ws['A{}'.format(row)].font = header_font
    ws['B{}'.format(row)].font = header_font
    row += 1

    cambios = [
        ('Intenciones Nuevas', metricas['new_ids']),
        ('Intenciones Eliminadas', metricas['removed_ids']),
        ('Intenciones Modificadas', metricas['modified_ids']),
        ('Topics Nuevos', metricas['new_topics']),
        ('Topics Eliminados', metricas['removed_topics']),
        ('Topics Modificados', metricas['modified_topics']),
    ]

    for dato, cantidad in cambios:
        ws['A{}'.format(row)] = dato
        ws['B{}'.format(row)] = cantidad
        row += 1

    # Ajustar anchos
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    wb.save(filepath)
    print("[OK] Excel creado: {}".format(filepath))


def create_or_update_dashboard_master(filepath, metricas, modo, mes, anio):
    """
    Crea el Excel Dashboard Master con el template estándar
    Llena las celdas D7 (Contenidos en Botmaker) y D8 (Contenidos Prendidos USUARIO)
    """

    print("    [INFO] Creando Dashboard Master...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    header_font = Font(bold=True)

    if modo == 'mes':
        header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
    else:
        header_fecha = 'rango'

    # FILA 1: Encabezados
    ws['B1'] = 'Indicador'
    ws['C1'] = 'Descripción/Detalle'
    ws['D1'] = header_fecha
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font

    # FILA 2-17: Indicadores
    ws['B2'] = 'Conversaciones'
    ws['C2'] = 'Q Conversaciones'

    ws['B3'] = 'Usuarios'
    ws['C3'] = 'Q Usuarios únicos'

    ws['B4'] = 'Sesiones abiertas por Pushes'
    ws['C4'] = 'Q Sesiones que se abrieron con una Push'

    ws['B5'] = 'Sesiones Alcanzadas por Pushes'
    ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'

    ws['B6'] = 'Mensajes Pushes Enviados'
    ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'

    ws['B7'] = 'Contenidos en Botmaker'
    ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
    ws['D7'] = metricas['current_active_topics']

    ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
    ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
    ws['D8'] = metricas['relevant_topic_current']

    ws['B9'] = 'Interacciones'
    ws['C9'] = 'Q Interacciones'

    ws['B10'] = 'Trámites, solicitudes y turnos'
    ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'

    ws['B11'] = 'contenidos mas consultados'
    ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'

    ws['B12'] = 'Derivaciones'
    ws['C12'] = 'Q Derivaciones'

    ws['B13'] = 'No entendimiento'
    ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'

    ws['B14'] = 'Tasa de Efectividad'
    ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'

    ws['B15'] = 'CES (Customer Effort Score)'
    ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'

    ws['B16'] = 'Satisfacción (CSAT)'
    ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'

    ws['B17'] = 'Uptime servidor'
    ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'

    # Ajustar anchos
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    wb.save(filepath)
    print("    [OK] Dashboard Master creado (D7 = {}, D8 = {})".format(
        metricas['current_active_topics'], metricas['relevant_topic_current']))


def imprimir_resultados(metricas):
    """Imprime los resultados en consola"""

    print("")
    print("=" * 60)
    print("RESULTADOS - CONTENIDOS DEL BOT")
    print("=" * 60)

    print("\n  MES ACTUAL:")
    print("    Total Intenciones:              {:,}".format(metricas['current_total_ids']))
    print("    Total Contenidos:               {:,}".format(metricas['original_topics_current']))
    print("    Intenciones Prendidas:          {:,}".format(metricas['current_active_intents']))
    print("    Intenciones Apagadas:           {:,}".format(metricas['current_inactive_intents']))
    print("    Contenidos Prendidos:           {:,}".format(metricas['current_active_topics']))
    print("    Contenidos Apagados:            {:,}".format(metricas['current_inactive_topics']))
    print("    Intenciones Relevantes:         {:,}".format(metricas['relevant_intentions_current']))
    print("    Contenidos Relevantes:          {:,}".format(metricas['relevant_topic_current']))

    print("\n  MES ANTERIOR:")
    print("    Total Intenciones:              {:,}".format(metricas['previous_total_ids']))
    print("    Total Contenidos:               {:,}".format(metricas['original_topics_previous']))
    print("    Intenciones Prendidas:          {:,}".format(metricas['previous_active_intents']))
    print("    Intenciones Apagadas:           {:,}".format(metricas['previous_inactive_intents']))
    print("    Contenidos Prendidos:           {:,}".format(metricas['previous_active_topics']))
    print("    Contenidos Apagados:            {:,}".format(metricas['previous_inactive_topics']))
    print("    Intenciones Relevantes:         {:,}".format(metricas['relevant_intentions_previous']))
    print("    Contenidos Relevantes:          {:,}".format(metricas['relevant_topic_previous']))

    print("\n  CAMBIOS ENTRE PERÍODOS:")
    print("    Intenciones Nuevas:             {:,}".format(metricas['new_ids']))
    print("    Intenciones Eliminadas:         {:,}".format(metricas['removed_ids']))
    print("    Intenciones Modificadas:        {:,}".format(metricas['modified_ids']))
    print("    Topics Nuevos:                  {:,}".format(metricas['new_topics']))
    print("    Topics Eliminados:              {:,}".format(metricas['removed_topics']))
    print("    Topics Modificados:             {:,}".format(metricas['modified_topics']))

    print("=" * 60)


# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("SCRIPT: CONTENIDOS DEL BOT")
    print("=" * 60)
    print("Compara contenidos entre dos exportaciones TSV de Botmaker")
    print("Genera: Excel con métricas + Dashboard Master (D7, D8)")
    print("")

    # Leer configuración de fechas
    print("Leyendo configuracion de fechas...")
    modo, mes, anio, descripcion = read_date_config(CONFIG['config_file'])

    if modo is None:
        print("[ERROR] No se pudo obtener la configuracion. Abortando.")
        exit(1)

    print("[OK] Período: {}".format(descripcion))

    # Buscar archivos TSV
    print("")
    print("Buscando archivos TSV...")
    path_current, path_previous = buscar_tsv_mas_recientes()

    if path_current is None:
        print("[ERROR] No se encontraron archivos TSV suficientes. Abortando.")
        exit(1)

    # Cargar datos
    print("")
    print("Cargando archivos TSV...")
    df_current = cargar_tsv(path_current)
    df_previous = cargar_tsv(path_previous)

    # Calcular métricas
    print("")
    print("Calculando métricas...")
    metricas = calcular_metricas(df_current, df_previous)

    # Imprimir resultados
    imprimir_resultados(metricas)

    # Generar archivos de salida
    output_folder = CONFIG['output_folder']
    os.makedirs(output_folder, exist_ok=True)

    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_excel = "contenidos_bot_detalle_{0}_{1}.xlsx".format(mes_nombre, anio)
        filename_dashboard = "contenidos_bot_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:
        fecha_inicio_fmt = descripcion.replace('/', '').replace(' al ', '_a_')
        filename_excel = "contenidos_bot_detalle_{0}.xlsx".format(fecha_inicio_fmt)
        filename_dashboard = "contenidos_bot_{0}.xlsx".format(fecha_inicio_fmt)

    filepath_excel = os.path.join(output_folder, filename_excel)
    filepath_dashboard = os.path.join(output_folder, filename_dashboard)

    print("")
    print("Generando archivos Excel...")
    crear_excel(filepath_excel, metricas, modo, mes, anio, descripcion)
    create_or_update_dashboard_master(filepath_dashboard, metricas, modo, mes, anio)

    print("")
    print("ARCHIVOS GENERADOS:")
    print("    [1] {}".format(filepath_excel))
    print("        - Métricas completas de contenidos")
    print("        - Comparación mes actual vs anterior")
    print("    [2] {}".format(filepath_dashboard))
    print("        - D7 = {} (Contenidos Prendidos)".format(metricas['current_active_topics']))
    print("        - D8 = {} (Contenidos Relevantes)".format(metricas['relevant_topic_current']))

    print("")
    print("=" * 60)
    print("PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 60)