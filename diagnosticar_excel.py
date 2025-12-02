#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script de Diagnóstico - Verificar Excel generados

Este script ayuda a diagnosticar problemas con el consolidador
mostrando exactamente qué hay en cada Excel generado.

Uso:
    python diagnosticar_excel.py
"""
import os
import glob
import openpyxl

# Módulos a verificar
MODULOS = {
    'Usuarios y Conversaciones': {
        'carpeta': 'Metricas_Boti_Conversaciones_Usuarios/output',
        'patron': 'usuarios_conversaciones_*.xlsx',
        'patron_alternativo': 'usuarios_*.xlsx',
        'celdas_esperadas': ['D2', 'D3']
    },
    'Sesiones Abiertas': {
        'carpeta': 'Sesiones_Abiertas_Pushes/output',
        'patron': 'sesiones_abiertas_pushes_*.xlsx',
        'patron_alternativo': 'sesiones_abiertas*.xlsx',
        'celdas_esperadas': ['D4']
    },
    'Sesiones Alcanzadas': {
        'carpeta': 'Sesiones_alcanzadas_pushes/output',
        'patron': 'sesiones_alcanzadas_pushes_*.xlsx',
        'patron_alternativo': 'sesiones_alcanzadas*.xlsx',
        'celdas_esperadas': ['D5']
    },
    'Pushes Enviadas': {
        'carpeta': 'Pushes_Enviadas/output',
        'patron': 'pushes_enviadas_*.xlsx',
        'patron_alternativo': '*pushes*.xlsx',
        'celdas_esperadas': ['D6']
    },
    'Disponibilidad WhatsApp': {
        'carpeta': 'Metricas_Boti_Disponibilidad/output',
        'patron': 'whatsapp_availability_*.xlsx',
        'patron_alternativo': '*availability*.xlsx',
        'celdas_esperadas': ['D17']
    }
}

def print_header(text):
    """Imprime un header formateado"""
    print("\n" + "=" * 70)
    print(f"  {text}")
    print("=" * 70 + "\n")

def buscar_excel_mas_reciente(carpeta, patron, patron_alternativo=None):
    """Busca el Excel más reciente"""
    ruta_completa = os.path.join(carpeta, patron)
    archivos = glob.glob(ruta_completa)
    
    # Si no encuentra con el patrón principal, probar con el alternativo
    if not archivos and patron_alternativo:
        ruta_completa = os.path.join(carpeta, patron_alternativo)
        archivos = glob.glob(ruta_completa)
    
    if not archivos:
        return None
    
    archivos.sort(key=os.path.getmtime, reverse=True)
    return archivos[0]

def leer_todas_celdas_relevantes(archivo_excel):
    """Lee todas las celdas relevantes de un Excel"""
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws = wb.active
        
        # Leer header
        header_d1 = ws['D1'].value
        
        # Leer celdas de datos (D2 a D17)
        celdas = {}
        for i in range(2, 18):
            celda = f'D{i}'
            valor = ws[celda].value
            indicador_b = ws[f'B{i}'].value
            detalle_c = ws[f'C{i}'].value
            
            celdas[celda] = {
                'valor': valor,
                'indicador': indicador_b,
                'detalle': detalle_c
            }
        
        wb.close()
        
        return header_d1, celdas
    except Exception as e:
        return None, {'error': str(e)}

def diagnosticar_modulo(nombre, config):
    """Diagnostica un módulo específico"""
    print(f"📊 {nombre}")
    print(f"   Carpeta: {config['carpeta']}")
    print(f"   Patrón: {config['patron']}")
    print(f"   Celdas esperadas: {', '.join(config['celdas_esperadas'])}")
    
    # Verificar si la carpeta existe
    if not os.path.exists(config['carpeta']):
        print(f"   ❌ La carpeta no existe!")
        return False
    
    print(f"   ✅ Carpeta existe")
    
    # Buscar Excel
    patron_alt = config.get('patron_alternativo')
    excel_path = buscar_excel_mas_reciente(config['carpeta'], config['patron'], patron_alt)
    
    if not excel_path:
        print(f"   ❌ No se encontró ningún archivo Excel")
        
        # Listar qué archivos hay en la carpeta
        todos_archivos = os.listdir(config['carpeta'])
        if todos_archivos:
            print(f"   📁 Archivos en la carpeta:")
            for archivo in todos_archivos:
                print(f"      • {archivo}")
        else:
            print(f"   📁 La carpeta está vacía")
        
        return False
    
    print(f"   ✅ Excel encontrado: {os.path.basename(excel_path)}")
    
    # Leer contenido
    print(f"\n   📖 Contenido del Excel:")
    header, celdas = leer_todas_celdas_relevantes(excel_path)
    
    if 'error' in celdas:
        print(f"   ❌ Error al leer Excel: {celdas['error']}")
        return False
    
    print(f"   📅 Header D1: {header}")
    
    # Mostrar solo las celdas esperadas
    print(f"\n   📊 Celdas esperadas:")
    for celda in config['celdas_esperadas']:
        info = celdas.get(celda, {})
        valor = info.get('valor', 'N/A')
        indicador = info.get('indicador', 'N/A')
        
        if valor is None or valor == '' or valor == 'N/A':
            estado = "❌ VACÍO"
        else:
            estado = "✅ OK"
        
        print(f"   {estado} {celda}: {valor}")
        print(f"       Indicador B{celda[1:]}: {indicador}")
    
    # Mostrar también el rango completo D2-D17 para contexto
    print(f"\n   📋 Todas las celdas D2-D17 (para contexto):")
    for i in range(2, 18):
        celda = f'D{i}'
        info = celdas.get(celda, {})
        valor = info.get('valor', 'N/A')
        indicador = info.get('indicador', 'N/A')
        
        if celda in config['celdas_esperadas']:
            marca = "⭐"
        else:
            marca = "  "
        
        print(f"   {marca} {celda}: {valor} | B{i}: {indicador}")
    
    print("\n" + "-" * 70)
    return True

def main():
    """Función principal"""
    print_header("DIAGNÓSTICO DE EXCEL - Metricas_Boti_Mensual")
    
    print("Este script verifica el contenido de todos los Excel generados")
    print("para ayudar a diagnosticar problemas con el consolidador.\n")
    
    print("🔍 Verificando módulos...")
    
    resultados = {}
    
    for nombre, config in MODULOS.items():
        resultado = diagnosticar_modulo(nombre, config)
        resultados[nombre] = resultado
        print("")
    
    # Resumen
    print_header("RESUMEN")
    
    exitosos = sum(1 for v in resultados.values() if v)
    total = len(resultados)
    
    print(f"📊 Módulos verificados: {total}")
    print(f"✅ Con Excel encontrado: {exitosos}")
    print(f"❌ Sin Excel: {total - exitosos}")
    
    if exitosos < total:
        print("\n⚠️  PROBLEMA DETECTADO:")
        print("   Algunos módulos no tienen Excel generados.")
        print("   Ejecutar primero: python run_all.py")
    
    print("\n💡 Notas:")
    print("   • Las celdas marcadas con ⭐ son las que el consolidador debe extraer")
    print("   • Si una celda esperada está VACÍA, el consolidador pondrá '-'")
    print("   • Verificar que los scripts individuales generen correctamente los valores")
    
    print("\n" + "=" * 70 + "\n")

if __name__ == "__main__":
    main()
