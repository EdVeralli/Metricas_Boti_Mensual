#!/usr/bin/env python3
# -*- coding: utf-8 -*-
'''
Consolidador de Excel - Metricas_Boti_Mensual

Consolida todos los reportes Excel parciales en un dashboard completo
ubicado en la raíz del proyecto.

Este script:
1. Busca los Excel más recientes en cada carpeta output/
2. Lee las métricas específicas de cada uno
3. Crea un dashboard consolidado con todas las métricas
4. Guarda el archivo en la raíz del proyecto

Uso:
    python consolidar_excel.py
'''
import os
import glob
from datetime import datetime
from calendar import monthrange
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== FUNCIONES DE UTILIDAD ====================

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, f'mes{mes}')

def leer_config_fechas(config_file='config_fechas.txt'):
    '''
    Lee el archivo de configuración de fechas y retorna el periodo como string
    para usar en el nombre del archivo.
    
    Retorna: string con el periodo (ej: "octubre_2025" o "2025-10-01_al_2025-10-15")
    '''
    try:
        if not os.path.exists(config_file):
            # Si no existe, usar fecha actual
            return datetime.now().strftime('%Y%m%d')
        
        mes = None
        anio = None
        fecha_inicio = None
        fecha_fin = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
                
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio = line.split('=')[1].strip()
                
                if line.startswith('FECHA_FIN='):
                    fecha_fin = line.split('=')[1].strip()
        
        # PRIORIDAD: Rango personalizado
        if fecha_inicio and fecha_fin:
            # Formato: 2025-10-01_al_2025-10-15
            return f"{fecha_inicio}_al_{fecha_fin}"
        
        # Mes completo
        if mes is not None and anio is not None:
            mes_nombre = get_month_name(mes)
            return f"{mes_nombre}_{anio}"
        
        # Fallback: fecha actual
        return datetime.now().strftime('%Y%m%d')
        
    except Exception as e:
        print(f"⚠️  Error al leer config: {str(e)}")
        return datetime.now().strftime('%Y%m%d')

# ==================== CONFIGURACIÓN ====================
MODULOS = {
    'usuarios_conversaciones': {
        'carpeta': 'Metricas_Boti_Conversaciones_Usuarios/output',
        'patron': 'usuarios_conversaciones_*.xlsx',
        'patron_alternativo': 'usuarios_*.xlsx',
        'celdas': {
            'D2': 'Conversaciones',
            'D3': 'Usuarios'
        }
    },
    'sesiones_abiertas': {
        'carpeta': 'Sesiones_Abiertas_Pushes/output',
        'patron': 'sesiones_abiertas_pushes_*.xlsx',
        'patron_alternativo': 'sesiones_abiertas*.xlsx',
        'celdas': {
            'D4': 'Sesiones Abiertas'
        }
    },
    'sesiones_alcanzadas': {
        'carpeta': 'Sesiones_alcanzadas_pushes/output',
        'patron': 'sesiones_alcanzadas_pushes_*.xlsx',
        'patron_alternativo': 'sesiones_alcanzadas*.xlsx',
        'celdas': {
            'D5': 'Sesiones Alcanzadas'
        }
    },
    'pushes_enviadas': {
        'carpeta': 'Pushes_Enviadas/output',
        'patron': 'pushes_enviadas_*.xlsx',
        'patron_alternativo': '*pushes*.xlsx',
        'celdas': {
            'D6': 'Pushes Enviadas'
        }
    },
    'contenidos_bot': {
        'carpeta': 'Contenidos_Bot/output',
        'patron': 'contenidos_bot_*.xlsx',
        'patron_alternativo': None,
        'celdas': {
            'D7': 'Contenidos Activos',
            'D8': 'Contenidos Relevantes'
        },
        'excluir_patron': '*_detalle_*'
    },
    'no_entendimiento': {
        'carpeta': 'No_Entendidos/output',
        'patron': 'no_entendimiento_*.xlsx',
        'patron_alternativo': None,
        'celdas': {
            'D13': 'No Entendimiento'
        },
        'excluir_patron': '*_detalle_*'
    },
    'feedback_efectividad': {
        'carpeta': 'Feedback_Efectividad/output',
        'patron': 'feedback_efectividad_*.xlsx',
        'patron_alternativo': None,
        'celdas': {
            'D14': 'Efectividad'
        },
        'excluir_patron': '*_detalle_*'  # Excluir archivos _detalle
    },
    'feedback_ces': {
        'carpeta': 'Feedback_CES/output',
        'patron': 'feedback_ces_*.xlsx',
        'patron_alternativo': None,
        'celdas': {
            'D15': 'CES'
        },
        'excluir_patron': '*_detalle_*'  # Excluir archivos _detalle
    },
    'feedback_csat': {
        'carpeta': 'Feedback_CSAT/output',
        'patron': 'feedback_csat_*.xlsx',
        'patron_alternativo': None,
        'celdas': {
            'D16': 'CSAT'
        },
        'excluir_patron': '*_detalle_*'  # Excluir archivos _detalle
    },
    'contenidos_consultados': {
        'carpeta': 'Contenidos_Consultados/output',
        'patron': 'contenidos_consultados_*.xlsx',
        'patron_alternativo': None,
        'celdas': {
            'D11': 'Contenidos Consultados'
        },
        'excluir_patron': '*_detalle_*'
    },
    'disponibilidad': {
        'carpeta': 'Metricas_Boti_Disponibilidad/output',
        'patron': 'whatsapp_availability_*.xlsx',
        'patron_alternativo': '*availability*.xlsx',
        'celdas': {
            'D17': 'Availability'
        }
    }
}

# Estructura del dashboard GCBA (17 filas)
ESTRUCTURA_DASHBOARD = [
    {'fila': 1, 'indicador': 'Indicador', 'detalle': 'Descripción/Detalle', 'es_header': True},
    {'fila': 2, 'indicador': 'Conversaciones', 'detalle': 'Q Conversaciones'},
    {'fila': 3, 'indicador': 'Usuarios', 'detalle': 'Q Usuarios únicos'},
    {'fila': 4, 'indicador': 'Sesiones abiertas por Pushes', 'detalle': 'Q Sesiones que se abrieron con una Push'},
    {'fila': 5, 'indicador': 'Sesiones Alcanzadas por Pushes', 'detalle': 'Q Sesiones que recibieron al menos 1 Push'},
    {'fila': 6, 'indicador': 'Mensajes Pushes Enviados', 'detalle': 'Q de mensajes enviados bajo el formato push'},
    {'fila': 7, 'indicador': 'Contenidos en Botmaker', 'detalle': 'Contenidos prendidos en botmaker'},
    {'fila': 8, 'indicador': 'Contenidos Prendidos para el USUARIO', 'detalle': 'Contenidos prendidos de cara al usuario'},
    {'fila': 9, 'indicador': 'Interacciones', 'detalle': 'Q Interacciones'},
    {'fila': 10, 'indicador': 'Trámites, solicitudes y turnos', 'detalle': 'Q Trámites, solicitudes y turnos disponibles'},
    {'fila': 11, 'indicador': 'contenidos mas consultados', 'detalle': 'Q Contenidos con más interacciones'},
    {'fila': 12, 'indicador': 'Derivaciones', 'detalle': 'Q Derivaciones'},
    {'fila': 13, 'indicador': 'No entendimiento', 'detalle': 'Performance motor de búsqueda'},
    {'fila': 14, 'indicador': 'Tasa de Efectividad', 'detalle': 'Mide el porcentaje de usuarios que lograron su objetivo'},
    {'fila': 15, 'indicador': 'CES (Customer Effort Score)', 'detalle': 'Mide la facilidad con la que los usuarios pueden interactuar'},
    {'fila': 16, 'indicador': 'Satisfacción (CSAT)', 'detalle': 'Mide la satisfacción usando una escala de 1 a 5'},
    {'fila': 17, 'indicador': 'Uptime servidor', 'detalle': 'Disponibilidad del servidor (% tiempo activo)'},
]

# ==================== FUNCIONES ====================

def print_header(text):
    '''Imprime un header formateado'''
    print("\n" + "=" * 70)
    print(f"  {text}")
    print("=" * 70 + "\n")

def buscar_excel_mas_reciente(carpeta, patron, patron_alternativo=None, excluir_patron=None):
    '''Busca el Excel más reciente en una carpeta, excluyendo archivos no deseados'''
    ruta_completa = os.path.join(carpeta, patron)
    print(f"   🔍 Buscando: {ruta_completa}")
    
    archivos = glob.glob(ruta_completa)
    
    # Si no encuentra con el patrón principal, probar con el alternativo
    if not archivos and patron_alternativo:
        ruta_alternativa = os.path.join(carpeta, patron_alternativo)
        print(f"   🔍 Probando patrón alternativo: {ruta_alternativa}")
        archivos = glob.glob(ruta_alternativa)
    
    # Excluir archivos según patrón de exclusión (ej: *_detalle_*)
    if excluir_patron and archivos:
        ruta_excluir = os.path.join(carpeta, excluir_patron)
        archivos_excluir = set(glob.glob(ruta_excluir))
        archivos = [f for f in archivos if f not in archivos_excluir]
        if archivos_excluir:
            print(f"   ⚠️  Excluidos {len(archivos_excluir)} archivos con patrón: {excluir_patron}")
    
    if not archivos:
        print(f"   ❌ No se encontraron archivos")
        return None
    
    print(f"   📁 Archivos encontrados: {len(archivos)}")
    for archivo in archivos:
        print(f"      • {os.path.basename(archivo)}")
    
    # Ordenar por fecha de modificación (más reciente primero)
    archivos.sort(key=os.path.getmtime, reverse=True)
    archivo_seleccionado = archivos[0]
    print(f"   ⭐ Seleccionado (más reciente): {os.path.basename(archivo_seleccionado)}")
    
    return archivo_seleccionado

def leer_valor_celda(archivo_excel, celda):
    '''Lee el valor de una celda específica de un Excel'''
    try:
        wb = openpyxl.load_workbook(archivo_excel, data_only=True)
        ws = wb.active
        
        valor = ws[celda].value
        
        # Debug: mostrar tipo de dato
        if valor is not None:
            tipo = type(valor).__name__
            print(f"       DEBUG: Celda {celda} = {valor} (tipo: {tipo})")
        else:
            print(f"       DEBUG: Celda {celda} = None")
        
        wb.close()
        
        return valor
    except Exception as e:
        print(f"    ⚠️  Error al leer {celda}: {str(e)}")
        return None

def extraer_metricas():
    '''Extrae todas las métricas de los Excel parciales'''
    print_header("EXTRAYENDO MÉTRICAS DE LOS REPORTES")
    
    metricas = {}
    periodo_detectado = None
    
    for modulo_key, config in MODULOS.items():
        print(f"\n📊 Módulo: {modulo_key}")
        print(f"   Carpeta: {config['carpeta']}")
        
        # Buscar Excel más reciente
        patron_alt = config.get('patron_alternativo')
        excluir_patron = config.get('excluir_patron')
        excel_path = buscar_excel_mas_reciente(
            config['carpeta'], 
            config['patron'], 
            patron_alt,
            excluir_patron
        )
        
        if not excel_path:
            print(f"   ⚠️  No se encontró archivo Excel - marcando como '-'")
            for celda in config['celdas'].keys():
                metricas[celda] = '-'
            continue
        
        print(f"   ✅ Excel encontrado")
        
        # Leer valores de las celdas
        for celda, nombre in config['celdas'].items():
            print(f"   📈 Leyendo {celda} ({nombre})...")
            valor = leer_valor_celda(excel_path, celda)
            metricas[celda] = valor if valor is not None else '-'
            print(f"   ✅ {celda} = {metricas[celda]}")
        
        # Intentar detectar el período del header
        if not periodo_detectado:
            try:
                wb = openpyxl.load_workbook(excel_path, data_only=True)
                ws = wb.active
                periodo_detectado = ws['D1'].value
                wb.close()
                print(f"   📅 Período detectado en header: {periodo_detectado}")
            except:
                pass
    
    print("\n" + "─" * 70)
    print(f"📅 Período detectado: {periodo_detectado if periodo_detectado else 'No disponible'}")
    
    return metricas, periodo_detectado

def crear_estilos():
    '''Crea los estilos para el Excel'''
    estilos = {
        'header': {
            'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
            'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'indicador': {
            'font': Font(name='Calibri', size=11, bold=True),
            'alignment': Alignment(horizontal='left', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'detalle': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='left', vertical='center', wrap_text=True),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        },
        'valor': {
            'font': Font(name='Calibri', size=11),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
    }
    return estilos

def crear_dashboard_consolidado(metricas, periodo):
    '''Crea el Excel consolidado con todas las métricas'''
    print_header("CREANDO DASHBOARD CONSOLIDADO")
    
    # Crear workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dashboard Boti"
    
    # Obtener estilos
    estilos = crear_estilos()
    
    # Configurar anchos de columna
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 20
    
    # Header del período
    periodo_texto = periodo if periodo else datetime.now().strftime('%b-%Y')
    
    # Crear estructura del dashboard
    for item in ESTRUCTURA_DASHBOARD:
        fila = item['fila']
        
        if item.get('es_header'):
            # Fila de encabezado
            ws[f'B{fila}'] = item['indicador']
            ws[f'C{fila}'] = item['detalle']
            ws[f'D{fila}'] = periodo_texto
            
            ws[f'B{fila}'].font = estilos['header']['font']
            ws[f'B{fila}'].fill = estilos['header']['fill']
            ws[f'B{fila}'].alignment = estilos['header']['alignment']
            ws[f'B{fila}'].border = estilos['header']['border']
            
            ws[f'C{fila}'].font = estilos['header']['font']
            ws[f'C{fila}'].fill = estilos['header']['fill']
            ws[f'C{fila}'].alignment = estilos['header']['alignment']
            ws[f'C{fila}'].border = estilos['header']['border']
            
            ws[f'D{fila}'].font = estilos['header']['font']
            ws[f'D{fila}'].fill = estilos['header']['fill']
            ws[f'D{fila}'].alignment = estilos['header']['alignment']
            ws[f'D{fila}'].border = estilos['header']['border']
        else:
            # Filas de datos
            ws[f'B{fila}'] = item['indicador']
            ws[f'C{fila}'] = item['detalle']
            
            # Obtener valor de las métricas extraídas
            celda_valor = f'D{fila}'
            valor = metricas.get(celda_valor, '-')
            ws[celda_valor] = valor
            
            # Aplicar estilos
            ws[f'B{fila}'].font = estilos['indicador']['font']
            ws[f'B{fila}'].alignment = estilos['indicador']['alignment']
            ws[f'B{fila}'].border = estilos['indicador']['border']
            
            ws[f'C{fila}'].font = estilos['detalle']['font']
            ws[f'C{fila}'].alignment = estilos['detalle']['alignment']
            ws[f'C{fila}'].border = estilos['detalle']['border']
            
            ws[celda_valor].font = estilos['valor']['font']
            ws[celda_valor].alignment = estilos['valor']['alignment']
            ws[celda_valor].border = estilos['valor']['border']
            
            # Aplicar formato específico según la celda
            if celda_valor == 'D11':  # Contenidos Consultados (texto multilínea Top 10)
                ws[celda_valor].alignment = Alignment(wrap_text=True, vertical='top')
            elif celda_valor == 'D13':  # No Entendimiento (porcentaje)
                ws[celda_valor].number_format = '0.00%'
            elif celda_valor == 'D14':  # Efectividad (porcentaje)
                ws[celda_valor].number_format = '0.00%'
            elif celda_valor == 'D15':  # CES (número decimal)
                ws[celda_valor].number_format = '0.00'
            elif celda_valor == 'D16':  # CSAT (porcentaje)
                ws[celda_valor].number_format = '0.00%'
            elif celda_valor == 'D17':  # Availability (porcentaje)
                ws[celda_valor].number_format = '0.00%'
    
    # Ajustar altura de filas
    for fila in range(1, 18):
        if fila == 1:
            ws.row_dimensions[fila].height = 30
        elif fila == 11:
            ws.row_dimensions[fila].height = 180  # Top 10 multilínea
        else:
            ws.row_dimensions[fila].height = 25
    
    # Generar nombre de archivo basado en el periodo del config
    periodo_config = leer_config_fechas()
    nombre_archivo = f'Boti_Consolidado_{periodo_config}.xlsx'
    
    # Guardar en raíz del proyecto
    wb.save(nombre_archivo)
    print(f"✅ Dashboard consolidado creado: {nombre_archivo}")
    
    return nombre_archivo

def mostrar_resumen(metricas, nombre_archivo):
    '''Muestra un resumen de las métricas consolidadas'''
    print_header("RESUMEN DE MÉTRICAS CONSOLIDADAS")
    
    print("📊 Métricas extraídas:\n")
    
    metricas_con_valor = 0
    metricas_sin_valor = 0
    
    # Lista de todas las celdas de métricas
    celdas_metricas = ['D2', 'D3', 'D4', 'D5', 'D6', 'D7', 'D8', 'D11', 'D13', 'D14', 'D15', 'D16', 'D17']
    
    for celda in celdas_metricas:
        valor = metricas.get(celda, '-')
        
        # Determinar el nombre de la métrica
        nombres = {
            'D2': 'Conversaciones',
            'D3': 'Usuarios',
            'D4': 'Sesiones Abiertas',
            'D5': 'Sesiones Alcanzadas',
            'D6': 'Pushes Enviadas',
            'D7': 'Contenidos Activos',
            'D8': 'Contenidos Relevantes',
            'D11': 'Contenidos Consultados',
            'D13': 'No Entendimiento',
            'D14': 'Efectividad',
            'D15': 'CES',
            'D16': 'CSAT',
            'D17': 'Availability'
        }
        
        nombre = nombres.get(celda, celda)
        estado = "✅" if valor != '-' else "⚠️"
        print(f"  {estado} {nombre} ({celda}): {valor}")
        
        if valor != '-':
            metricas_con_valor += 1
        else:
            metricas_sin_valor += 1
    
    print("\n" + "─" * 70)
    print(f"📈 Total de métricas: {metricas_con_valor + metricas_sin_valor}")
    print(f"✅ Con valor: {metricas_con_valor}")
    print(f"⚠️  Sin valor: {metricas_sin_valor}")
    
    print("\n" + "─" * 70)
    print(f"📁 Archivo guardado en la raíz del proyecto:")
    print(f"   {nombre_archivo}")
    
    if metricas_sin_valor > 0:
        print("\n⚠️  NOTA: Algunas métricas no tienen valor.")
        print("   Verificar que todos los scripts se ejecutaron correctamente.")

def main():
    '''Función principal'''
    print_header("CONSOLIDADOR DE EXCEL - Metricas_Boti_Mensual")
    
    print("Este script consolidará todos los reportes Excel en un dashboard único.")
    print("\n📋 Módulos a consolidar:")
    print("  • Usuarios y Conversaciones (D2, D3)")
    print("  • Sesiones Abiertas (D4)")
    print("  • Sesiones Alcanzadas (D5)")
    print("  • Pushes Enviadas (D6)")
    print("  • Contenidos del Bot (D7, D8)")
    print("  • Contenidos Consultados (D11)")
    print("  • No Entendimiento (D13)")
    print("  • Feedback - Efectividad (D14)")
    print("  • Feedback - CES (D15)")
    print("  • Feedback - CSAT (D16)")
    print("  • Disponibilidad WhatsApp (D17)")
    
    print("\n⚠️  El archivo consolidado se guardará en la raíz del proyecto")
    
    # Iniciar consolidación automáticamente
    print("\n" + "=" * 70)
    print("🚀 Iniciando consolidación automática...")
    print("=" * 70)
    
    # Leer periodo del config para mostrar
    periodo_config = leer_config_fechas()
    print(f"\n📅 Periodo detectado del config: {periodo_config.replace('_', ' ')}")
    print(f"📝 Nombre del archivo: Boti_Consolidado_{periodo_config}.xlsx")
    
    # Extraer métricas
    metricas, periodo = extraer_metricas()
    
    # Crear dashboard consolidado
    nombre_archivo = crear_dashboard_consolidado(metricas, periodo)
    
    # Mostrar resumen
    mostrar_resumen(metricas, nombre_archivo)
    
    print("\n" + "=" * 70)
    print("✨ CONSOLIDACIÓN COMPLETADA")
    print("=" * 70 + "\n")

if __name__ == "__main__":
    main()
