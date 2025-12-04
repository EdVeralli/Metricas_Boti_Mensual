# -*- coding: utf-8 -*-
'''
Script para ejecutar query de Feedback - CSAT (Customer Satisfaction) en Athena
Calcula la satisfacción del cliente basado en respuestas de conformidad
Guarda resultado en CSV y genera Excel con análisis detallado y cálculo de CSAT
Lee configuracion de fechas desde archivo config_fechas.txt

MODOS SOPORTADOS:
1. MES COMPLETO: Especificar MES y AÑO (comportamiento original)
2. RANGO PERSONALIZADO: Especificar FECHA_INICIO y FECHA_FIN

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti
'''
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': '../config_fechas.txt'  # Config centralizado en raiz del proyecto
}

# Reglas específicas para el cálculo de CSAT (Customer Satisfaction)
REGLAS_CSAT = {
    'conforme_integraciones': 'CXF01CUX01 Conforme Integraciones',
    'inconforme_integraciones': 'CXF01CUX01 Inconforme Integraciones',
    'muy_conforme_integraciones': 'CXF01CUX01 Muy conforme Integraciones',
    'muy_inconforme_integraciones': 'CXF01CUX01 Muy inconforme Integraciones',
    'no_se_integraciones': 'CXF01CUX01 No sé Integraciones ',  # Nota: hay espacio al final
    'conforme_estaticos': 'CXF01CUX02 Conforme Estáticos',
    'inconforme_estaticos': 'CXF01CUX02 Inconforme Estáticos',
    'muy_conforme_estaticos': 'CXF01CUX02 Muy conforme Estáticos',
    'muy_inconforme_estaticos': 'CXF01CUX02 Muy inconforme Estáticos',
    'no_se_estaticos': 'CXF01CUX02 No sé Estáticos',
    'conforme_pushes': 'CXF01CUX03 Conforme Pushes',
    'inconforme_pushes': 'CXF01CUX03 Inconforme Pushes',
    'muy_conforme_pushes': 'CXF01CUX03 Muy conforme Pushes',
    'muy_inconforme_pushes': 'CXF01CUX03 Muy inconforme Pushes',
    'no_se_pushes': 'CXF01CUX03 No sé Pushes'
}

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    '''
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)
    
    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    '''
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            print("    Creando archivo de ejemplo...")
            
            with open(config_file, 'w', encoding='utf-8') as f:
                f.write("# Configuracion de fechas para el reporte\n")
                f.write("# MODO 1: Mes completo\n")
                f.write("MES=10\n")
                f.write("AÑO=2025\n\n")
                f.write("# MODO 2: Rango personalizado\n")
                f.write("# FECHA_INICIO=2025-10-01\n")
                f.write("# FECHA_FIN=2025-10-15\n")
            
            print("    Archivo creado: {}".format(config_file))
            return 'mes', None, None, 10, 2025, "octubre 2025"
        
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
                
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()
        
        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                
                if fecha_inicio > fecha_fin:
                    print("[ERROR] FECHA_INICIO no puede ser posterior a FECHA_FIN")
                    return None, None, None, None, None, None
                
                descripcion = "{} al {}".format(
                    fecha_inicio.strftime('%d/%m/%Y'),
                    fecha_fin.strftime('%d/%m/%Y')
                )
                
                print("[INFO] Modo: RANGO PERSONALIZADO")
                return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion
                
            except ValueError as e:
                print("[ERROR] Formato de fecha invalido. Use YYYY-MM-DD (ej: 2025-10-01)")
                return None, None, None, None, None, None
        
        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            if mes < 1 or mes > 12:
                print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
                return None, None, None, None, None, None
            
            if anio < 2020 or anio > 2030:
                print("[ADVERTENCIA] Año inusual: {}".format(anio))
            
            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
            
            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)
            
            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion
        
        print("[ERROR] El archivo {} no contiene configuracion valida".format(config_file))
        return None, None, None, None, None, None
        
    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    '''Retorna la abreviatura del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_query(fecha_inicio, fecha_fin):
    '''Construye la query de Feedback - CSAT con el rango de fechas especificado'''
    
    query = """SELECT 
rule_name,
count(distinct(session_id)) as Cant_Sesiones 
FROM "caba-piba-consume-zone-db"."boti_message_metrics_2"
WHERE CAST(session_creation_time AS DATE) BETWEEN date '{fecha_inicio}' and date '{fecha_fin}'
and rule_name like ('%CXF%')
group by rule_name""".format(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    
    return query

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    '''Genera los nombres de archivos basados en el modo y las fechas'''
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "feedback_csat_{0}_{1}.csv".format(mes_nombre, anio)
        filename_excel_detalle = "feedback_csat_detalle_{0}_{1}.xlsx".format(mes_nombre, anio)
        filename_dashboard = "feedback_csat_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "feedback_csat_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_excel_detalle = "feedback_csat_detalle_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_dashboard = "feedback_csat_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
    
    return filename_csv, filename_excel_detalle, filename_dashboard

def extraer_valores_csat(df):
    '''
    Extrae los valores de las 15 reglas específicas para el cálculo de CSAT
    Retorna un diccionario con los valores encontrados
    '''
    valores = {}
    
    for key, rule_name in REGLAS_CSAT.items():
        # Buscar la regla en el DataFrame (con trim para manejar espacios)
        fila = df[df['rule_name'].str.strip() == rule_name.strip()]
        
        if len(fila) > 0:
            valores[key] = int(fila['cant_sesiones'].iloc[0])
        else:
            # Si no se encuentra, intentar sin el espacio final
            fila = df[df['rule_name'].str.strip() == rule_name.strip()]
            if len(fila) > 0:
                valores[key] = int(fila['cant_sesiones'].iloc[0])
            else:
                print("    [ADVERTENCIA] No se encontró la regla: {}".format(rule_name))
                valores[key] = 0
    
    return valores

def calcular_csat(valores):
    '''
    Calcula el CSAT (Customer Satisfaction) basado en los valores extraídos
    
    CSAT se calcula como:
    - Positivas = Conforme + Muy conforme (de las 3 categorías)
    - Negativas = Inconforme + Muy inconforme + No sé (de las 3 categorías)
    - CSAT = Positivas / Total
    '''
    # Respuestas Positivas (Conforme + Muy conforme)
    positivas = (
        valores['conforme_integraciones'] +
        valores['muy_conforme_integraciones'] +
        valores['conforme_estaticos'] +
        valores['muy_conforme_estaticos'] +
        valores['conforme_pushes'] +
        valores['muy_conforme_pushes']
    )
    
    # Respuestas Negativas (Inconforme + Muy inconforme + No sé)
    negativas = (
        valores['inconforme_integraciones'] +
        valores['muy_inconforme_integraciones'] +
        valores['no_se_integraciones'] +
        valores['inconforme_estaticos'] +
        valores['muy_inconforme_estaticos'] +
        valores['no_se_estaticos'] +
        valores['inconforme_pushes'] +
        valores['muy_inconforme_pushes'] +
        valores['no_se_pushes']
    )
    
    # Total
    total = positivas + negativas
    
    # CSAT (%)
    if total > 0:
        csat = positivas / total
    else:
        csat = 0
    
    return {
        'positivas': positivas,
        'negativas': negativas,
        'total': total,
        'csat': csat
    }

def create_excel_with_csat(filepath, df, valores, calculos, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea un Excel con la estructura de "CSAT 2025"
    Incluye las 15 reglas específicas y los cálculos de CSAT
    '''
    
    print("    [INFO] Creando Excel detallado con cálculo de CSAT...")
    
    wb = openpyxl.Workbook()
    
    # ==================== HOJA 1: BASE CRUDA ====================
    ws_base = wb.active
    ws_base.title = 'base cruda'
    
    # Estilos para base cruda
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Headers
    ws_base['A1'] = 'rule_name'
    ws_base['B1'] = 'Cant_Sesiones'
    ws_base['A1'].font = header_font
    ws_base['B1'].font = header_font
    ws_base['A1'].fill = header_fill
    ws_base['B1'].fill = header_fill
    
    # Datos - ordenados por cantidad de sesiones
    df_sorted = df.sort_values('cant_sesiones', ascending=False)
    for idx, row in enumerate(df_sorted.iterrows(), start=2):
        ws_base['A{}'.format(idx)] = row[1]['rule_name']
        ws_base['B{}'.format(idx)] = int(row[1]['cant_sesiones'])
    
    # Ajustar anchos
    ws_base.column_dimensions['A'].width = 50
    ws_base.column_dimensions['B'].width = 15
    
    # ==================== HOJA 2: CSAT 2025 ====================
    ws = wb.create_sheet('CSAT 2025')
    
    # Estilos
    title_font = Font(bold=True, size=12)
    category_font = Font(bold=True, size=11)
    csat_font = Font(bold=True, size=12, color="FFFFFF")
    csat_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Determinar el header de fecha
    if modo == 'mes':
        header_fecha = '{} {}'.format(get_month_name(mes), anio)
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{} al {}'.format(
            fecha_inicio_obj.strftime('%d/%m/%Y'),
            fecha_fin_obj.strftime('%d/%m/%Y')
        )
    
    # FILA 1: Header
    ws['A1'] = header_fecha
    ws['A1'].font = title_font
    
    # FILA 2: Headers de columna
    ws['A2'] = 'rule_name'
    ws['B2'] = 'sesiones'
    ws['A2'].font = Font(bold=True)
    ws['B2'].font = Font(bold=True)
    
    row = 3
    
    # INTEGRACIONES
    ws['A{}'.format(row)] = REGLAS_CSAT['conforme_integraciones']
    ws['B{}'.format(row)] = valores['conforme_integraciones']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['inconforme_integraciones']
    ws['B{}'.format(row)] = valores['inconforme_integraciones']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['muy_conforme_integraciones']
    ws['B{}'.format(row)] = valores['muy_conforme_integraciones']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['muy_inconforme_integraciones']
    ws['B{}'.format(row)] = valores['muy_inconforme_integraciones']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['no_se_integraciones']
    ws['B{}'.format(row)] = valores['no_se_integraciones']
    row += 1
    
    # ESTÁTICOS
    ws['A{}'.format(row)] = REGLAS_CSAT['conforme_estaticos']
    ws['B{}'.format(row)] = valores['conforme_estaticos']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['inconforme_estaticos']
    ws['B{}'.format(row)] = valores['inconforme_estaticos']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['muy_conforme_estaticos']
    ws['B{}'.format(row)] = valores['muy_conforme_estaticos']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['muy_inconforme_estaticos']
    ws['B{}'.format(row)] = valores['muy_inconforme_estaticos']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['no_se_estaticos']
    ws['B{}'.format(row)] = valores['no_se_estaticos']
    row += 1
    
    # PUSHES
    ws['A{}'.format(row)] = REGLAS_CSAT['conforme_pushes']
    ws['B{}'.format(row)] = valores['conforme_pushes']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['inconforme_pushes']
    ws['B{}'.format(row)] = valores['inconforme_pushes']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['muy_conforme_pushes']
    ws['B{}'.format(row)] = valores['muy_conforme_pushes']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['muy_inconforme_pushes']
    ws['B{}'.format(row)] = valores['muy_inconforme_pushes']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_CSAT['no_se_pushes']
    ws['B{}'.format(row)] = valores['no_se_pushes']
    row += 1
    
    # Línea en blanco
    row += 1
    
    # RESUMEN
    ws['A{}'.format(row)] = 'Respuestas'
    ws['C{}'.format(row)] = 'Satisfacción'
    ws['A{}'.format(row)].font = category_font
    ws['C{}'.format(row)].font = category_font
    row += 1
    
    # Positivas
    ws['A{}'.format(row)] = 'Positivas'
    ws['B{}'.format(row)] = calculos['positivas']
    ws['C{}'.format(row)] = calculos['csat']
    ws['C{}'.format(row)].number_format = '0.00%'
    row += 1
    
    # Negativas
    ws['A{}'.format(row)] = 'Negativo'
    ws['B{}'.format(row)] = calculos['negativas']
    ws['C{}'.format(row)] = 1 - calculos['csat']
    ws['C{}'.format(row)].number_format = '0.00%'
    row += 1
    
    # Total
    ws['B{}'.format(row)] = calculos['total']
    ws['B{}'.format(row)].font = Font(bold=True)
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # ==================== HOJA 3: RESUMEN EJECUTIVO ====================
    ws_resumen = wb.create_sheet('Resumen')
    
    # Título
    ws_resumen['A1'] = 'FEEDBACK - CSAT (Customer Satisfaction)'
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A2'] = 'Período: {}'.format(header_fecha)
    ws_resumen['A2'].font = Font(size=11)
    
    # Métrica principal
    ws_resumen['A4'] = '🎯 CSAT (Customer Satisfaction)'
    ws_resumen['A4'].font = Font(bold=True, size=12)
    
    ws_resumen['A5'] = 'Satisfacción:'
    ws_resumen['B5'] = calculos['csat']
    ws_resumen['B5'].number_format = '0.00%'
    ws_resumen['B5'].font = csat_font
    ws_resumen['B5'].fill = csat_fill
    ws_resumen['B5'].alignment = Alignment(horizontal='center')
    
    # Desglose
    ws_resumen['A7'] = 'DESGLOSE'
    ws_resumen['A7'].font = Font(bold=True, size=11)
    
    ws_resumen['A8'] = 'Respuestas Positivas:'
    ws_resumen['B8'] = calculos['positivas']
    ws_resumen['A9'] = '  • Conforme'
    ws_resumen['A10'] = '  • Muy conforme'
    
    ws_resumen['A12'] = 'Respuestas Negativas:'
    ws_resumen['B12'] = calculos['negativas']
    ws_resumen['A13'] = '  • Inconforme'
    ws_resumen['A14'] = '  • Muy inconforme'
    ws_resumen['A15'] = '  • No sé'
    
    ws_resumen['A17'] = 'Total de respuestas:'
    ws_resumen['B17'] = calculos['total']
    ws_resumen['B17'].font = Font(bold=True)
    
    # Detalle por categoría
    ws_resumen['A19'] = 'DETALLE POR CATEGORÍA'
    ws_resumen['A19'].font = Font(bold=True, size=11)
    
    integ_pos = valores['conforme_integraciones'] + valores['muy_conforme_integraciones']
    integ_neg = valores['inconforme_integraciones'] + valores['muy_inconforme_integraciones'] + valores['no_se_integraciones']
    
    estat_pos = valores['conforme_estaticos'] + valores['muy_conforme_estaticos']
    estat_neg = valores['inconforme_estaticos'] + valores['muy_inconforme_estaticos'] + valores['no_se_estaticos']
    
    pushes_pos = valores['conforme_pushes'] + valores['muy_conforme_pushes']
    pushes_neg = valores['inconforme_pushes'] + valores['muy_inconforme_pushes'] + valores['no_se_pushes']
    
    ws_resumen['A20'] = 'Integraciones:'
    ws_resumen['B20'] = 'Pos: {:,} | Neg: {:,}'.format(integ_pos, integ_neg)
    
    ws_resumen['A21'] = 'Estáticos:'
    ws_resumen['B21'] = 'Pos: {:,} | Neg: {:,}'.format(estat_pos, estat_neg)
    
    ws_resumen['A22'] = 'Pushes:'
    ws_resumen['B22'] = 'Pos: {:,} | Neg: {:,}'.format(pushes_pos, pushes_neg)
    
    # Interpretación
    ws_resumen['A24'] = 'INTERPRETACIÓN'
    ws_resumen['A24'].font = Font(bold=True, size=11)
    
    if calculos['csat'] >= 0.8:
        interpretacion = "Excelente - Alta satisfacción"
        color = "70AD47"
    elif calculos['csat'] >= 0.6:
        interpretacion = "Bueno - Satisfacción adecuada"
        color = "92D050"
    elif calculos['csat'] >= 0.4:
        interpretacion = "Regular - Requiere mejoras"
        color = "FFC000"
    else:
        interpretacion = "Crítico - Baja satisfacción"
        color = "C00000"
    
    ws_resumen['A25'] = interpretacion
    ws_resumen['A25'].font = Font(bold=True, size=11, color=color)
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 25
    
    # Guardar
    wb.save(filepath)
    print("    [OK] Excel detallado creado con {} hojas".format(len(wb.sheetnames)))

def create_or_update_dashboard_master(filepath, csat_valor, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea o actualiza el Excel Dashboard Master
    Llena SOLO la celda D16 (CSAT - Customer Satisfaction)
    '''
    
    # Verificar si el archivo ya existe
    if os.path.exists(filepath):
        print("    [INFO] Dashboard Master existe, actualizando celda D16...")
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Actualizar SOLO D16
        ws['D16'] = csat_valor
        ws['D16'].number_format = '0.00%'
        
        wb.save(filepath)
        print("    [OK] Dashboard Master actualizado (D16 = {:.2f}%)".format(csat_valor * 100))
    else:
        print("    [INFO] Dashboard Master no existe, creando desde cero...")
        
        # Crear nuevo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dashboard'
        
        header_font = Font(bold=True)
        
        # Determinar el header de fecha
        if modo == 'mes':
            header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
        else:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            header_fecha = '{}-{}'.format(
                fecha_inicio_obj.strftime('%d/%m'),
                fecha_fin_obj.strftime('%d/%m/%y')
            )
        
        # FILA 1: Encabezados
        ws['B1'] = 'Indicador'
        ws['C1'] = 'Descripción/Detalle'
        ws['D1'] = header_fecha
        ws['B1'].font = header_font
        ws['C1'].font = header_font
        ws['D1'].font = header_font
        
        # FILA 2-13: Otros indicadores (vacíos)
        ws['B2'] = 'Conversaciones'
        ws['C2'] = 'Q Conversaciones'
        
        ws['B3'] = 'Usuarios'
        ws['C3'] = 'Q Usuarios únicos'
        
        ws['B4'] = 'Sesiones abiertas por Pushes'
        ws['C4'] = 'Q Sesiones que se abrieron con una Push'
        
        ws['B5'] = 'Sesiones Alcanzadas por Pushes'
        ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'
        
        ws['B6'] = 'Mensajes Pushes Enviados'
        ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'
        
        ws['B7'] = 'Contenidos en Botmaker'
        ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
        
        ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
        ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
        
        ws['B9'] = 'Interacciones'
        ws['C9'] = 'Q Interacciones'
        
        ws['B10'] = 'Trámites, solicitudes y turnos'
        ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'
        
        ws['B11'] = 'contenidos mas consultados'
        ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'
        
        ws['B12'] = 'Derivaciones'
        ws['C12'] = 'Q Derivaciones'
        
        ws['B13'] = 'No entendimiento'
        ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'
        
        # FILA 14-15: Otros indicadores (vacíos)
        ws['B14'] = 'Tasa de Efectividad'
        ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'
        
        ws['B15'] = 'CES (Customer Effort Score)'
        ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'
        
        # FILA 16: CSAT - AQUI VA NUESTRO VALOR
        ws['B16'] = 'Satisfacción (CSAT)'
        ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'
        ws['D16'] = csat_valor
        ws['D16'].number_format = '0.00%'
        
        # FILA 17: Otros indicadores (vacíos)
        ws['B17'] = 'Uptime servidor'
        ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'
        
        # Ajustar anchos
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        
        wb.save(filepath)
        print("    [OK] Dashboard Master creado (D16 = {:.2f}%)".format(csat_valor * 100))

def execute_query_and_save():
    '''Función principal que ejecuta la query y guarda los resultados'''
    
    print("Leyendo configuracion de fechas...")
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = read_date_config(CONFIG['config_file'])
    
    if modo is None:
        print("[ERROR] No se pudo obtener la configuracion. Abortando.")
        return None
    
    print("[OK] Configuracion leida correctamente")
    print("    Modo: {}".format(modo.upper()))
    print("    Periodo: {}".format(descripcion))
    print("    Fecha inicio: {}".format(fecha_inicio))
    print("    Fecha fin: {}".format(fecha_fin))
    
    query = build_query(fecha_inicio, fecha_fin)
    
    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))
    
    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))
    
    try:
        session = boto3.Session(region_name=CONFIG['region'])
        
        print("")
        print("Ejecutando consulta...")
        
        try:
            df = wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                workgroup=CONFIG['workgroup'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False
            )
        except Exception as e:
            if 'workgroup' in str(e).lower():
                print("[ADVERTENCIA] Intentando sin workgroup...")
                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            else:
                raise e
        
        print("[OK] Consulta ejecutada exitosamente!")
        
        # Verificar resultados
        if len(df) == 0:
            print("[ADVERTENCIA] La query no retornó resultados")
            print("    Esto puede significar que no hay reglas CXF en el período especificado")
            return None
        
        # Normalizar nombres de columnas
        df.columns = df.columns.str.lower()
        
        print("")
        print("=" * 60)
        print("RESULTADO - {}".format(descripcion.upper()))
        print("=" * 60)
        print("Total de reglas CXF encontradas: {}".format(len(df)))
        print("Total de sesiones: {:,}".format(df['cant_sesiones'].sum()))
        
        # Extraer valores para CSAT
        print("")
        print("Extrayendo reglas específicas para cálculo de CSAT...")
        valores = extraer_valores_csat(df)
        
        print("")
        print("REGLAS PARA CSAT:")
        print("-" * 60)
        print("  INTEGRACIONES:")
        print("    Conforme:         {:,}".format(valores['conforme_integraciones']))
        print("    Inconforme:       {:,}".format(valores['inconforme_integraciones']))
        print("    Muy conforme:     {:,}".format(valores['muy_conforme_integraciones']))
        print("    Muy inconforme:   {:,}".format(valores['muy_inconforme_integraciones']))
        print("    No sé:            {:,}".format(valores['no_se_integraciones']))
        print("")
        print("  ESTÁTICOS:")
        print("    Conforme:         {:,}".format(valores['conforme_estaticos']))
        print("    Inconforme:       {:,}".format(valores['inconforme_estaticos']))
        print("    Muy conforme:     {:,}".format(valores['muy_conforme_estaticos']))
        print("    Muy inconforme:   {:,}".format(valores['muy_inconforme_estaticos']))
        print("    No sé:            {:,}".format(valores['no_se_estaticos']))
        print("")
        print("  PUSHES:")
        print("    Conforme:         {:,}".format(valores['conforme_pushes']))
        print("    Inconforme:       {:,}".format(valores['inconforme_pushes']))
        print("    Muy conforme:     {:,}".format(valores['muy_conforme_pushes']))
        print("    Muy inconforme:   {:,}".format(valores['muy_inconforme_pushes']))
        print("    No sé:            {:,}".format(valores['no_se_pushes']))
        
        # Calcular CSAT
        calculos = calcular_csat(valores)
        
        print("")
        print("=" * 60)
        print("CÁLCULO DE CSAT (Customer Satisfaction)")
        print("=" * 60)
        print("  Respuestas Positivas:  {:,}".format(calculos['positivas']))
        print("    (Conforme + Muy conforme)")
        print("")
        print("  Respuestas Negativas:  {:,}".format(calculos['negativas']))
        print("    (Inconforme + Muy inconforme + No sé)")
        print("")
        print("  Total de Respuestas:   {:,}".format(calculos['total']))
        print("")
        print("  🎯 CSAT:               {:.2f}%".format(calculos['csat'] * 100))
        print("=" * 60)
        
        filename_csv, filename_excel_detalle, filename_dashboard = generate_filename(modo, mes, anio, fecha_inicio, fecha_fin)
        output_folder = CONFIG['output_folder']
        
        os.makedirs(output_folder, exist_ok=True)
        
        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel_detalle = os.path.join(output_folder, filename_excel_detalle)
        local_path_dashboard = os.path.join(output_folder, filename_dashboard)
        
        print("")
        print("Guardando CSV...")
        df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        
        print("Generando Excel detallado...")
        create_excel_with_csat(local_path_excel_detalle, df, valores, calculos, modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("Generando Dashboard Master...")
        create_or_update_dashboard_master(local_path_dashboard, calculos['csat'], modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("")
        print("ARCHIVOS GENERADOS:")
        print("    [CSV] {}".format(filename_csv))
        print("")
        print("    [EXCEL DETALLE] {}".format(filename_excel_detalle))
        print("            - Hoja 'base cruda': {} reglas".format(len(df)))
        print("            - Hoja 'CSAT 2025': Cálculos detallados")
        print("            - Hoja 'Resumen': Métricas principales")
        print("            - CSAT: {:.2f}%".format(calculos['csat'] * 100))
        print("")
        print("    [DASHBOARD] {}".format(filename_dashboard))
        print("            - Celda D16 = {:.2f}% (CSAT - Customer Satisfaction)".format(calculos['csat'] * 100))
        
        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)
        
        return df
        
    except Exception as e:
        print("")
        print("[ERROR] {}".format(str(e)))
        import traceback
        traceback.print_exc()
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("SCRIPT: FEEDBACK - CSAT (Customer Satisfaction) - QUERY ATHENA")
    print("=" * 60)
    print("Rol requerido: PIBAConsumeBoti")
    print("Salida: CSV + Excel Detalle + Dashboard Master")
    print("Query: Reglas CXF con conteo de sesiones")
    print("")
    print("ARCHIVOS GENERADOS:")
    print("  - CSV: feedback_csat_[fecha].csv")
    print("  - Excel Detalle: feedback_csat_detalle_[fecha].xlsx (3 hojas)")
    print("  - Dashboard: feedback_csat_[fecha].xlsx (celda D16)")
    print("")
    print("CÁLCULOS:")
    print("  - Extrae 15 reglas específicas (Integraciones, Estáticos, Pushes)")
    print("  - Positivas = Conforme + Muy conforme")
    print("  - Negativas = Inconforme + Muy inconforme + No sé")
    print("  - CSAT (%) = Positivas / Total")
    print("")
    print("MODOS:")
    print("  [1] MES COMPLETO: MES + AÑO")
    print("  [2] RANGO PERSONALIZADO: FECHA_INICIO + FECHA_FIN")
    print("=" * 60)
    print("")
    
    result = execute_query_and_save()
    
    if result is not None:
        print("")
        print("Para procesar otro periodo, edita config_fechas.txt")
    else:
        print("")
        print("La ejecucion fallo. Revisa los errores arriba.")
        