# -*- coding: utf-8 -*-
"""
Script AUTOMATICO para extraer Availability de WhatsApp Business API Status
Usa Selenium + ChromeDriver para manejar JavaScript

URL: https://metastatus.com/whatsapp-business-api
Celda Excel: D17 (Uptime servidor / Availability)
"""
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.common.exceptions import WebDriverException
import pandas as pd
from datetime import datetime
import os
import openpyxl
from openpyxl.styles import Font
import re
import time

# ==================== CONFIGURACION ====================
CONFIG = {
    'url': 'https://metastatus.com/whatsapp-business-api',
    'output_folder': 'output',
    'headless': True,  # True = sin ventana, False = con ventana visible
    'wait_time': 10  # Segundos para esperar que cargue la página
}

# ==================== FUNCIONES ====================

def get_month_abbr(mes):
    """Retorna la abreviatura del mes en español"""
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def setup_chrome_driver():
    """Configura el driver de Chrome con opciones"""
    try:
        chrome_options = Options()
        
        if CONFIG['headless']:
            chrome_options.add_argument('--headless')
        
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument('--window-size=1920,1080')
        chrome_options.add_argument('--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36')
        
        driver = webdriver.Chrome(options=chrome_options)
        return driver
        
    except WebDriverException as e:
        print("[ERROR] No se pudo iniciar ChromeDriver")
        print("    Error: {}".format(str(e)))
        print("")
        print("SOLUCIÓN:")
        print("    1. Instalar Google Chrome")
        print("    2. Instalar Selenium: pip install selenium")
        print("    3. ChromeDriver se instala automáticamente con Selenium 4+")
        print("       (Si falla, descarga manual de: https://chromedriver.chromium.org/)")
        return None

def extract_availability_selenium(url):
    """
    Extrae el porcentaje de Availability usando Selenium
    Retorna: (porcentaje, fecha_consulta)
    """
    driver = None
    try:
        print("Iniciando navegador Chrome...")
        driver = setup_chrome_driver()
        
        if driver is None:
            return None, None
        
        print("[OK] Navegador iniciado")
        print("Accediendo a la página...")
        print("    URL: {}".format(url))
        
        driver.get(url)
        
        print("    Esperando {} segundos a que cargue el contenido JavaScript...".format(CONFIG['wait_time']))
        time.sleep(CONFIG['wait_time'])
        
        print("[OK] Página cargada")
        print("Extrayendo porcentaje de Availability...")
        
        # Obtener todo el texto de la página
        page_text = driver.find_element(By.TAG_NAME, 'body').text
        
        # Estrategia 1: Buscar patrones de availability
        patterns = [
            r'(?:Availability|availability|AVAILABILITY)[\s:]*(\d{1,3}\.?\d*)\s*%',
            r'(\d{2,3}\.\d+)\s*%',
            r'Uptime[\s:]*(\d{1,3}\.?\d*)\s*%'
        ]
        
        for pattern in patterns:
            matches = re.findall(pattern, page_text)
            if matches:
                for match in matches:
                    percentage = float(match)
                    if 90 <= percentage <= 100:
                        print("[OK] Porcentaje extraído: {}%".format(percentage))
                        driver.quit()
                        return percentage, datetime.now()
        
        # Estrategia 2: Buscar en elementos específicos
        try:
            elements = driver.find_elements(By.XPATH, "//*[contains(@class, 'metric') or contains(@class, 'availability') or contains(@class, 'uptime')]")
            
            for element in elements:
                text = element.text
                match = re.search(r'(\d{2,3}\.\d+)\s*%', text)
                if match:
                    percentage = float(match.group(1))
                    if 90 <= percentage <= 100:
                        print("[OK] Porcentaje extraído: {}%".format(percentage))
                        driver.quit()
                        return percentage, datetime.now()
        except:
            pass
        
        # Si llegamos aquí, no se encontró
        print("")
        print("[ERROR] =====================================================")
        print("[ERROR] NO SE PUDO EXTRAER EL PORCENTAJE DE AVAILABILITY")
        print("[ERROR] =====================================================")
        print("[ERROR] Primeros 500 caracteres de la página:")
        print(page_text[:500])
        print("[ERROR] ...")
        print("[ERROR] =====================================================")
        
        driver.quit()
        return None, None
        
    except Exception as e:
        print("")
        print("[ERROR] =====================================================")
        print("[ERROR] ERROR DURANTE LA EXTRACCIÓN")
        print("[ERROR] =====================================================")
        print("[ERROR] {}".format(str(e)))
        print("[ERROR] =====================================================")
        
        if driver:
            driver.quit()
        
        return None, None

def create_excel_with_dashboard(filepath, availability_percentage):
    """
    Crea Excel con estructura EXACTA del modelo GCBA
    Solo llena la celda D17 (Uptime servidor / Availability)
    """
    
    print("    [INFO] Creando Excel con estructura Dashboard...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard'
    
    header_font = Font(bold=True)
    
    now = datetime.now()
    mes_abbr = get_month_abbr(now.month)
    header_fecha = '{}-{}'.format(mes_abbr, str(now.year)[-2:])
    
    # FILA 1: Encabezados
    ws['B1'] = 'Indicador'
    ws['C1'] = 'Descripción/Detalle'
    ws['D1'] = header_fecha
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font
    
    # FILA 2: Conversaciones
    ws['B2'] = 'Conversaciones'
    ws['C2'] = 'Q Conversaciones'
    
    # FILA 3: Usuarios
    ws['B3'] = 'Usuarios'
    ws['C3'] = 'Q Usuarios únicos'
    
    # FILA 4: Sesiones abiertas por Pushes
    ws['B4'] = 'Sesiones abiertas por Pushes'
    ws['C4'] = 'Q Sesiones que se abrieron con una Push'
    
    # FILA 5: Sesiones Alcanzadas por Pushes
    ws['B5'] = 'Sesiones Alcanzadas por Pushes'
    ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'
    
    # FILA 6: Mensajes Pushes Enviados
    ws['B6'] = 'Mensajes Pushes Enviados'
    ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'
    
    # FILA 7: Contenidos en Botmaker
    ws['B7'] = 'Contenidos en Botmaker'
    ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
    
    # FILA 8: Contenidos Prendidos para el USUARIO
    ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
    ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
    
    # FILA 9: Interacciones
    ws['B9'] = 'Interacciones'
    ws['C9'] = 'Q Interacciones'
    
    # FILA 10: Trámites, solicitudes y turnos
    ws['B10'] = 'Trámites, solicitudes y turnos'
    ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'
    
    # FILA 11: contenidos mas consultados
    ws['B11'] = 'contenidos mas consultados'
    ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'
    
    # FILA 12: Derivaciones
    ws['B12'] = 'Derivaciones'
    ws['C12'] = 'Q Derivaciones'
    
    # FILA 13: No entendimiento
    ws['B13'] = 'No entendimiento'
    ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'
    
    # FILA 14: Tasa de Efectividad
    ws['B14'] = 'Tasa de Efectividad'
    ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'
    
    # FILA 15: CES (Customer Effort Score)
    ws['B15'] = 'CES (Customer Effort Score)'
    ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'
    
    # FILA 16: Satisfacción (CSAT)
    ws['B16'] = 'Satisfacción (CSAT)'
    ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'
    
    # FILA 17: Uptime servidor (AVAILABILITY) - AQUI VA EL VALOR
    ws['B17'] = 'Uptime servidor'
    ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'
    ws['D17'] = '{}%'.format(availability_percentage)  # ← VALOR DE AVAILABILITY
    
    # Ajustar anchos
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    
    wb.save(filepath)
    print("    [OK] Excel generado: {}".format(filepath))

def generate_filename():
    """Genera nombre de archivo con timestamp"""
    now = datetime.now()
    timestamp = now.strftime('%Y%m%d_%H%M%S')
    filename_csv = "whatsapp_availability_{}.csv".format(timestamp)
    filename_excel = "whatsapp_availability_{}.xlsx".format(timestamp)
    return filename_csv, filename_excel

def execute_and_save():
    """Funcion principal"""
    
    print("=" * 60)
    print("SCRIPT: WHATSAPP BUSINESS API - AVAILABILITY (AUTOMATICO)")
    print("=" * 60)
    print("Método: Selenium + ChromeDriver")
    print("=" * 60)
    print("")
    
    # Extraer availability con Selenium
    availability_percentage, fecha_consulta = extract_availability_selenium(CONFIG['url'])
    
    if availability_percentage is None or fecha_consulta is None:
        print("")
        print("=" * 60)
        print("PROCESO ABORTADO")
        print("=" * 60)
        print("No se pudo extraer el porcentaje de Availability.")
        print("")
        print("REQUISITOS:")
        print("    1. Google Chrome instalado")
        print("    2. pip install selenium")
        print("")
        print("Revisa los errores arriba.")
        print("=" * 60)
        return None
    
    print("")
    print("=" * 60)
    print("RESULTADO")
    print("=" * 60)
    print("Availability: {}%".format(availability_percentage))
    print("Fecha: {}".format(fecha_consulta.strftime('%Y-%m-%d %H:%M:%S')))
    print("=" * 60)
    
    # CSV
    df = pd.DataFrame({
        'Fecha_Consulta': [fecha_consulta.strftime('%Y-%m-%d %H:%M:%S')],
        'Availability_Percentage': [availability_percentage],
        'URL': [CONFIG['url']]
    })
    
    filename_csv, filename_excel = generate_filename()
    output_folder = CONFIG['output_folder']
    
    os.makedirs(output_folder, exist_ok=True)
    
    local_path_csv = os.path.join(output_folder, filename_csv)
    local_path_excel = os.path.join(output_folder, filename_excel)
    
    print("")
    print("Guardando archivos...")
    
    df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
    print("    [CSV] {}".format(filename_csv))
    
    create_excel_with_dashboard(local_path_excel, availability_percentage)
    print("    [EXCEL] {}".format(filename_excel))
    
    print("")
    print("ARCHIVOS GENERADOS:")
    print("    Carpeta: {}/".format(output_folder))
    print("    Celda D17 (Uptime servidor / Availability) = {}%".format(availability_percentage))
    
    print("")
    print("=" * 60)
    print("PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 60)
    
    return df

# ==================== EJECUCION ====================

if __name__ == "__main__":
    print("")
    result = execute_and_save()
    
    if result is not None:
        print("")
        print("Listo! Archivos guardados en output/")
    else:
        print("")
        print("Hubo un problema. Revisa los mensajes arriba.")
