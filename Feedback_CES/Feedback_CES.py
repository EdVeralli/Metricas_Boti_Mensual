# -*- coding: utf-8 -*-
'''
Script para ejecutar query de Feedback - CES (Customer Effort Score) en Athena
Calcula el esfuerzo percibido por los usuarios en escala de 1 a 5
Guarda resultado en CSV y genera Excel con análisis detallado y cálculo de CES
Lee configuracion de fechas desde archivo config_fechas.txt

MODOS SOPORTADOS:
1. MES COMPLETO: Especificar MES y AÑO (comportamiento original)
2. RANGO PERSONALIZADO: Especificar FECHA_INICIO y FECHA_FIN

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti
'''
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': '../config_fechas.txt'  # Config centralizado en raiz del proyecto
}

# Reglas específicas para el cálculo de CES (Customer Effort Score)
REGLAS_CES = {
    'muy_dificil_integraciones': 'CXF01CUX01 Muy difícil Integraciones',
    'dificil_integraciones': 'CXF01CUX01 Difícil Integraciones',
    'mas_o_menos_integraciones': 'CXF01CUX01 Más o menos Integraciones',
    'facil_integraciones': 'CXF01CUX01 Fácil Integraciones',
    'muy_facil_integraciones': 'CXF01CUX01 Muy fácil Integraciones',
    'muy_dificil_estaticos': 'CXF01CUX02 Muy difícil Estáticos',
    'dificil_estaticos': 'CXF01CUX02 Difícil Estáticos',
    'mas_o_menos_estaticos': 'CXF01CUX02 Más o menos Estáticos',
    'facil_estaticos': 'CXF01CUX02 Fácil Estáticos',
    'muy_facil_estaticos': 'CXF01CUX02 Muy fácil Estáticos'
}

# Valores de CES para cada nivel
VALORES_CES = {
    'muy_dificil': 5,
    'dificil': 4,
    'mas_o_menos': 3,
    'facil': 2,
    'muy_facil': 1
}

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    '''
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)
    
    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    '''
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            print("    Creando archivo de ejemplo...")
            
            with open(config_file, 'w', encoding='utf-8') as f:
                f.write("# Configuracion de fechas para el reporte\n")
                f.write("# MODO 1: Mes completo\n")
                f.write("MES=10\n")
                f.write("AÑO=2025\n\n")
                f.write("# MODO 2: Rango personalizado\n")
                f.write("# FECHA_INICIO=2025-10-01\n")
                f.write("# FECHA_FIN=2025-10-15\n")
            
            print("    Archivo creado: {}".format(config_file))
            return 'mes', None, None, 10, 2025, "octubre 2025"
        
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
                
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()
        
        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                
                if fecha_inicio > fecha_fin:
                    print("[ERROR] FECHA_INICIO no puede ser posterior a FECHA_FIN")
                    return None, None, None, None, None, None
                
                descripcion = "{} al {}".format(
                    fecha_inicio.strftime('%d/%m/%Y'),
                    fecha_fin.strftime('%d/%m/%Y')
                )
                
                print("[INFO] Modo: RANGO PERSONALIZADO")
                return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion
                
            except ValueError as e:
                print("[ERROR] Formato de fecha invalido. Use YYYY-MM-DD (ej: 2025-10-01)")
                return None, None, None, None, None, None
        
        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            if mes < 1 or mes > 12:
                print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
                return None, None, None, None, None, None
            
            if anio < 2020 or anio > 2030:
                print("[ADVERTENCIA] Año inusual: {}".format(anio))
            
            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
            
            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)
            
            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion
        
        print("[ERROR] El archivo {} no contiene configuracion valida".format(config_file))
        return None, None, None, None, None, None
        
    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    '''Retorna la abreviatura del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_query(fecha_inicio, fecha_fin):
    '''Construye la query de Feedback - CES con el rango de fechas especificado'''
    
    query = """SELECT 
rule_name,
count(distinct(session_id)) as Cant_Sesiones 
FROM "caba-piba-consume-zone-db"."boti_message_metrics_2"
WHERE CAST(session_creation_time AS DATE) BETWEEN date '{fecha_inicio}' and date '{fecha_fin}'
and rule_name like ('%CXF%')
group by rule_name""".format(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    
    return query

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    '''Genera los nombres de archivos basados en el modo y las fechas'''
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "feedback_ces_{0}_{1}.csv".format(mes_nombre, anio)
        filename_excel_detalle = "feedback_ces_detalle_{0}_{1}.xlsx".format(mes_nombre, anio)
        filename_dashboard = "feedback_ces_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "feedback_ces_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_excel_detalle = "feedback_ces_detalle_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_dashboard = "feedback_ces_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
    
    return filename_csv, filename_excel_detalle, filename_dashboard

def extraer_valores_ces(df):
    '''
    Extrae los valores de las 10 reglas específicas para el cálculo de CES
    Retorna un diccionario con los valores encontrados
    '''
    valores = {}
    
    for key, rule_name in REGLAS_CES.items():
        # Buscar la regla en el DataFrame
        fila = df[df['rule_name'] == rule_name]
        
        if len(fila) > 0:
            valores[key] = int(fila['cant_sesiones'].iloc[0])
        else:
            print("    [ADVERTENCIA] No se encontró la regla: {}".format(rule_name))
            valores[key] = 0
    
    return valores

def calcular_ces(valores):
    '''
    Calcula el CES (Customer Effort Score) basado en los valores extraídos
    
    CES se calcula como promedio ponderado:
    - Muy difícil = 5 (peor)
    - Difícil = 4
    - Más o menos = 3
    - Fácil = 2
    - Muy fácil = 1 (mejor)
    '''
    # Sumar sesiones por nivel (Integraciones + Estáticos)
    muy_dificil_total = valores['muy_dificil_integraciones'] + valores['muy_dificil_estaticos']
    dificil_total = valores['dificil_integraciones'] + valores['dificil_estaticos']
    mas_o_menos_total = valores['mas_o_menos_integraciones'] + valores['mas_o_menos_estaticos']
    facil_total = valores['facil_integraciones'] + valores['facil_estaticos']
    muy_facil_total = valores['muy_facil_integraciones'] + valores['muy_facil_estaticos']
    
    # Calcular ponderados (valor * sesiones)
    muy_dificil_ponderado = VALORES_CES['muy_dificil'] * muy_dificil_total
    dificil_ponderado = VALORES_CES['dificil'] * dificil_total
    mas_o_menos_ponderado = VALORES_CES['mas_o_menos'] * mas_o_menos_total
    facil_ponderado = VALORES_CES['facil'] * facil_total
    muy_facil_ponderado = VALORES_CES['muy_facil'] * muy_facil_total
    
    # Totales
    total_sesiones = (muy_dificil_total + dificil_total + mas_o_menos_total + 
                     facil_total + muy_facil_total)
    total_ponderado = (muy_dificil_ponderado + dificil_ponderado + mas_o_menos_ponderado + 
                      facil_ponderado + muy_facil_ponderado)
    
    # CES = promedio ponderado
    if total_sesiones > 0:
        ces = total_ponderado / total_sesiones
    else:
        ces = 0
    
    return {
        'muy_dificil_total': muy_dificil_total,
        'dificil_total': dificil_total,
        'mas_o_menos_total': mas_o_menos_total,
        'facil_total': facil_total,
        'muy_facil_total': muy_facil_total,
        'muy_dificil_ponderado': muy_dificil_ponderado,
        'dificil_ponderado': dificil_ponderado,
        'mas_o_menos_ponderado': mas_o_menos_ponderado,
        'facil_ponderado': facil_ponderado,
        'muy_facil_ponderado': muy_facil_ponderado,
        'total_sesiones': total_sesiones,
        'total_ponderado': total_ponderado,
        'ces': ces
    }

def create_excel_with_ces(filepath, df, valores, calculos, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea un Excel con la estructura de "esfuerzo CES"
    Incluye las 10 reglas específicas y los cálculos de CES
    '''
    
    print("    [INFO] Creando Excel detallado con cálculo de CES...")
    
    wb = openpyxl.Workbook()
    
    # ==================== HOJA 1: BASE CRUDA ====================
    ws_base = wb.active
    ws_base.title = 'base cruda'
    
    # Estilos para base cruda
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Headers
    ws_base['A1'] = 'rule_name'
    ws_base['B1'] = 'Cant_Sesiones'
    ws_base['A1'].font = header_font
    ws_base['B1'].font = header_font
    ws_base['A1'].fill = header_fill
    ws_base['B1'].fill = header_fill
    
    # Datos - ordenados por cantidad de sesiones
    df_sorted = df.sort_values('cant_sesiones', ascending=False)
    for idx, row in enumerate(df_sorted.iterrows(), start=2):
        ws_base['A{}'.format(idx)] = row[1]['rule_name']
        ws_base['B{}'.format(idx)] = int(row[1]['cant_sesiones'])
    
    # Ajustar anchos
    ws_base.column_dimensions['A'].width = 50
    ws_base.column_dimensions['B'].width = 15
    
    # ==================== HOJA 2: ESFUERZO CES ====================
    ws = wb.create_sheet('esfuerzo CES')
    
    # Estilos
    title_font = Font(bold=True, size=12)
    header_font_table = Font(bold=True, size=11)
    ces_font = Font(bold=True, size=12, color="FFFFFF")
    ces_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Determinar el header de fecha
    if modo == 'mes':
        header_fecha = '{} {}'.format(get_month_name(mes), anio)
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{} al {}'.format(
            fecha_inicio_obj.strftime('%d/%m/%Y'),
            fecha_fin_obj.strftime('%d/%m/%Y')
        )
    
    # FILA 1: Header
    ws['A1'] = header_fecha
    ws['A1'].font = title_font
    
    # FILA 2: Headers de tabla
    ws['A2'] = 'Respuesta'
    ws['B2'] = 'Valores CES'
    ws['C2'] = 'Total'
    ws['D2'] = 'Ponderado'
    ws['A2'].font = header_font_table
    ws['B2'].font = header_font_table
    ws['C2'].font = header_font_table
    ws['D2'].font = header_font_table
    
    # FILAS 3-7: Tabla de resultados
    ws['A3'] = 'Muy difícil'
    ws['B3'] = VALORES_CES['muy_dificil']
    ws['C3'] = calculos['muy_dificil_total']
    ws['D3'] = calculos['muy_dificil_ponderado']
    
    ws['A4'] = 'Difícil'
    ws['B4'] = VALORES_CES['dificil']
    ws['C4'] = calculos['dificil_total']
    ws['D4'] = calculos['dificil_ponderado']
    
    ws['A5'] = 'Más o menos'
    ws['B5'] = VALORES_CES['mas_o_menos']
    ws['C5'] = calculos['mas_o_menos_total']
    ws['D5'] = calculos['mas_o_menos_ponderado']
    
    ws['A6'] = 'Fácil'
    ws['B6'] = VALORES_CES['facil']
    ws['C6'] = calculos['facil_total']
    ws['D6'] = calculos['facil_ponderado']
    
    ws['A7'] = 'Muy fácil'
    ws['B7'] = VALORES_CES['muy_facil']
    ws['C7'] = calculos['muy_facil_total']
    ws['D7'] = calculos['muy_facil_ponderado']
    
    # FILA 8: Totales
    ws['A8'] = 'Totales:'
    ws['A8'].font = Font(bold=True)
    ws['C8'] = calculos['total_sesiones']
    ws['D8'] = calculos['total_ponderado']
    ws['C8'].font = Font(bold=True)
    ws['D8'].font = Font(bold=True)
    
    # FILA 9: CES
    ws['A9'] = 'CES'
    ws['A9'].font = Font(bold=True, size=12)
    ws['C9'] = calculos['ces']
    ws['C9'].font = ces_font
    ws['C9'].fill = ces_fill
    ws['C9'].number_format = '0.00'
    
    # Línea en blanco
    row = 11
    
    # DETALLE DE REGLAS
    ws['A{}'.format(row)] = 'DETALLE POR REGLA'
    ws['A{}'.format(row)].font = Font(bold=True)
    row += 1
    
    # Integraciones
    ws['A{}'.format(row)] = 'INTEGRACIONES:'
    ws['A{}'.format(row)].font = Font(bold=True, size=10)
    row += 1
    
    for key in ['dificil_integraciones', 'facil_integraciones', 'mas_o_menos_integraciones', 
                'muy_dificil_integraciones', 'muy_facil_integraciones']:
        ws['A{}'.format(row)] = REGLAS_CES[key]
        ws['B{}'.format(row)] = valores[key]
        row += 1
    
    # Estáticos
    row += 1
    ws['A{}'.format(row)] = 'ESTÁTICOS:'
    ws['A{}'.format(row)].font = Font(bold=True, size=10)
    row += 1
    
    for key in ['dificil_estaticos', 'facil_estaticos', 'mas_o_menos_estaticos',
                'muy_dificil_estaticos', 'muy_facil_estaticos']:
        ws['A{}'.format(row)] = REGLAS_CES[key]
        ws['B{}'.format(row)] = valores[key]
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # ==================== HOJA 3: RESUMEN EJECUTIVO ====================
    ws_resumen = wb.create_sheet('Resumen')
    
    # Título
    ws_resumen['A1'] = 'FEEDBACK - CES (Customer Effort Score)'
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A2'] = 'Período: {}'.format(header_fecha)
    ws_resumen['A2'].font = Font(size=11)
    
    # Métrica principal
    ws_resumen['A4'] = '🎯 CES (Customer Effort Score)'
    ws_resumen['A4'].font = Font(bold=True, size=12)
    
    ws_resumen['A5'] = 'Puntuación CES:'
    ws_resumen['B5'] = calculos['ces']
    ws_resumen['B5'].number_format = '0.00'
    ws_resumen['B5'].font = ces_font
    ws_resumen['B5'].fill = ces_fill
    ws_resumen['B5'].alignment = Alignment(horizontal='center')
    
    ws_resumen['A6'] = 'Escala: 1 (muy fácil) a 5 (muy difícil)'
    ws_resumen['A6'].font = Font(size=9, italic=True)
    
    # Distribución
    ws_resumen['A8'] = 'DISTRIBUCIÓN DE RESPUESTAS'
    ws_resumen['A8'].font = Font(bold=True, size=11)
    
    ws_resumen['A9'] = 'Muy fácil (1):'
    ws_resumen['B9'] = calculos['muy_facil_total']
    
    ws_resumen['A10'] = 'Fácil (2):'
    ws_resumen['B10'] = calculos['facil_total']
    
    ws_resumen['A11'] = 'Más o menos (3):'
    ws_resumen['B11'] = calculos['mas_o_menos_total']
    
    ws_resumen['A12'] = 'Difícil (4):'
    ws_resumen['B12'] = calculos['dificil_total']
    
    ws_resumen['A13'] = 'Muy difícil (5):'
    ws_resumen['B13'] = calculos['muy_dificil_total']
    
    ws_resumen['A15'] = 'Total de respuestas:'
    ws_resumen['B15'] = calculos['total_sesiones']
    ws_resumen['B15'].font = Font(bold=True)
    
    # Interpretación
    ws_resumen['A17'] = 'INTERPRETACIÓN'
    ws_resumen['A17'].font = Font(bold=True, size=11)
    
    if calculos['ces'] <= 2:
        interpretacion = "Excelente - Muy fácil de usar"
        color = "70AD47"
    elif calculos['ces'] <= 3:
        interpretacion = "Bueno - Experiencia satisfactoria"
        color = "FFC000"
    elif calculos['ces'] <= 4:
        interpretacion = "Regular - Requiere mejoras"
        color = "FF6600"
    else:
        interpretacion = "Crítico - Experiencia muy difícil"
        color = "C00000"
    
    ws_resumen['A18'] = interpretacion
    ws_resumen['A18'].font = Font(bold=True, size=11, color=color)
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 35
    ws_resumen.column_dimensions['B'].width = 20
    
    # Guardar
    wb.save(filepath)
    print("    [OK] Excel detallado creado con {} hojas".format(len(wb.sheetnames)))

def create_or_update_dashboard_master(filepath, ces_valor, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea o actualiza el Excel Dashboard Master
    Llena SOLO la celda D15 (CES - Customer Effort Score)
    '''
    
    # Verificar si el archivo ya existe
    if os.path.exists(filepath):
        print("    [INFO] Dashboard Master existe, actualizando celda D15...")
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Actualizar SOLO D15
        ws['D15'] = ces_valor
        ws['D15'].number_format = '0.00'
        
        wb.save(filepath)
        print("    [OK] Dashboard Master actualizado (D15 = {:.2f})".format(ces_valor))
    else:
        print("    [INFO] Dashboard Master no existe, creando desde cero...")
        
        # Crear nuevo workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dashboard'
        
        header_font = Font(bold=True)
        
        # Determinar el header de fecha
        if modo == 'mes':
            header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
        else:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            header_fecha = '{}-{}'.format(
                fecha_inicio_obj.strftime('%d/%m'),
                fecha_fin_obj.strftime('%d/%m/%y')
            )
        
        # FILA 1: Encabezados
        ws['B1'] = 'Indicador'
        ws['C1'] = 'Descripción/Detalle'
        ws['D1'] = header_fecha
        ws['B1'].font = header_font
        ws['C1'].font = header_font
        ws['D1'].font = header_font
        
        # FILA 2-13: Otros indicadores (vacíos)
        ws['B2'] = 'Conversaciones'
        ws['C2'] = 'Q Conversaciones'
        
        ws['B3'] = 'Usuarios'
        ws['C3'] = 'Q Usuarios únicos'
        
        ws['B4'] = 'Sesiones abiertas por Pushes'
        ws['C4'] = 'Q Sesiones que se abrieron con una Push'
        
        ws['B5'] = 'Sesiones Alcanzadas por Pushes'
        ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'
        
        ws['B6'] = 'Mensajes Pushes Enviados'
        ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'
        
        ws['B7'] = 'Contenidos en Botmaker'
        ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
        
        ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
        ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
        
        ws['B9'] = 'Interacciones'
        ws['C9'] = 'Q Interacciones'
        
        ws['B10'] = 'Trámites, solicitudes y turnos'
        ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'
        
        ws['B11'] = 'contenidos mas consultados'
        ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'
        
        ws['B12'] = 'Derivaciones'
        ws['C12'] = 'Q Derivaciones'
        
        ws['B13'] = 'No entendimiento'
        ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'
        
        # FILA 14: Tasa de Efectividad (vacío)
        ws['B14'] = 'Tasa de Efectividad'
        ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'
        
        # FILA 15: CES - AQUI VA NUESTRO VALOR
        ws['B15'] = 'CES (Customer Effort Score)'
        ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'
        ws['D15'] = ces_valor
        ws['D15'].number_format = '0.00'
        
        # FILA 16-17: Otros indicadores (vacíos)
        ws['B16'] = 'Satisfacción (CSAT)'
        ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'
        
        ws['B17'] = 'Uptime servidor'
        ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'
        
        # Ajustar anchos
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        
        wb.save(filepath)
        print("    [OK] Dashboard Master creado (D15 = {:.2f})".format(ces_valor))

def execute_query_and_save():
    '''Función principal que ejecuta la query y guarda los resultados'''
    
    print("Leyendo configuracion de fechas...")
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = read_date_config(CONFIG['config_file'])
    
    if modo is None:
        print("[ERROR] No se pudo obtener la configuracion. Abortando.")
        return None
    
    print("[OK] Configuracion leida correctamente")
    print("    Modo: {}".format(modo.upper()))
    print("    Periodo: {}".format(descripcion))
    print("    Fecha inicio: {}".format(fecha_inicio))
    print("    Fecha fin: {}".format(fecha_fin))
    
    query = build_query(fecha_inicio, fecha_fin)
    
    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))
    
    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))
    
    try:
        session = boto3.Session(region_name=CONFIG['region'])
        
        print("")
        print("Ejecutando consulta...")
        
        try:
            df = wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                workgroup=CONFIG['workgroup'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False
            )
        except Exception as e:
            if 'workgroup' in str(e).lower():
                print("[ADVERTENCIA] Intentando sin workgroup...")
                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            else:
                raise e
        
        print("[OK] Consulta ejecutada exitosamente!")
        
        # Verificar resultados
        if len(df) == 0:
            print("[ADVERTENCIA] La query no retornó resultados")
            print("    Esto puede significar que no hay reglas CXF en el período especificado")
            return None
        
        # Normalizar nombres de columnas
        df.columns = df.columns.str.lower()
        
        print("")
        print("=" * 60)
        print("RESULTADO - {}".format(descripcion.upper()))
        print("=" * 60)
        print("Total de reglas CXF encontradas: {}".format(len(df)))
        print("Total de sesiones: {:,}".format(df['cant_sesiones'].sum()))
        
        # Extraer valores para CES
        print("")
        print("Extrayendo reglas específicas para cálculo de CES...")
        valores = extraer_valores_ces(df)
        
        print("")
        print("REGLAS PARA CES:")
        print("-" * 60)
        print("  INTEGRACIONES:")
        print("    Muy difícil:   {:,}".format(valores['muy_dificil_integraciones']))
        print("    Difícil:       {:,}".format(valores['dificil_integraciones']))
        print("    Más o menos:   {:,}".format(valores['mas_o_menos_integraciones']))
        print("    Fácil:         {:,}".format(valores['facil_integraciones']))
        print("    Muy fácil:     {:,}".format(valores['muy_facil_integraciones']))
        print("")
        print("  ESTÁTICOS:")
        print("    Muy difícil:   {:,}".format(valores['muy_dificil_estaticos']))
        print("    Difícil:       {:,}".format(valores['dificil_estaticos']))
        print("    Más o menos:   {:,}".format(valores['mas_o_menos_estaticos']))
        print("    Fácil:         {:,}".format(valores['facil_estaticos']))
        print("    Muy fácil:     {:,}".format(valores['muy_facil_estaticos']))
        
        # Calcular CES
        calculos = calcular_ces(valores)
        
        print("")
        print("=" * 60)
        print("CÁLCULO DE CES (Customer Effort Score)")
        print("=" * 60)
        print("  Muy difícil (5):   {:,} sesiones → {:,} ponderado".format(
            calculos['muy_dificil_total'], calculos['muy_dificil_ponderado']))
        print("  Difícil (4):       {:,} sesiones → {:,} ponderado".format(
            calculos['dificil_total'], calculos['dificil_ponderado']))
        print("  Más o menos (3):   {:,} sesiones → {:,} ponderado".format(
            calculos['mas_o_menos_total'], calculos['mas_o_menos_ponderado']))
        print("  Fácil (2):         {:,} sesiones → {:,} ponderado".format(
            calculos['facil_total'], calculos['facil_ponderado']))
        print("  Muy fácil (1):     {:,} sesiones → {:,} ponderado".format(
            calculos['muy_facil_total'], calculos['muy_facil_ponderado']))
        print("")
        print("  Total sesiones:    {:,}".format(calculos['total_sesiones']))
        print("  Total ponderado:   {:,}".format(calculos['total_ponderado']))
        print("")
        print("  🎯 CES:            {:.2f}".format(calculos['ces']))
        print("     (Escala: 1=muy fácil, 5=muy difícil)")
        print("=" * 60)
        
        filename_csv, filename_excel_detalle, filename_dashboard = generate_filename(modo, mes, anio, fecha_inicio, fecha_fin)
        output_folder = CONFIG['output_folder']
        
        os.makedirs(output_folder, exist_ok=True)
        
        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel_detalle = os.path.join(output_folder, filename_excel_detalle)
        local_path_dashboard = os.path.join(output_folder, filename_dashboard)
        
        print("")
        print("Guardando CSV...")
        df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        
        print("Generando Excel detallado...")
        create_excel_with_ces(local_path_excel_detalle, df, valores, calculos, modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("Generando Dashboard Master...")
        create_or_update_dashboard_master(local_path_dashboard, calculos['ces'], modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("")
        print("ARCHIVOS GENERADOS:")
        print("    [CSV] {}".format(filename_csv))
        print("")
        print("    [EXCEL DETALLE] {}".format(filename_excel_detalle))
        print("            - Hoja 'base cruda': {} reglas".format(len(df)))
        print("            - Hoja 'esfuerzo CES': Cálculos detallados")
        print("            - Hoja 'Resumen': Métricas principales")
        print("            - CES: {:.2f}".format(calculos['ces']))
        print("")
        print("    [DASHBOARD] {}".format(filename_dashboard))
        print("            - Celda D15 = {:.2f} (CES - Customer Effort Score)".format(calculos['ces']))
        
        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)
        
        return df
        
    except Exception as e:
        print("")
        print("[ERROR] {}".format(str(e)))
        import traceback
        traceback.print_exc()
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("SCRIPT: FEEDBACK - CES (Customer Effort Score) - QUERY ATHENA")
    print("=" * 60)
    print("Rol requerido: PIBAConsumeBoti")
    print("Salida: CSV + Excel Detalle + Dashboard Master")
    print("Query: Reglas CXF con conteo de sesiones")
    print("")
    print("ARCHIVOS GENERADOS:")
    print("  - CSV: feedback_ces_[fecha].csv")
    print("  - Excel Detalle: feedback_ces_detalle_[fecha].xlsx (3 hojas)")
    print("  - Dashboard: feedback_ces_[fecha].xlsx (celda D15)")
    print("")
    print("CÁLCULOS:")
    print("  - Extrae 10 reglas específicas (Integraciones y Estáticos)")
    print("  - Calcula puntuación ponderada (1-5)")
    print("  - CES = Total ponderado / Total sesiones")
    print("  - 1 = Muy fácil (mejor), 5 = Muy difícil (peor)")
    print("")
    print("MODOS:")
    print("  [1] MES COMPLETO: MES + AÑO")
    print("  [2] RANGO PERSONALIZADO: FECHA_INICIO + FECHA_FIN")
    print("=" * 60)
    print("")
    
    result = execute_query_and_save()
    
    if result is not None:
        print("")
        print("Para procesar otro periodo, edita config_fechas.txt")
    else:
        print("")
        print("La ejecucion fallo. Revisa los errores arriba.")