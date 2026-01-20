"""
Calculo de Tasa de Efectividad WEB+BOTI

Este script calcula la tasa de efectividad combinada entre Boti y Web,
leyendo datos de ambos repositorios y generando un Excel con los resultados.

Fuentes de datos:
- Metricas_Boti_Mensual: Feedback_Efectividad/output/feedback_efectividad_{mes}_{año}_efectividad.xlsx
- Metricas_Web_Mensual: Satisfaccion/data/conteo_completo_{mes}_{año}.xlsx

Salida:
- efectividad_web_boti/efectividad_web_boti_{mes}_{año}.xlsx
"""

import os
import re
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Mapeo de numero de mes a nombre en español
MESES = {
    1: "enero", 2: "febrero", 3: "marzo", 4: "abril",
    5: "mayo", 6: "junio", 7: "julio", 8: "agosto",
    9: "septiembre", 10: "octubre", 11: "noviembre", 12: "diciembre"
}


def leer_config_fechas(config_path: str) -> dict:
    """Lee el archivo de configuracion de fechas."""
    config = {}

    with open(config_path, 'r', encoding='utf-8') as f:
        for linea in f:
            linea = linea.strip()
            # Ignorar comentarios y lineas vacias
            if not linea or linea.startswith('#'):
                continue

            if '=' in linea:
                clave, valor = linea.split('=', 1)
                clave = clave.strip()
                valor = valor.strip()

                if clave in ['MES', 'AÑO']:
                    config[clave] = int(valor)
                elif clave in ['FECHA_INICIO', 'FECHA_FIN']:
                    config[clave] = valor

    return config


def obtener_mes_año(config: dict) -> tuple:
    """Obtiene mes y año de la configuracion."""
    # Si hay rango personalizado, usar el mes/año de FECHA_INICIO
    if 'FECHA_INICIO' in config and 'FECHA_FIN' in config:
        fecha = datetime.strptime(config['FECHA_INICIO'], '%Y-%m-%d')
        return fecha.month, fecha.year
    else:
        return config['MES'], config['AÑO']


def leer_valor_celda(ruta_excel: str, celda: str, hoja: int = 0) -> any:
    """Lee el valor de una celda especifica de un Excel."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active
    valor = ws[celda].value
    wb.close()
    return valor


def leer_valor_columna(ruta_excel: str, nombre_columna: str, fila_dato: int = 2) -> any:
    """Lee el valor de una columna por nombre de header."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb.active

    # Buscar la columna por nombre
    valor = None
    for cell in ws[1]:
        if cell.value == nombre_columna:
            valor = ws.cell(row=fila_dato, column=cell.column).value
            break

    wb.close()
    return valor


def calcular_efectividad_web_boti(
    efectividad_positiva_boti: float,
    total_boti: int,
    total_web: int,
    tasa_efectividad_web: float,
    total_general_web: int
) -> dict:
    """
    Calcula la Tasa de Efectividad WEB+BOTI.

    Retorna un diccionario con todos los calculos intermedios y el resultado final.
    """
    # Total General
    total_general = total_boti + total_web

    # Ponderacion Feedback Boti
    ponderacion_feedback_boti = total_boti / total_general

    # Primer Parcial General
    primer_parcial_general = efectividad_positiva_boti * ponderacion_feedback_boti

    # Ponderacion Feedback WEB
    ponderacion_feedback_web = total_general_web / total_general

    # Convertir tasa efectividad web a decimal si viene en porcentaje
    if tasa_efectividad_web > 1:
        tasa_efectividad_web_decimal = tasa_efectividad_web / 100
    else:
        tasa_efectividad_web_decimal = tasa_efectividad_web

    # Segundo Parcial General
    segundo_parcial_general = ponderacion_feedback_web * tasa_efectividad_web_decimal

    # Tasa de Efectividad WEB+BOTI
    tasa_efectividad_web_boti = primer_parcial_general + segundo_parcial_general

    return {
        'efectividad_positiva_boti': efectividad_positiva_boti,
        'total_boti': total_boti,
        'total_web': total_web,
        'tasa_efectividad_web': tasa_efectividad_web,
        'tasa_efectividad_web_decimal': tasa_efectividad_web_decimal,
        'total_general_web': total_general_web,
        'total_general': total_general,
        'ponderacion_feedback_boti': ponderacion_feedback_boti,
        'primer_parcial_general': primer_parcial_general,
        'ponderacion_feedback_web': ponderacion_feedback_web,
        'segundo_parcial_general': segundo_parcial_general,
        'tasa_efectividad_web_boti': tasa_efectividad_web_boti
    }


def generar_excel_resultado(resultados: dict, mes: int, año: int, ruta_salida: str):
    """Genera el Excel con los resultados y calculos intermedios."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Efectividad WEB+BOTI"

    # Estilos
    titulo_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
    header_font = Font(name='Arial', size=11, bold=True)
    valor_font = Font(name='Arial', size=11)
    resultado_font = Font(name='Arial', size=14, bold=True, color='FFFFFF')

    titulo_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    resultado_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')

    nombre_mes = MESES[mes].capitalize()

    # Titulo principal
    ws.merge_cells('A1:C1')
    ws['A1'] = f"Tasa de Efectividad WEB+BOTI - {nombre_mes} {año}"
    ws['A1'].font = titulo_font
    ws['A1'].fill = titulo_fill
    ws['A1'].alignment = center_align

    # Fecha de generacion
    ws.merge_cells('A2:C2')
    ws['A2'] = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = center_align

    # Seccion: Valores de Entrada
    fila = 4
    ws.merge_cells(f'A{fila}:C{fila}')
    ws[f'A{fila}'] = "VALORES DE ENTRADA"
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].fill = header_fill
    ws[f'A{fila}'].alignment = center_align

    datos_entrada = [
        ("Efectividad Positiva Boti (C28)", resultados['efectividad_positiva_boti'], True),
        ("Total Boti (B30)", resultados['total_boti'], False),
        ("Total Web (S2)", resultados['total_web'], False),
        ("Tasa Efectividad WEB (T2)", resultados['tasa_efectividad_web'], True),
        ("Total General WEB (S2)", resultados['total_general_web'], False),
    ]

    fila += 1
    ws[f'A{fila}'] = "Variable"
    ws[f'B{fila}'] = "Valor"
    ws[f'C{fila}'] = "Fuente"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{fila}'].font = header_font
        ws[f'{col}{fila}'].border = border

    fuentes = [
        "Metricas_Boti_Mensual",
        "Metricas_Boti_Mensual",
        "Metricas_Web_Mensual",
        "Metricas_Web_Mensual",
        "Metricas_Web_Mensual",
    ]

    for i, (nombre, valor, es_porcentaje) in enumerate(datos_entrada):
        fila += 1
        ws[f'A{fila}'] = nombre
        ws[f'A{fila}'].alignment = left_align
        ws[f'A{fila}'].border = border

        if es_porcentaje:
            ws[f'B{fila}'] = valor
            ws[f'B{fila}'].number_format = '0.00%'
        else:
            ws[f'B{fila}'] = valor
            ws[f'B{fila}'].number_format = '#,##0'
        ws[f'B{fila}'].alignment = center_align
        ws[f'B{fila}'].border = border

        ws[f'C{fila}'] = fuentes[i]
        ws[f'C{fila}'].alignment = center_align
        ws[f'C{fila}'].border = border

    # Seccion: Calculos Intermedios
    fila += 2
    ws.merge_cells(f'A{fila}:C{fila}')
    ws[f'A{fila}'] = "CALCULOS INTERMEDIOS"
    ws[f'A{fila}'].font = header_font
    ws[f'A{fila}'].fill = header_fill
    ws[f'A{fila}'].alignment = center_align

    calculos = [
        ("Total General", resultados['total_general'], False, "Total Boti + Total Web"),
        ("Ponderacion Feedback Boti", resultados['ponderacion_feedback_boti'], True, "Total Boti / Total General"),
        ("Primer Parcial General", resultados['primer_parcial_general'], True, "Efectividad Positiva Boti × Ponderacion Feedback Boti"),
        ("Ponderacion Feedback WEB", resultados['ponderacion_feedback_web'], True, "Total General WEB / Total General"),
        ("Segundo Parcial General", resultados['segundo_parcial_general'], True, "Ponderacion Feedback WEB × Tasa Efectividad WEB"),
    ]

    fila += 1
    ws[f'A{fila}'] = "Calculo"
    ws[f'B{fila}'] = "Valor"
    ws[f'C{fila}'] = "Formula"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{fila}'].font = header_font
        ws[f'{col}{fila}'].border = border

    for nombre, valor, es_porcentaje, formula in calculos:
        fila += 1
        ws[f'A{fila}'] = nombre
        ws[f'A{fila}'].alignment = left_align
        ws[f'A{fila}'].border = border

        if es_porcentaje:
            ws[f'B{fila}'] = valor
            ws[f'B{fila}'].number_format = '0.00%'
        else:
            ws[f'B{fila}'] = valor
            ws[f'B{fila}'].number_format = '#,##0'
        ws[f'B{fila}'].alignment = center_align
        ws[f'B{fila}'].border = border

        ws[f'C{fila}'] = formula
        ws[f'C{fila}'].alignment = left_align
        ws[f'C{fila}'].border = border

    # Seccion: Resultado Final
    fila += 2
    ws.merge_cells(f'A{fila}:C{fila}')
    ws[f'A{fila}'] = "RESULTADO FINAL"
    ws[f'A{fila}'].font = resultado_font
    ws[f'A{fila}'].fill = resultado_fill
    ws[f'A{fila}'].alignment = center_align

    fila += 1
    ws.merge_cells(f'A{fila}:B{fila}')
    ws[f'A{fila}'] = "Tasa de Efectividad WEB+BOTI"
    ws[f'A{fila}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{fila}'].alignment = left_align
    ws[f'A{fila}'].border = border

    ws[f'C{fila}'] = resultados['tasa_efectividad_web_boti']
    ws[f'C{fila}'].font = Font(name='Arial', size=14, bold=True)
    ws[f'C{fila}'].number_format = '0.00%'
    ws[f'C{fila}'].alignment = center_align
    ws[f'C{fila}'].fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')
    ws[f'C{fila}'].border = border

    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 50

    # Guardar
    os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
    wb.save(ruta_salida)
    print(f"Excel generado: {ruta_salida}")


def main():
    # Rutas base
    repo_boti = Path(__file__).parent
    repo_web = repo_boti.parent / "Metricas_Web_Mensual"

    # Leer configuracion de fechas
    config_path = repo_boti / "config_fechas.txt"
    config = leer_config_fechas(str(config_path))
    mes, año = obtener_mes_año(config)
    nombre_mes = MESES[mes]

    print(f"Procesando: {nombre_mes.capitalize()} {año}")
    print("=" * 50)

    # Rutas de archivos fuente
    excel_boti = repo_boti / "Feedback_Efectividad" / "output" / f"feedback_efectividad_{nombre_mes}_{año}_efectividad.xlsx"
    excel_web = repo_web / "Satisfaccion" / "data" / f"conteo_completo_{nombre_mes}_{año}.xlsx"

    # Verificar que existen los archivos
    if not excel_boti.exists():
        raise FileNotFoundError(f"No se encuentra: {excel_boti}")
    if not excel_web.exists():
        raise FileNotFoundError(f"No se encuentra: {excel_web}")

    print(f"Leyendo: {excel_boti.name}")
    print(f"Leyendo: {excel_web.name}")
    print()

    # Leer valores de entrada
    efectividad_positiva_boti = leer_valor_celda(str(excel_boti), 'C28')
    total_boti = leer_valor_celda(str(excel_boti), 'B30')

    total_web = leer_valor_columna(str(excel_web), 'Total_General', fila_dato=2)
    tasa_efectividad_web = leer_valor_columna(str(excel_web), 'Tasa_Efectividad', fila_dato=2)
    total_general_web = leer_valor_columna(str(excel_web), 'Total_General', fila_dato=2)

    print("Valores de entrada:")
    print(f"  Efectividad Positiva Boti: {efectividad_positiva_boti:.4f} ({efectividad_positiva_boti*100:.2f}%)")
    print(f"  Total Boti: {total_boti:,}")
    print(f"  Total Web: {total_web:,}")
    print(f"  Tasa Efectividad WEB: {tasa_efectividad_web}")
    print(f"  Total General WEB: {total_general_web:,}")
    print()

    # Calcular
    resultados = calcular_efectividad_web_boti(
        efectividad_positiva_boti=efectividad_positiva_boti,
        total_boti=total_boti,
        total_web=total_web,
        tasa_efectividad_web=tasa_efectividad_web,
        total_general_web=total_general_web
    )

    print("Calculos intermedios:")
    print(f"  Total General: {resultados['total_general']:,}")
    print(f"  Ponderacion Feedback Boti: {resultados['ponderacion_feedback_boti']:.4f} ({resultados['ponderacion_feedback_boti']*100:.2f}%)")
    print(f"  Primer Parcial General: {resultados['primer_parcial_general']:.4f}")
    print(f"  Ponderacion Feedback WEB: {resultados['ponderacion_feedback_web']:.4f} ({resultados['ponderacion_feedback_web']*100:.2f}%)")
    print(f"  Segundo Parcial General: {resultados['segundo_parcial_general']:.4f}")
    print()

    print("=" * 50)
    print(f"TASA DE EFECTIVIDAD WEB+BOTI: {resultados['tasa_efectividad_web_boti']*100:.2f}%")
    print("=" * 50)

    # Generar Excel
    ruta_salida = repo_boti / "efectividad_web_boti" / f"efectividad_web_boti_{nombre_mes}_{año}.xlsx"
    generar_excel_resultado(resultados, mes, año, str(ruta_salida))

    print()
    print("Proceso completado exitosamente.")


if __name__ == "__main__":
    main()
