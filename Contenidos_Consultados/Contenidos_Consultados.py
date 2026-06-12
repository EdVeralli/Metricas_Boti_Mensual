# -*- coding: utf-8 -*-
'''
Script para descargar y procesar Contenidos Consultados desde Athena
Descarga la vista boti_vw_buscador_rulename y genera tabla completa de contenidos.

Implementa filtrado en 2 capas como el Power BI original:
  - CAPA 1: Filtro dinámico por patrones (CONTAINS) - extraído del PDF de lógica
  - CAPA 2: Lista fija de exclusiones manuales para casos puntuales

Lógica extraída del Power BI "Consultas por dia 1.pbix":
1. Filtrar por rango de fechas del período
2. CAPA 1: Excluir por patrones dinámicos (CONTAINS)
3. CAPA 2: Excluir por lista fija de nombres exactos
4. Extraer prefijo (RulenameUnique) y agrupar por (prefijo, fecha)
5. Quedarse con el rulename con más sesiones por (prefijo, fecha)
6. Sumar sesiones por rulename y calcular %GT
7. Generar serie temporal diaria (Histórico)

MODOS SOPORTADOS:
1. MES COMPLETO: Especificar MES y AÑO
2. RANGO PERSONALIZADO: Especificar FECHA_INICIO y FECHA_FIN

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBADataScientist
'''
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import re
import csv
import glob
import sys
import time
from contextlib import contextmanager
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ==================== HELPERS DE LOGGING (verbose) ====================
# Forzar flush en cada print porque Windows bufferea la salida y los pasos
# largos parecen colgados. Tambien agregar timestamp para visibilidad.

def log(msg=""):
    '''Print con timestamp y flush inmediato (clave en consolas Windows).'''
    if msg:
        print("[{:%H:%M:%S}] {}".format(datetime.now(), msg), flush=True)
    else:
        print("", flush=True)

@contextmanager
def step(nombre):
    '''
    Context manager que imprime inicio/fin de un paso pesado con tiempo
    transcurrido. Asi el usuario sabe si esta avanzando o se trabo.
    '''
    t0 = time.time()
    log(">>> INICIO: {}".format(nombre))
    try:
        yield
    finally:
        elapsed = time.time() - t0
        log("<<< FIN:    {} ({:.1f}s)".format(nombre, elapsed))

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': '../config_fechas.txt'  # Config centralizado en raiz del proyecto
}

# Flag para activar/desactivar las exclusiones de contenidos.
# Poner en False para ver TODOS los contenidos sin filtrar.
APLICAR_EXCLUSIONES = True

# ==================== CAPA 1: PATRONES DINAMICOS ====================
# Extraídos del PDF "Lógica de armado de contenidos más consultados"
# Se excluye toda rule_name que CONTENGA alguno de estos textos (case-insensitive)
PATRONES_EXCLUIR = [
    'push',
    'recordatorio',
    'confirmación',
    'confirmacion',
    'cancelacion',
    'cancelación',
    'CXF',
    'CAT',
    'api',
    'onboarding',
    'menú',
    'menu',
    'Invocar',
    'Bifurcador',
    'No entendidos',
    'No, nada de eso',
    'No entend',
    'Orquestador',
    'CTA',
    'Atiende',
    'Agente',
    'V2',
    'v2',
    'Instancia',
    'SIGECI',
    'USIG',
    'Botonera',
    '_autenticacion',
    '_',        # NOTA: Excluye todo lo que contenga guión bajo (muy agresivo, viene del PDF)
    'Solicitud',
    'Reconocer',
    'Terminar',
    'Cierre',
    'Chau',
    'miBA',
    'BOT0',
]

# Compilar regex una sola vez para mejor rendimiento
# Se usa re.IGNORECASE para que coincida sin importar mayúsculas/minúsculas
# Se escapan los patrones para evitar problemas con caracteres especiales
_PATRON_REGEX = re.compile(
    '|'.join(re.escape(p) for p in PATRONES_EXCLUIR),
    re.IGNORECASE
)

# ==================== CAPA 2: EXCLUSIONES MANUALES ====================
# Para contenidos puntuales que no caen en ningún patrón de la Capa 1
# pero igual deben excluirse.
# NOTA: Muchos de los items del script original ya caen por la Capa 1,
# acá solo van los que NO serían atrapados por los patrones.
CONTENIDOS_EXCLUIR_MANUAL = [
    # Coyunturas (no caen en patrones)
    'Coyuntura (cierre de estación Carlos Gardel línea B)',

    # Login y miBA (no contienen ningún patrón de Capa 1)
    '3. Login miBA',
    '3.1 Login miBA',
    'miBA - Login exitoso',

    # Navegación
    'Más temas post Menú 2.0',  # ya cae por 'menu' pero se deja por seguridad

    # Atención
    'Transferir con un agente',  # ya cae por 'Agente' pero se deja por seguridad
    '147 - ¿Te puedo ayudar en algo más?',
    'Cancelar',

    # No entendidos
    'No entendió letra no existente en WA',  # ya cae por 'No entend' pero se deja
    'X. Buscaba otra cosa',

    # Otros internos
    'TUR01CUX13 Preguntar género',
    'MO05CUX01 - Sexo',
    'Puede estacionar CTA',  # ya cae por 'CTA' pero se deja por seguridad

    # Subcontenidos de Licencias (no deben aparecer en el ranking)
    # "Licencias" (MO05CUX02 Apertura) es la entrada principal que agrupa el trámite.
    # Los subcontenidos de ese flujo se excluyen para evitar duplicación en el top 10:
    'Licencia prorroga  > Consultar',  # subcontenido de Licencias - fue feb-2026 puesto 9 (62.240) - DOBLE ESPACIO es como viene en los datos
    'MO05CUX01 > Tiene que renovar',   # subcontenido de Licencias - fue feb-2026 puesto 10 (47.990)
    'LIC00CUX00 Validaciones',         # subcontenido de Licencias

    # Subcontenidos de Motovehiculos (no deben aparecer en el ranking)
    # "Trámites Vehículos SAP" (MO00CUX01 Auto o moto) es la entrada principal.
    'MO08CUX03 Apertura',              # subcontenido de Trámites Vehículos SAP - fue ene-2026 puesto 10 (28.077)

    # Subcontenidos de Salud / Turnos (no deben aparecer en el ranking)
    'SA06CUX03 Desambiguar confirmarción',  # subcontenido del flujo SA06CUX03

]

# ==================== NOMBRES AMIGABLES ====================
# Mapeo de rulename técnico a nombre amigable para el reporte.
# Se aplica al resultado final (Excel detallado + Dashboard).
NOMBRES_AMIGABLES = {
    'SA06CUX03 Confirmar turno Sí, dale': 'Confirmación Turnos Salud',
    'TUR00CUX02 Turnos para salud': 'Turnos Salud',
    'Infracciones * Apertura': 'Infracciones',
    'TUR00CUX02 Mensaje inicial - Turnos': 'Turnos SAP',
    'SA06CUX02 Apertura': 'Mis profesionales',
    'MO00CUX01 Auto o moto': 'Trámites Vehículos SAP',
    'SA10CUX01 Hospitales': 'Hospitales',
    'MO05CUX02 Apertura': 'Licencias',
    'TR00CUX01 Apertura': 'Trámites SAP',
    'Busca donde está permitido estacionar': 'Busca donde está permitido estacionar',

    # Agregados desde Trasco.csv (mapeo oficial rulename -> Topico)
    'SA00CUX01 Superapertura': 'SAP de Salud',
    'LIC00CUX01 Telefono': 'Otorgamiento de Licencias Exp unificada',
    'TUR00CUX02 Hospitales y CeSAC': 'Bifurcador Salud BAX',
    'TUR00CUX02 Turnos para trámites': 'Turnos para trámites',
    'TUR00CUX02 Para vehículos': 'Turnos para vehículos',
    'MO00CUX01 Estacionar en BA': 'Estacionar en BA',
    'MO00CUX01 Taxi': 'BA Taxi',
    'DE00CUX01 Apertura': 'Deportes SAP',
    'VI00CUX01 Apertura': 'Viviendas SAP',
    'ED00CUX01 Superapertura educación': 'Educación SAP',
    'CLI00CUX01 Apertura': 'Clima SAP',
}

# ==================== AGRUPACIÓN ESPECIAL POR PREFIJO ====================
# Lista de prefijos cuyo flujo (varios rulenames con el mismo prefijo) debe
# colapsarse a un solo rulename: el de mayor cantidad de sesiones gana, los
# demas se eliminan del ranking.
# DEJAR LISTA VACIA si no se quiere agrupar ningun prefijo (todos los rulenames
# aparecen separados, como muestra el Power BI historico).
#
# Nota: TUR00CUX02 estaba aca historicamente, pero ahora se desea que sus
# distintos rulenames (Turnos para salud / Mensaje inicial - Turnos / etc.)
# aparezcan separados en el ranking, asi que la lista quedo vacia.
PREFIJOS_AGRUPAR = []

# La regla de agrupacion recien empieza a aplicar desde esta fecha (inclusive, YYYY-MM-DD).
# Solo tiene efecto cuando PREFIJOS_AGRUPAR no esta vacia.
PREFIJOS_AGRUPAR_DESDE = '2026-05-01'

# ==================== CAPA 3: REPORTAR Y EXCLUIR ====================
# Rulenames que se procesan COMO SI estuvieran incluidos para calcular su posicion
# en el ranking, y luego se excluyen del resultado final.
# Util para saber si un contenido excluido habria entrado al top 10 (sin que aparezca).
# Match EXACTO sobre el rulename completo.
RULENAMES_REPORTAR_Y_EXCLUIR = [
    'SA06CUX02 Mensaje de error general',
]

# ==================== RULENAMES_PRESERVAR (paso 5) ====================
# Por defecto, el paso 5 consolida rulenames del mismo prefijo dejando solo
# el de mayor sesiones por dia (ej: para SA06CUX03 se queda solo el ganador
# diario y se descartan las otras variantes).
#
# Los rulenames en esta lista quedan EXENTOS de esa consolidacion: siempre se
# preservan TODAS sus filas, asi mantienen el conteo de sesiones completo y
# aparecen en el ranking con su volumen real.
# Match EXACTO sobre el rulename completo.
RULENAMES_PRESERVAR = [
    'TUR00CUX02 Mensaje inicial - Turnos',  # mapea a 'Turnos SAP' en NOMBRES_AMIGABLES
    'SA06CUX02 Apertura',                   # mapea a 'Mis profesionales' en NOMBRES_AMIGABLES

]

# ==================== DESCRIPCIONES AUTOMATICAS DESDE TSV ====================
# Cuando un rulename empieza con un codigo del estilo SA03CUX01 (sin nombre amigable manual),
# se busca ese codigo en la columna 'Topic' del TSV de reglas mas reciente y se usa
# la descripcion que sigue al codigo.
# Ejemplo: rulename "SA03CUX01 Apertura" + Topic "01. SA03CUX01 - Vacunacion"
#          --> Nombre Amigable = "Vacunacion"
_PATRON_CODIGO = re.compile(r'^[A-Z]{2,4}\d{2}CUX\d{2}$')
_PATRON_TOPIC_CODIGO = re.compile(r'\b([A-Z]{2,4}\d{2}CUX\d{2})\b\s*[-:.]*\s*(.+)$')
_DESCRIPCIONES_TSV_CACHE = None  # se carga lazy en la primera llamada
_NAME_A_TOPIC_CACHE = None       # cache para el lookup por columna Name (TSV mas reciente gana)

# Disclaimer que se agrega cuando un rulename no tiene nombre amigable en ningun lado.
DISCLAIMER_FALTANTE = '[NOMBRE AMIGABLE FALTANTE]'

# ==================== FUNCIONES ====================

def _find_all_tsvs():
    '''Busca TODOS los TSV de reglas en el directorio del script (rules-*.tsv).
    Devuelve la lista ordenada ascendentemente (mas viejos primero), para que al
    consolidar los mas nuevos sobreescriban a los mas viejos.'''
    script_dir = os.path.dirname(os.path.abspath(__file__))
    return sorted(glob.glob(os.path.join(script_dir, 'rules-*.tsv')))

def cargar_descripciones_tsv():
    '''
    Lee TODOS los TSV de reglas (rules-*.tsv) del directorio del script y arma
    un dict {codigo: descripcion} a partir de la columna 'Topic'.

    En caso de que un mismo codigo aparezca en varios TSV con descripciones
    distintas, gana el TSV mas reciente (los archivos se procesan en orden
    ascendente, por eso los posteriores sobreescriben).

    Cachea el resultado en memoria para no releer los archivos en cada llamada.
    Encoding latin-1 (ISO-8859) que es como vienen los exports de Botmaker.
    '''
    global _DESCRIPCIONES_TSV_CACHE
    if _DESCRIPCIONES_TSV_CACHE is not None:
        return _DESCRIPCIONES_TSV_CACHE

    descripciones = {}
    tsv_paths = _find_all_tsvs()
    if not tsv_paths:
        print("    [WARN] No se encontro ningun rules-*.tsv para descripciones automaticas")
        _DESCRIPCIONES_TSV_CACHE = descripciones
        return descripciones

    archivos_info = []  # (filename, codigos_en_archivo, codigos_nuevos_aportados, codigos_actualizados)
    for tsv_path in tsv_paths:
        codigos_archivo = 0
        codigos_nuevos = 0
        codigos_actualizados = 0
        try:
            with open(tsv_path, 'r', encoding='latin-1', newline='') as f:
                reader = csv.DictReader(f, delimiter='\t')
                vistos_en_archivo = set()
                for row in reader:
                    topic = (row.get('Topic') or '').strip()
                    if not topic:
                        continue
                    m = _PATRON_TOPIC_CODIGO.search(topic)
                    if m:
                        code = m.group(1).upper()
                        desc = m.group(2).strip().rstrip('.,;:')
                        if not desc:
                            continue
                        if code in vistos_en_archivo:
                            # Mismo codigo aparece varias veces en este TSV, ya lo procesamos
                            continue
                        vistos_en_archivo.add(code)
                        codigos_archivo += 1
                        if code not in descripciones:
                            codigos_nuevos += 1
                        elif descripciones[code] != desc:
                            codigos_actualizados += 1
                        # Latest wins: el actual (mas nuevo) sobreescribe al anterior
                        descripciones[code] = desc
            archivos_info.append((os.path.basename(tsv_path), codigos_archivo, codigos_nuevos, codigos_actualizados))
        except Exception as e:
            print("    [WARN] Error leyendo {}: {}".format(tsv_path, e))

    print("    [INFO] Descripciones automaticas consolidadas desde {} TSV (total: {} codigos unicos):".format(
        len(archivos_info), len(descripciones)
    ))
    for nombre, total_archivo, nuevos, actualizados in archivos_info:
        extras = ""
        if actualizados > 0:
            extras = ", {} actualizados por nuevo".format(actualizados)
        print("           - {}: {} codigos en archivo ({} nuevos{})".format(
            nombre, total_archivo, nuevos, extras
        ))

    _DESCRIPCIONES_TSV_CACHE = descripciones
    return descripciones

def aplicar_descripcion_automatica(rulename, descripciones_tsv):
    '''
    Si el rulename empieza con un codigo conocido en el TSV, devuelve la descripcion.
    Si no matchea el patron de codigo o no esta en el dict, devuelve None.
    '''
    if not rulename or ' ' not in str(rulename):
        return None
    primer_token = str(rulename).split(' ')[0]
    if not _PATRON_CODIGO.fullmatch(primer_token):
        return None
    return descripciones_tsv.get(primer_token.upper())

def cargar_mapping_name_a_topic():
    '''
    Lee TODOS los TSV de reglas (rules-*.tsv) y arma un dict:
        Name (rulename) -> (Topic, Topic path)

    Empezando por el TSV MAS NUEVO: si un mismo rulename aparece en varios TSV,
    gana el del archivo mas reciente.

    Cachea en memoria. Encoding latin-1 (export Botmaker).
    '''
    global _NAME_A_TOPIC_CACHE
    if _NAME_A_TOPIC_CACHE is not None:
        return _NAME_A_TOPIC_CACHE

    mapping = {}
    # Los TSV vienen ordenados ASC (mas viejos primero). Recorrer en orden DESC
    # asi el mas nuevo se procesa primero y gana al hacer setdefault.
    tsv_paths = list(reversed(_find_all_tsvs()))
    if not tsv_paths:
        print("    [WARN] No se encontro ningun rules-*.tsv para mapping name->topic")
        _NAME_A_TOPIC_CACHE = mapping
        return mapping

    for tsv_path in tsv_paths:
        try:
            with open(tsv_path, 'r', encoding='latin-1', newline='') as f:
                reader = csv.DictReader(f, delimiter='\t')
                for row in reader:
                    name = (row.get('Name') or '').strip()
                    if not name:
                        continue
                    # setdefault asegura que el TSV mas nuevo (procesado primero) gane
                    if name not in mapping:
                        topic = (row.get('Topic') or '').strip()
                        topic_path = (row.get('Topic path') or '').strip()
                        mapping[name] = (topic, topic_path)
        except Exception as e:
            print("    [WARN] Error leyendo {} para mapping Name: {}".format(tsv_path, e))

    print("    [INFO] Mapping Name->Topic cargado: {} rulenames unicos".format(len(mapping)))
    _NAME_A_TOPIC_CACHE = mapping
    return mapping

def resolver_nombre_amigable_por_name(rulename, mapping_name_a_topic):
    '''
    Logica NUEVA: busca el rulename en la columna 'Name' del TSV y arma el nombre
    amigable a partir de las columnas 'Topic' y 'Topic path'.

    Reglas:
      - Si el rulename no esta en el mapping  --> None (cae al disclaimer)
      - Si Topic esta vacio                    --> None (cae al disclaimer)
      - Si Topic == Topic path                 --> "Topic [rulename]"
      - Si Topic != Topic path                 --> "Topic" (tal cual)

    Para que coincida con strip (los rulenames pueden tener espacios extra), se
    intenta primero el match exacto y despues el match con strip.
    '''
    if not rulename:
        return None
    # Match exacto
    if rulename in mapping_name_a_topic:
        topic, topic_path = mapping_name_a_topic[rulename]
    else:
        # Match con strip (algunos rulenames tienen espacios al final/inicio)
        rn_strip = str(rulename).strip()
        clave_match = None
        for k in mapping_name_a_topic:
            if k.strip() == rn_strip:
                clave_match = k
                break
        if clave_match is None:
            return None
        topic, topic_path = mapping_name_a_topic[clave_match]

    if not topic:
        return None

    if topic == topic_path:
        return "{} [{}]".format(topic, rulename)
    return topic

def read_date_config(config_file):
    '''
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)

    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    '''
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            print("    Creando archivo de ejemplo...")

            with open(config_file, 'w', encoding='utf-8') as f:
                f.write("# Configuracion de fechas para el reporte\n")
                f.write("# MODO 1: Mes completo\n")
                f.write("MES=1\n")
                f.write("AÑO=2026\n\n")
                f.write("# MODO 2: Rango personalizado\n")
                f.write("# FECHA_INICIO=2026-01-01\n")
                f.write("# FECHA_FIN=2026-01-31\n")

            print("    Archivo creado: {}".format(config_file))
            return 'mes', None, None, 1, 2026, "enero 2026"

        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None

        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)

                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)

                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()

                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()

        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')

                if fecha_inicio > fecha_fin:
                    print("[ERROR] FECHA_INICIO no puede ser posterior a FECHA_FIN")
                    return None, None, None, None, None, None

                descripcion = "{} al {}".format(
                    fecha_inicio.strftime('%d/%m/%Y'),
                    fecha_fin.strftime('%d/%m/%Y')
                )

                print("[INFO] Modo: RANGO PERSONALIZADO")
                return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion

            except ValueError as e:
                print("[ERROR] Formato de fecha invalido. Use YYYY-MM-DD (ej: 2026-01-01)")
                return None, None, None, None, None, None

        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            if mes < 1 or mes > 12:
                print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
                return None, None, None, None, None, None

            if anio < 2020 or anio > 2030:
                print("[ADVERTENCIA] Año inusual: {}".format(anio))

            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)

            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)

            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion

        print("[ERROR] El archivo {} no contiene configuracion valida".format(config_file))
        return None, None, None, None, None, None

    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    '''Retorna la abreviatura del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_query():
    '''Construye la query para descargar la vista de contenidos consultados'''
    query = 'SELECT * FROM "caba-piba-consume-zone-db"."boti_vw_buscador_rulename"'
    return query

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    '''Genera los nombres de archivos basados en el modo y las fechas'''
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "contenidos_consultados_{0}_{1}.csv".format(mes_nombre, anio)
        filename_detalle = "contenidos_consultados_detalle_{0}_{1}.xlsx".format(mes_nombre, anio)
        filename_dashboard = "contenidos_consultados_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "contenidos_consultados_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_detalle = "contenidos_consultados_detalle_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_dashboard = "contenidos_consultados_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)

    return filename_csv, filename_detalle, filename_dashboard

def filtrar_por_patrones(df, rulename_col):
    '''
    CAPA 1: Filtra contenidos usando patrones dinámicos (CONTAINS).
    Replica la lógica del Power BI documentada en el PDF.

    Excluye toda rule_name que contenga alguno de los patrones definidos en PATRONES_EXCLUIR.

    Retorna: (df_filtrado, cantidad_excluidos, detalle_por_patron)
    '''
    total_antes = len(df)

    # Crear máscara: True si el rulename contiene algún patrón
    mascara_excluir = df[rulename_col].astype(str).str.contains(_PATRON_REGEX, na=False)

    # Detalle: contar cuántos excluye cada patrón individual (para el log)
    detalle = {}
    for patron in PATRONES_EXCLUIR:
        patron_regex = re.compile(re.escape(patron), re.IGNORECASE)
        coincidencias = df[rulename_col].astype(str).str.contains(patron_regex, na=False).sum()
        if coincidencias > 0:
            detalle[patron] = coincidencias

    df_filtrado = df[~mascara_excluir].copy()
    cantidad_excluidos = total_antes - len(df_filtrado)

    return df_filtrado, cantidad_excluidos, detalle

def filtrar_por_lista_manual(df, rulename_col):
    '''
    CAPA 2: Filtra contenidos usando la lista fija de exclusiones manuales.
    Para los casos puntuales que no caen en ningún patrón de la Capa 1.

    Retorna: (df_filtrado, cantidad_excluidos)
    '''
    total_antes = len(df)
    df_filtrado = df[~df[rulename_col].isin(CONTENIDOS_EXCLUIR_MANUAL)].copy()
    cantidad_excluidos = total_antes - len(df_filtrado)
    return df_filtrado, cantidad_excluidos

def procesar_contenidos(df, fecha_inicio, fecha_fin):
    '''
    Procesa el DataFrame descargado y calcula las métricas de contenidos consultados.

    Lógica (extraída del Power BI "Consultas por dia 1" + PDF de lógica):
    1. Filtrar por rango de fechas
    2. CAPA 1: Excluir por patrones dinámicos (como el PBI)
    3. CAPA 2: Excluir por lista fija de nombres exactos (respaldo)
    4. Sumar sesiones por rulename individual
    5. Agrupación especial TUR00CUX02: quedarse solo con la de mayor valor
    6. Aplicar nombres amigables y calcular %GT
    7. Serie temporal diaria (Histórico)

    Retorna: (df_contenidos, total_contenidos, df_historico)
    '''
    print("    [INFO] Procesando datos...")

    # Normalizar nombres de columnas
    df.columns = df.columns.str.lower().str.strip()

    # Mostrar columnas disponibles
    print("    [INFO] Columnas disponibles: {}".format(', '.join(df.columns.tolist())))

    # Detectar columna de fecha
    fecha_col = None
    for col in ['fecha', 'date', 'session_date', 'dia']:
        if col in df.columns:
            fecha_col = col
            break

    # Detectar columna de rulename
    rulename_col = None
    for col in ['rulename', 'rule_name', 'rulename_con_max_sesiones', 'contenido']:
        if col in df.columns:
            rulename_col = col
            break

    # Detectar columna de sesiones
    sesiones_col = None
    for col in ['suma_de_sesiones_diarias', 'sesiones_diarias', 'max_sesiones_diarias', 'sesiones', 'cant_sesiones', 'count']:
        if col in df.columns:
            sesiones_col = col
            break

    if rulename_col is None:
        print("    [ERROR] No se encontró columna de rulename")
        return None, None, None

    if sesiones_col is None:
        print("    [ERROR] No se encontró columna de sesiones")
        return None, None, None

    print("    [INFO] Usando columnas: rulename='{}', sesiones='{}'".format(rulename_col, sesiones_col))

    df_filtrado = df.copy()

    # 1. Filtrar por fecha si existe la columna
    if fecha_col:
        print("    [INFO] Filtrando por fecha: {} a {}".format(fecha_inicio, fecha_fin))
        df_filtrado[fecha_col] = pd.to_datetime(df_filtrado[fecha_col])
        fecha_inicio_dt = pd.to_datetime(fecha_inicio)
        fecha_fin_dt = pd.to_datetime(fecha_fin)
        df_filtrado = df_filtrado[
            (df_filtrado[fecha_col] >= fecha_inicio_dt) &
            (df_filtrado[fecha_col] <= fecha_fin_dt)
        ]
        print("    [INFO] Registros después de filtro de fecha: {:,}".format(len(df_filtrado)))

    # 2. CAPA 1: Excluir por patrones dinámicos
    if APLICAR_EXCLUSIONES:
        log("")
        log("    === CAPA 1: Filtro por patrones dinámicos (CONTAINS) ===")
        log("    Patrones configurados: {}".format(len(PATRONES_EXCLUIR)))
        with step("    CAPA 1 - aplicando {} patrones".format(len(PATRONES_EXCLUIR))):
            df_filtrado, excluidos_capa1, detalle_patrones = filtrar_por_patrones(df_filtrado, rulename_col)
        log("    Registros excluidos por patrones: {:,}".format(excluidos_capa1))
        log("    Registros restantes: {:,}".format(len(df_filtrado)))

        # Mostrar top 10 patrones que más excluyeron
        if detalle_patrones:
            top_patrones = sorted(detalle_patrones.items(), key=lambda x: x[1], reverse=True)[:10]
            print("    [INFO] Top patrones con más coincidencias:")
            for patron, count in top_patrones:
                print("           '{}': {:,} registros".format(patron, count))

        # 3. CAPA 2: Excluir por lista fija manual
        log("")
        log("    === CAPA 2: Filtro por lista fija manual ===")
        log("    Items en lista manual: {}".format(len(CONTENIDOS_EXCLUIR_MANUAL)))
        with step("    CAPA 2 - aplicando lista manual"):
            df_filtrado, excluidos_capa2 = filtrar_por_lista_manual(df_filtrado, rulename_col)
        log("    Registros excluidos por lista manual: {:,}".format(excluidos_capa2))
        log("    Registros restantes: {:,}".format(len(df_filtrado)))

        total_excluidos = excluidos_capa1 + excluidos_capa2
        print("")
        print("    [INFO] TOTAL excluidos (ambas capas): {:,}".format(total_excluidos))
    else:
        print("    [INFO] Exclusiones DESACTIVADAS (APLICAR_EXCLUSIONES = False)")

    # 4. Extraer prefijo (RulenameUnique) - lógica del PBI:
    #    Primer palabra antes del espacio, o el rulename completo si no tiene espacio
    df_filtrado['_RulenameUnique'] = df_filtrado[rulename_col].apply(
        lambda x: str(x).split(' ')[0] if ' ' in str(x) else str(x)
    )

    # 5. Para cada (prefijo, fecha): quedarse con el rulename con más sesiones ese día
    #    Esto evita que aparezcan múltiples rulenames del mismo grupo (ej: SA06CUX03)
    #    EXCEPCIÓN: los rulenames en RULENAMES_PRESERVAR quedan exentos de la
    #    consolidacion: se preservan TODAS sus filas, asi conservan su volumen total.
    if fecha_col:
        with step("    Agrupacion por (prefijo, fecha) - max por dia"):
            idx_max = df_filtrado.groupby(['_RulenameUnique', fecha_col])[sesiones_col].idxmax()
            indices_finales = set(idx_max.tolist())
            preservados = 0
            if RULENAMES_PRESERVAR:
                mask_preservar = df_filtrado[rulename_col].isin(RULENAMES_PRESERVAR)
                idx_preservar = df_filtrado[mask_preservar].index
                indices_nuevos = set(idx_preservar.tolist()) - indices_finales
                preservados = len(indices_nuevos)
                indices_finales.update(idx_preservar.tolist())
            df_filtrado = df_filtrado.loc[sorted(indices_finales)].copy()
        log("    Registros tras agrupación por prefijo (max por día): {:,}".format(len(df_filtrado)))
        if preservados > 0:
            log("    {} filas adicionales preservadas (rulenames en RULENAMES_PRESERVAR)".format(preservados))

    # 6. Sumar sesiones por rulename (tras la agrupación por prefijo)
    df_agrupado = df_filtrado.groupby(rulename_col)[sesiones_col].sum().reset_index()
    df_agrupado.columns = ['Rulename', 'Suma de Sesiones']

    # 7. Agrupación especial para prefijos en PREFIJOS_AGRUPAR (ej: TUR00CUX02)
    #    Después de la agrupación por prefijo, pueden quedar múltiples rulenames
    #    del mismo prefijo (porque cada una ganó en días distintos).
    #    Para estos prefijos: quedarse SOLO con la de mayor valor total.
    #    NOTA: Solo se aplica desde PREFIJOS_AGRUPAR_DESDE en adelante.
    aplicar_agrupar = bool(PREFIJOS_AGRUPAR) and fecha_inicio and fecha_inicio >= PREFIJOS_AGRUPAR_DESDE
    if not PREFIJOS_AGRUPAR:
        print("    [INFO] PREFIJOS_AGRUPAR vacio: ningun prefijo se agrupa, todos los rulenames aparecen separados")
    elif not aplicar_agrupar:
        print("    [INFO] Regla PREFIJOS_AGRUPAR NO aplicada (período < {}): los prefijos agrupados aparecen separados".format(PREFIJOS_AGRUPAR_DESDE))
    else:
        for prefijo in PREFIJOS_AGRUPAR:
            mascara = df_agrupado['Rulename'].str.startswith(prefijo)
            df_prefijo = df_agrupado[mascara]
            if len(df_prefijo) > 1:
                idx_max = df_prefijo['Suma de Sesiones'].idxmax()
                ganadora = df_agrupado.loc[idx_max, 'Rulename']
                eliminadas = df_prefijo[df_prefijo.index != idx_max]
                print("    [INFO] Agrupación {}: ganó '{}' ({:,})".format(
                    prefijo, ganadora, int(df_agrupado.loc[idx_max, 'Suma de Sesiones'])))
                for _, row in eliminadas.iterrows():
                    print("           Eliminada: '{}' ({:,})".format(row['Rulename'], int(row['Suma de Sesiones'])))
                df_agrupado = df_agrupado.drop(eliminadas.index)

    # 6. Aplicar nombres amigables (cascada: NOMBRES_AMIGABLES manual > descripcion del TSV > rulename original)
    df_agrupado['Nombre Amigable'] = df_agrupado['Rulename'].map(NOMBRES_AMIGABLES)

    # Fallback 1: para los que no estan en NOMBRES_AMIGABLES, intentar lookup por codigo en TSV.
    # Cuando hay match en el TSV, se muestra el rulename original + " - " + descripcion del TSV
    # para que en la planilla "Buscador de contenidos" se vea tanto el codigo tecnico como el
    # nombre familiar derivado del Topic. Ejemplo:
    #   "SA03CUX01 Apertura - Vacunacion"
    with step("    Cargando descripciones desde TSV (rules-*.tsv)"):
        descripciones_tsv = cargar_descripciones_tsv()
    if descripciones_tsv:
        mask_sin_amigable = df_agrupado['Nombre Amigable'].isna()
        if mask_sin_amigable.any():
            def _resolver_tsv(rulename):
                desc = aplicar_descripcion_automatica(rulename, descripciones_tsv)
                if desc:
                    return "{} - {}".format(rulename, desc)
                return None
            df_agrupado.loc[mask_sin_amigable, 'Nombre Amigable'] = (
                df_agrupado.loc[mask_sin_amigable, 'Rulename'].apply(_resolver_tsv)
            )

    # Fallback 2: NUEVA logica - lookup por columna 'Name' del TSV.
    # Si el rulename existe en la columna Name del TSV mas reciente, usar la
    # columna 'Topic' como nombre amigable.
    #   - Si Topic == Topic path  ->  "Topic [rulename]"  (Topic es muy generico)
    #   - Si Topic != Topic path  ->  usar Topic tal cual
    # Sino, cae al fallback 3 (disclaimer).
    with step("    Cargando mapping Name->Topic desde TSV"):
        mapping_name = cargar_mapping_name_a_topic()
    if mapping_name:
        mask_sin_amigable = df_agrupado['Nombre Amigable'].isna()
        if mask_sin_amigable.any():
            df_agrupado.loc[mask_sin_amigable, 'Nombre Amigable'] = (
                df_agrupado.loc[mask_sin_amigable, 'Rulename'].apply(
                    lambda r: resolver_nombre_amigable_por_name(r, mapping_name)
                )
            )

    # Fallback 3: disclaimer. Si nada lo resolvio, mostrar rulename + disclaimer
    # para que sea OBVIO en el reporte que falta agregar el nombre amigable.
    mask_aun_sin = df_agrupado['Nombre Amigable'].isna()
    if mask_aun_sin.any():
        cant_faltantes = int(mask_aun_sin.sum())
        log("    [WARN] {} rulenames quedaron sin nombre amigable (marcados con disclaimer)".format(cant_faltantes))
        df_agrupado.loc[mask_aun_sin, 'Nombre Amigable'] = (
            df_agrupado.loc[mask_aun_sin, 'Rulename'].apply(
                lambda r: "{} {}".format(r, DISCLAIMER_FALTANTE)
            )
        )

    # Ordenar descendente
    df_agrupado = df_agrupado.sort_values('Suma de Sesiones', ascending=False).reset_index(drop=True)

    # ==== CAPA 3: Reportar posicion en ranking y luego excluir ====
    # Para cada rulename marcado: mostrar en que puesto habria estado y si entraba al top 10,
    # y despues quitarlo del df para que no aparezca en el output final.
    if APLICAR_EXCLUSIONES and RULENAMES_REPORTAR_Y_EXCLUIR:
        print("")
        print("    [INFO] === CAPA 3: Reporte de posicion (rulenames marcados para excluir tras calcular su rank) ===")
        indices_a_eliminar = []
        for rn in RULENAMES_REPORTAR_Y_EXCLUIR:
            match = df_agrupado[df_agrupado['Rulename'] == rn]
            if not match.empty:
                idx_orig = match.index[0]
                posicion = idx_orig + 1  # 1-indexed
                sesiones_rn = int(match.iloc[0]['Suma de Sesiones'])
                top10_flag = "SI ENTRA AL TOP 10" if posicion <= 10 else "fuera del top 10"
                print("    [INFO] '{}'".format(rn))
                print("           -> habria estado en el puesto #{} con {:,} sesiones ({})".format(
                    posicion, sesiones_rn, top10_flag
                ))
                indices_a_eliminar.append(idx_orig)
            else:
                print("    [INFO] '{}' no aparece en el dataset del periodo (no entra al ranking)".format(rn))
        if indices_a_eliminar:
            df_agrupado = df_agrupado.drop(indices_a_eliminar).reset_index(drop=True)
            print("    [INFO] Excluidos del ranking final: {}".format(len(indices_a_eliminar)))

    total_contenidos = len(df_agrupado)
    total_sesiones = df_agrupado['Suma de Sesiones'].sum()
    print("    [INFO] Total contenidos relevantes únicos: {:,}".format(total_contenidos))
    print("    [INFO] Total sesiones: {:,}".format(int(total_sesiones)))

    # Calcular %GT (porcentaje del gran total)
    if total_sesiones > 0:
        df_agrupado['% del Total'] = df_agrupado['Suma de Sesiones'] / total_sesiones
    else:
        df_agrupado['% del Total'] = 0

    # 7. Serie temporal diaria (Histórico)
    df_historico = None
    if fecha_col:
        df_historico = df_filtrado.groupby(fecha_col)[sesiones_col].sum().reset_index()
        df_historico.columns = ['Fecha', 'Sesiones diarias']
        df_historico = df_historico.sort_values('Fecha').reset_index(drop=True)
        print("    [INFO] Días en serie temporal: {}".format(len(df_historico)))

    # Mostrar Top 10 en consola (con nombres amigables)
    print("")
    print("    TOP 10 CONTENIDOS MÁS CONSULTADOS:")
    print("    " + "-" * 70)
    for idx, row in df_agrupado.head(10).iterrows():
        print("    {:2d}. {:45s} {:>10,}  ({:.2f}%)".format(
            idx + 1,
            row['Nombre Amigable'][:45],
            int(row['Suma de Sesiones']),
            row['% del Total'] * 100
        ))

    return df_agrupado, total_contenidos, df_historico

def create_detail_excel(filepath, df_contenidos, df_historico, descripcion):
    '''Crea un Excel con 2 hojas: Buscador de contenidos + Historico'''

    print("    [INFO] Creando Excel detallado...")

    wb = openpyxl.Workbook()

    # ===== HOJA 1: Buscador de contenidos =====
    ws = wb.active
    ws.title = 'Buscador de contenidos'

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True, size=11)
    card_font = Font(bold=True, size=12)
    card_value_font = Font(bold=True, size=14)

    total_contenidos = len(df_contenidos)
    total_sesiones = int(df_contenidos['Suma de Sesiones'].sum())

    # Cards en filas 1-2
    ws['A1'] = 'Cantidad de Rulenames'
    ws['A1'].font = card_font
    ws['B1'] = total_contenidos
    ws['B1'].font = card_value_font
    ws['B1'].number_format = '#,##0'

    ws['D1'] = 'Total Sesiones'
    ws['D1'].font = card_font
    ws['E1'] = total_sesiones
    ws['E1'].font = card_value_font
    ws['E1'].number_format = '#,##0'

    ws['A2'] = 'Período: {}'.format(descripcion)
    ws['A2'].font = Font(size=10, italic=True)

    # Headers de tabla (fila 4)
    ws['A4'] = 'Contenido'
    ws['B4'] = 'Suma de Sesiones'
    ws['C4'] = '% del Total'

    for col in ['A', 'B', 'C']:
        ws['{}4'.format(col)].font = header_font
        ws['{}4'.format(col)].fill = header_fill

    # Datos (todos los contenidos)
    for idx, row in df_contenidos.iterrows():
        fila = 5 + idx
        ws['A{}'.format(fila)] = row['Nombre Amigable']
        ws['B{}'.format(fila)] = int(row['Suma de Sesiones'])
        ws['B{}'.format(fila)].number_format = '#,##0'
        ws['C{}'.format(fila)] = row['% del Total']
        ws['C{}'.format(fila)].number_format = '0.00%'

    # Fila de totales
    fila_total = 5 + total_contenidos
    ws['A{}'.format(fila_total)] = 'TOTAL'
    ws['A{}'.format(fila_total)].font = bold_font
    ws['B{}'.format(fila_total)] = total_sesiones
    ws['B{}'.format(fila_total)].font = bold_font
    ws['B{}'.format(fila_total)].number_format = '#,##0'
    ws['C{}'.format(fila_total)] = 1.0
    ws['C{}'.format(fila_total)].font = bold_font
    ws['C{}'.format(fila_total)].number_format = '0.00%'

    # Ajustar anchos
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # ===== HOJA 2: Historico =====
    if df_historico is not None and len(df_historico) > 0:
        ws2 = wb.create_sheet('Historico')

        # Headers
        ws2['A1'] = 'Fecha'
        ws2['B1'] = 'Sesiones diarias'
        ws2['A1'].font = header_font
        ws2['A1'].fill = header_fill
        ws2['B1'].font = header_font
        ws2['B1'].fill = header_fill

        # Datos
        for idx, row in df_historico.iterrows():
            fila = 2 + idx
            ws2['A{}'.format(fila)] = row['Fecha']
            ws2['A{}'.format(fila)].number_format = 'DD/MM/YYYY'
            ws2['B{}'.format(fila)] = int(row['Sesiones diarias'])
            ws2['B{}'.format(fila)].number_format = '#,##0'

        # Fila de total
        fila_total_hist = 2 + len(df_historico)
        ws2['A{}'.format(fila_total_hist)] = 'TOTAL'
        ws2['A{}'.format(fila_total_hist)].font = bold_font
        ws2['B{}'.format(fila_total_hist)] = int(df_historico['Sesiones diarias'].sum())
        ws2['B{}'.format(fila_total_hist)].font = bold_font
        ws2['B{}'.format(fila_total_hist)].number_format = '#,##0'

        # Ajustar anchos
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 18

    wb.save(filepath)
    print("    [OK] Excel detallado creado ({} contenidos, {} hojas)".format(
        total_contenidos,
        len(wb.sheetnames)
    ))

def format_top10_text(df_contenidos):
    '''
    Genera el texto formateado del Top 10 para la celda D11 del dashboard.
    Formato: "1- Nombre contenido: (123.456)\n2- Otro contenido: (78.901)\n..."
    '''
    top10 = df_contenidos.head(10)
    lines = []
    for idx, row in top10.iterrows():
        sesiones = int(row['Suma de Sesiones'])
        # Formato con punto como separador de miles (estilo argentino)
        sesiones_fmt = '{:,}'.format(sesiones).replace(',', '.')
        lines.append('{}- {}: ({})'.format(idx + 1, row['Nombre Amigable'], sesiones_fmt))
    return '\n'.join(lines)

def create_dashboard(filepath, valor_d11, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea el Excel Dashboard con la estructura estándar.
    Llena SOLO la celda D11 (Contenidos más consultados - Top 10 formateado)
    '''

    print("    [INFO] Creando Dashboard...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    header_font = Font(bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determinar el header de fecha
    if modo == 'mes':
        header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{}-{}'.format(
            fecha_inicio_obj.strftime('%d/%m'),
            fecha_fin_obj.strftime('%d/%m/%y')
        )

    # FILA 1: Encabezados
    ws['B1'] = 'Indicador'
    ws['C1'] = 'Descripción/Detalle'
    ws['D1'] = header_fecha
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font

    # Filas 2-17: Indicadores
    ws['B2'] = 'Conversaciones'
    ws['C2'] = 'Q Conversaciones'

    ws['B3'] = 'Usuarios'
    ws['C3'] = 'Q Usuarios únicos'

    ws['B4'] = 'Sesiones abiertas por Pushes'
    ws['C4'] = 'Q Sesiones que se abrieron con una Push'

    ws['B5'] = 'Sesiones Alcanzadas por Pushes'
    ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'

    ws['B6'] = 'Mensajes Pushes Enviados'
    ws['C6'] = 'Q de mensajes enviados bajo el formato push'

    ws['B7'] = 'Contenidos en Botmaker'
    ws['C7'] = 'Contenidos prendidos en botmaker'

    ws['B8'] = 'Contenidos Prendidos para el USUARIO'
    ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes)'

    ws['B9'] = 'Interacciones'
    ws['C9'] = 'Q Interacciones'

    ws['B10'] = 'Trámites, solicitudes y turnos'
    ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'

    ws['B11'] = 'Contenidos más consultados'
    ws['C11'] = 'Q Contenidos con más interacciones en el mes'
    if valor_d11 is not None:
        ws['D11'] = valor_d11
        ws['D11'].alignment = Alignment(wrap_text=True, vertical='top')

    ws['B12'] = 'Derivaciones'
    ws['C12'] = 'Q Derivaciones'

    ws['B13'] = 'No entendimiento'
    ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'

    ws['B14'] = 'Tasa de Efectividad'
    ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'

    ws['B15'] = 'CES (Customer Effort Score)'
    ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar'

    ws['B16'] = 'Satisfacción (CSAT)'
    ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'

    ws['B17'] = 'Uptime servidor'
    ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'

    # Aplicar bordes a todas las celdas con datos
    for row in range(1, 18):
        for col in ['B', 'C', 'D']:
            ws['{}{}'.format(col, row)].border = border

    # Ajustar anchos
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 50

    wb.save(filepath)

    if valor_d11 is not None:
        # Mostrar preview del Top 10
        preview = valor_d11.split('\n')[0]
        print("    [OK] Dashboard creado (D11 = Top 10, ej: {})".format(preview))
    else:
        print("    [OK] Dashboard creado (D11 vacío)")

def execute_query_and_save():
    '''Función principal que ejecuta la query y guarda los resultados'''

    print("Leyendo configuracion de fechas...")
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = read_date_config(CONFIG['config_file'])

    if modo is None:
        print("[ERROR] No se pudo obtener la configuracion. Abortando.")
        return None

    print("[OK] Configuracion leida correctamente")
    print("    Modo: {}".format(modo.upper()))
    print("    Periodo: {}".format(descripcion))
    print("    Fecha inicio: {}".format(fecha_inicio))
    print("    Fecha fin: {}".format(fecha_fin))

    query = build_query()

    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))

    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))

    try:
        session = boto3.Session(region_name=CONFIG['region'])

        log("")
        log("ATENCION: la query escanea TODA la vista boti_vw_buscador_rulename")
        log("          (sin filtro de fecha). Puede tardar VARIOS MINUTOS - es normal.")
        log("          El proceso NO esta colgado; estoy esperando que Athena responda.")

        with step("Query Athena (boti_vw_buscador_rulename)"):
            try:
                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    workgroup=CONFIG['workgroup'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            except Exception as e:
                if 'workgroup' in str(e).lower():
                    log("[ADVERTENCIA] Error con workgroup; reintentando sin workgroup...")
                    df = wr.athena.read_sql_query(
                        sql=query,
                        database=CONFIG['database'],
                        boto3_session=session,
                        ctas_approach=False,
                        unload_approach=False
                    )
                else:
                    raise e

        log("[OK] Consulta ejecutada - {:,} filas descargadas".format(len(df)))

        # Verificar resultados
        if len(df) == 0:
            print("[ADVERTENCIA] La query no retornó resultados")
            return None

        print("")
        print("=" * 60)
        print("RESULTADO - {}".format(descripcion.upper()))
        print("=" * 60)
        print("Total de filas descargadas: {:,}".format(len(df)))

        # Generar nombres de archivo
        filename_csv, filename_detalle, filename_dashboard = generate_filename(modo, mes, anio, fecha_inicio, fecha_fin)
        output_folder = CONFIG['output_folder']

        os.makedirs(output_folder, exist_ok=True)

        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_detalle = os.path.join(output_folder, filename_detalle)
        local_path_dashboard = os.path.join(output_folder, filename_dashboard)

        # Guardar CSV con datos crudos
        log("")
        with step("Guardando CSV crudo ({:,} filas)".format(len(df))):
            df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        log("    [OK] CSV guardado: {}".format(filename_csv))

        # Procesar datos
        log("")
        with step("Procesando contenidos (filtrado + agrupacion + TSV)"):
            df_contenidos, total_contenidos, df_historico = procesar_contenidos(df, fecha_inicio, fecha_fin)

        if df_contenidos is None:
            log("[ERROR] No se pudo procesar los datos")
            return None

        # Generar Excel detallado (2 hojas)
        log("")
        with step("Generando Excel detallado ({} filas)".format(len(df_contenidos))):
            create_detail_excel(local_path_detalle, df_contenidos, df_historico, descripcion)

        # Generar Dashboard
        log("")
        with step("Generando Dashboard"):
            # D11 = Top 10 contenidos formateado como texto multilínea
            top10_text = format_top10_text(df_contenidos)
            create_dashboard(local_path_dashboard, top10_text, modo, mes, anio, fecha_inicio, fecha_fin)

        print("")
        print("=" * 60)
        print("ARCHIVOS GENERADOS:")
        print("=" * 60)
        print("    [CSV] {}".format(filename_csv))
        print("          - Datos crudos de la vista Athena")
        print("")
        print("    [EXCEL DETALLE] {}".format(filename_detalle))
        print("          - Hoja 'Buscador de contenidos': {} contenidos con %GT".format(total_contenidos))
        if df_historico is not None:
            print("          - Hoja 'Historico': {} días de serie temporal".format(len(df_historico)))
        print("          - Total sesiones: {:,}".format(int(df_contenidos['Suma de Sesiones'].sum())))
        print("")
        print("    [DASHBOARD] {}".format(filename_dashboard))
        print("          - Celda D11 = Top 10 contenidos más consultados")

        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)

        return df

    except Exception as e:
        print("")
        print("[ERROR] {}".format(str(e)))
        import traceback
        traceback.print_exc()
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    # Forzar line-buffering en stdout/stderr para que los prints salgan en
    # tiempo real (sino Windows puede bufferear y parece que el script se trabo)
    try:
        sys.stdout.reconfigure(line_buffering=True)
        sys.stderr.reconfigure(line_buffering=True)
    except AttributeError:
        # Python < 3.7 - dejar como esta
        pass

    print("")
    print("=" * 60)
    print("  IMPORTANTE: LOGIN AWS REQUERIDO")
    print("=" * 60)
    print("  Antes de continuar, asegurate de estar logueado en AWS.")
    print("  Si tu token expiró, ejecutá en otra terminal:")
    print("")
    print("      aws-azure-login --profile default --mode=gui")
    print("")
    print("=" * 60)
    try:
        input("  Presioná ENTER para continuar (o Ctrl+C para cancelar)...")
    except KeyboardInterrupt:
        print("")
        print("  Cancelado por el usuario.")
        raise SystemExit(0)
    print("")

    print("")
    print("=" * 60)
    print("SCRIPT: CONTENIDOS CONSULTADOS - QUERY ATHENA")
    print("=" * 60)
    print("Rol requerido: PIBADataScientist")
    print("Salida: CSV + Excel Detalle (2 hojas) + Dashboard (celda D11)")
    print("Query: SELECT * FROM boti_vw_buscador_rulename")
    print("")
    print("LÓGICA (filtrado en 2 capas como el PBI):")
    print("  1. Descargar vista completa de Athena")
    print("  2. Filtrar por rango de fechas del período")
    if APLICAR_EXCLUSIONES:
        print("  3. CAPA 1: Excluir por {} patrones dinámicos (CONTAINS)".format(len(PATRONES_EXCLUIR)))
        print("  4. CAPA 2: Excluir por {} nombres exactos (lista manual)".format(len(CONTENIDOS_EXCLUIR_MANUAL)))
    else:
        print("  3-4. Exclusiones DESACTIVADAS")
    print("  5. Extraer prefijo y agrupar por (prefijo, fecha)")
    print("  6. Quedarse con rulename de más sesiones por (prefijo, fecha)")
    print("  7. Sumar sesiones por rulename y calcular % del total")
    print("  8. Generar serie temporal diaria (Histórico)")
    print("")
    print("MODOS:")
    print("  [1] MES COMPLETO: MES + AÑO")
    print("  [2] RANGO PERSONALIZADO: FECHA_INICIO + FECHA_FIN")
    print("=" * 60)
    print("")

    result = execute_query_and_save()

    if result is not None:
        print("")
        print("Para procesar otro periodo, edita config_fechas.txt")
    else:
        print("")
        print("La ejecucion fallo. Revisa los errores arriba.")
