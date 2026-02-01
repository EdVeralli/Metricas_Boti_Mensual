# -*- coding: utf-8 -*-
'''
Script para ejecutar query de Feedback - Efectividad en Athena
Cuenta sesiones por rule_name que contengan 'CXF'
Guarda resultado en CSV y genera Excel con análisis detallado y cálculo de Efectividad
Lee configuracion de fechas desde archivo config_fechas.txt

MODIFICACIÓN: Incluye SUBTOTALES después de cada categoría en el Excel

MODOS SOPORTADOS:
1. MES COMPLETO: Especificar MES y AÑO (comportamiento original)
2. RANGO PERSONALIZADO: Especificar FECHA_INICIO y FECHA_FIN

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti
'''
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': '../config_fechas.txt'  # Config centralizado en raiz del proyecto
}

# Reglas específicas para el cálculo de Efectividad
REGLAS_EFECTIVIDAD = {
    'integraciones_no': 'CXF01CUX01 No Integraciones',
    'integraciones_si': 'CXF01CUX01 Sí Integraciones',
    'estaticos_no': 'CXF01CUX02 No Estáticos',
    'estaticos_si': 'CXF01CUX02 Sí Estáticos',
    'pushes_no': 'CXF01CUX03 No Pushes',
    'pushes_si': 'CXF01CUX03 Sí Pushes',
    'cats_no': 'CXF01CUX04 No CATs',
    'cats_si': 'CXF01CUX04 Sí CATs'
}

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    '''
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)
    
    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    '''
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            print("    Creando archivo de ejemplo...")
            
            with open(config_file, 'w', encoding='utf-8') as f:
                f.write("# Configuracion de fechas para el reporte\n")
                f.write("# MODO 1: Mes completo\n")
                f.write("MES=10\n")
                f.write("AÑO=2025\n\n")
                f.write("# MODO 2: Rango personalizado\n")
                f.write("# FECHA_INICIO=2025-10-01\n")
                f.write("# FECHA_FIN=2025-10-15\n")
            
            print("    Archivo creado: {}".format(config_file))
            return 'mes', None, None, 10, 2025, "octubre 2025"
        
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
                
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()
        
        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
                
                if fecha_inicio > fecha_fin:
                    print("[ERROR] FECHA_INICIO no puede ser posterior a FECHA_FIN")
                    return None, None, None, None, None, None
                
                descripcion = "{} al {}".format(
                    fecha_inicio.strftime('%d/%m/%Y'),
                    fecha_fin.strftime('%d/%m/%Y')
                )
                
                print("[INFO] Modo: RANGO PERSONALIZADO")
                return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion
                
            except ValueError as e:
                print("[ERROR] Formato de fecha invalido. Use YYYY-MM-DD (ej: 2025-10-01)")
                return None, None, None, None, None, None
        
        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            if mes < 1 or mes > 12:
                print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
                return None, None, None, None, None, None
            
            if anio < 2020 or anio > 2030:
                print("[ADVERTENCIA] Año inusual: {}".format(anio))
            
            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
            
            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)
            
            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion
        
        print("[ERROR] El archivo {} no contiene configuracion valida".format(config_file))
        return None, None, None, None, None, None
        
    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    '''Retorna la abreviatura del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_query(fecha_inicio, fecha_fin):
    '''Construye la query de Feedback - Efectividad con el rango de fechas especificado'''
    
    query = """SELECT 
rule_name,
count(distinct(session_id)) as Cant_Sesiones 
FROM "caba-piba-consume-zone-db"."boti_message_metrics_2"
WHERE CAST(session_creation_time AS DATE) BETWEEN date '{fecha_inicio}' and date '{fecha_fin}'
and rule_name like ('%CXF%')
group by rule_name""".format(fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
    
    return query

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    '''Genera los nombres de archivos basados en el modo y las fechas'''
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "feedback_efectividad_{0}_{1}.csv".format(mes_nombre, anio)
        filename_excel_detalle = "feedback_efectividad_detalle_{0}_{1}.xlsx".format(mes_nombre, anio)
        filename_dashboard = "feedback_efectividad_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "feedback_efectividad_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_excel_detalle = "feedback_efectividad_detalle_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_dashboard = "feedback_efectividad_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
    
    return filename_csv, filename_excel_detalle, filename_dashboard

def extraer_valores_efectividad(df):
    '''
    Extrae los valores de las 6 reglas específicas para el cálculo de Efectividad
    Retorna un diccionario con los valores encontrados
    '''
    valores = {}
    
    for key, rule_name in REGLAS_EFECTIVIDAD.items():
        # Buscar la regla en el DataFrame
        fila = df[df['rule_name'] == rule_name]
        
        if len(fila) > 0:
            valores[key] = int(fila['cant_sesiones'].iloc[0])
        else:
            print("    [ADVERTENCIA] No se encontró la regla: {}".format(rule_name))
            valores[key] = 0
    
    return valores

def calcular_efectividad(valores):
    '''
    Calcula las métricas de efectividad basadas en los valores extraídos
    Incluye: Integraciones, Estáticos, Pushes y CATs
    '''
    # Respuestas Positivas (Sí)
    positivas = (
        valores['integraciones_si'] +
        valores['estaticos_si'] +
        valores['pushes_si'] +
        valores['cats_si']
    )
    
    # Respuestas Negativas (No)
    negativas = (
        valores['integraciones_no'] +
        valores['estaticos_no'] +
        valores['pushes_no'] +
        valores['cats_no']
    )
    
    # Total
    total = positivas + negativas
    
    # Efectividad (%)
    if total > 0:
        efectividad = positivas / total
    else:
        efectividad = 0
    
    return {
        'positivas': positivas,
        'negativas': negativas,
        'total': total,
        'efectividad': efectividad
    }

def create_excel_with_efectividad(filepath, df, valores, calculos, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea 3 Excel separados: base cruda, detalle con efectividad, y resumen
    ✨ MODIFICADO: Genera archivos separados en lugar de hojas múltiples
    '''
    
    print("    [INFO] Creando 3 archivos Excel separados...")
    
    # Generar nombres de archivos
    base_name = filepath.replace('.xlsx', '').replace('_detalle', '')
    filepath_base_cruda = base_name + '_base_cruda.xlsx'
    filepath_efectividad = base_name + '_efectividad.xlsx'
    filepath_resumen = base_name + '_resumen.xlsx'
    
    # Determinar el header de fecha
    if modo == 'mes':
        header_fecha = '{} {}'.format(get_month_name(mes), anio)
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{} al {}'.format(
            fecha_inicio_obj.strftime('%d/%m/%Y'),
            fecha_fin_obj.strftime('%d/%m/%Y')
        )
    
    # Estilos comunes
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    title_font = Font(bold=True, size=12)
    category_font = Font(bold=True, size=11)
    normal_font = Font(size=11)
    result_font = Font(bold=True, size=11)
    efectividad_font = Font(bold=True, size=12, color="FFFFFF")
    efectividad_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    subtotal_font = Font(bold=True, size=10, italic=True)
    subtotal_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    total_font = Font(bold=True, size=12, color="FFFFFF")
    total_no_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    total_si_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    
    # ==================== ARCHIVO 1: BASE CRUDA ====================
    print("    [1/3] Generando: base cruda...")
    wb1 = openpyxl.Workbook()
    ws_base = wb1.active
    ws_base.title = 'base cruda'
    
    # Headers
    ws_base['A1'] = 'rule_name'
    ws_base['B1'] = 'Cant_Sesiones'
    ws_base['A1'].font = header_font
    ws_base['B1'].font = header_font
    ws_base['A1'].fill = header_fill
    ws_base['B1'].fill = header_fill
    
    # Datos - ordenados por cantidad de sesiones
    df_sorted = df.sort_values('cant_sesiones', ascending=False)
    for idx, row in enumerate(df_sorted.iterrows(), start=2):
        ws_base['A{}'.format(idx)] = row[1]['rule_name']
        ws_base['B{}'.format(idx)] = int(row[1]['cant_sesiones'])
    
    # Ajustar anchos
    ws_base.column_dimensions['A'].width = 50
    ws_base.column_dimensions['B'].width = 15
    
    # Guardar
    wb1.save(filepath_base_cruda)
    wb1.close()
    
    # ==================== ARCHIVO 2: EFECTIVIDAD ====================
    print("    [2/3] Generando: efectividad...")
    wb2 = openpyxl.Workbook()
    ws = wb2.active
    ws.title = 'Efectividad 2025'
    
    # FILA 1: Header
    ws['A1'] = header_fecha
    ws['A1'].font = title_font
    
    row = 2
    
    # ============== INTEGRACIONES ==============
    ws['A{}'.format(row)] = 'Integraciones'
    ws['A{}'.format(row)].font = category_font
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['integraciones_no']
    ws['B{}'.format(row)] = valores['integraciones_no']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['integraciones_si']
    ws['B{}'.format(row)] = valores['integraciones_si']
    row += 1
    
    # SUBTOTAL INTEGRACIONES
    ws['A{}'.format(row)] = '  Subtotal Integraciones'
    ws['B{}'.format(row)] = valores['integraciones_no'] + valores['integraciones_si']
    ws['A{}'.format(row)].font = subtotal_font
    ws['B{}'.format(row)].font = Font(bold=True, size=10)
    ws['A{}'.format(row)].fill = subtotal_fill
    ws['B{}'.format(row)].fill = subtotal_fill
    row += 1
    
    # Línea en blanco
    row += 1
    
    # ============== ESTÁTICOS ==============
    ws['A{}'.format(row)] = 'Estáticos'
    ws['A{}'.format(row)].font = category_font
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['estaticos_no']
    ws['B{}'.format(row)] = valores['estaticos_no']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['estaticos_si']
    ws['B{}'.format(row)] = valores['estaticos_si']
    row += 1
    
    # SUBTOTAL ESTÁTICOS
    ws['A{}'.format(row)] = '  Subtotal Estáticos'
    ws['B{}'.format(row)] = valores['estaticos_no'] + valores['estaticos_si']
    ws['A{}'.format(row)].font = subtotal_font
    ws['B{}'.format(row)].font = Font(bold=True, size=10)
    ws['A{}'.format(row)].fill = subtotal_fill
    ws['B{}'.format(row)].fill = subtotal_fill
    row += 1
    
    # Línea en blanco
    row += 1
    
    # ============== PUSHES ==============
    ws['A{}'.format(row)] = 'Pushes'
    ws['A{}'.format(row)].font = category_font
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['pushes_no']
    ws['B{}'.format(row)] = valores['pushes_no']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['pushes_si']
    ws['B{}'.format(row)] = valores['pushes_si']
    row += 1
    
    # SUBTOTAL PUSHES
    ws['A{}'.format(row)] = '  Subtotal Pushes'
    ws['B{}'.format(row)] = valores['pushes_no'] + valores['pushes_si']
    ws['A{}'.format(row)].font = subtotal_font
    ws['B{}'.format(row)].font = Font(bold=True, size=10)
    ws['A{}'.format(row)].fill = subtotal_fill
    ws['B{}'.format(row)].fill = subtotal_fill
    row += 1
    
    # Línea en blanco
    row += 1
    
    # ============== CATS ==============
    ws['A{}'.format(row)] = 'CATs'
    ws['A{}'.format(row)].font = category_font
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['cats_no']
    ws['B{}'.format(row)] = valores['cats_no']
    row += 1
    
    ws['A{}'.format(row)] = REGLAS_EFECTIVIDAD['cats_si']
    ws['B{}'.format(row)] = valores['cats_si']
    row += 1
    
    # SUBTOTAL CATS
    ws['A{}'.format(row)] = '  Subtotal CATs'
    ws['B{}'.format(row)] = valores['cats_no'] + valores['cats_si']
    ws['A{}'.format(row)].font = subtotal_font
    ws['B{}'.format(row)].font = Font(bold=True, size=10)
    ws['A{}'.format(row)].fill = subtotal_fill
    ws['B{}'.format(row)].fill = subtotal_fill
    row += 1
    
    # Doble línea en blanco antes de totales
    row += 2
    
    # ============== TOTALES GENERALES ==============
    # SUMA TOTAL DE "NO"
    ws['A{}'.format(row)] = 'SUMA TOTAL DE "NO"'
    ws['B{}'.format(row)] = valores['integraciones_no'] + valores['estaticos_no'] + valores['pushes_no'] + valores['cats_no']
    ws['A{}'.format(row)].font = total_font
    ws['B{}'.format(row)].font = total_font
    ws['A{}'.format(row)].fill = total_no_fill
    ws['B{}'.format(row)].fill = total_no_fill
    ws['A{}'.format(row)].alignment = Alignment(horizontal='left')
    ws['B{}'.format(row)].alignment = Alignment(horizontal='center')
    row += 1
    
    # SUMA TOTAL DE "SÍ"
    ws['A{}'.format(row)] = 'SUMA TOTAL DE "SÍ"'
    ws['B{}'.format(row)] = valores['integraciones_si'] + valores['estaticos_si'] + valores['pushes_si'] + valores['cats_si']
    ws['A{}'.format(row)].font = total_font
    ws['B{}'.format(row)].font = total_font
    ws['A{}'.format(row)].fill = total_si_fill
    ws['B{}'.format(row)].fill = total_si_fill
    ws['A{}'.format(row)].alignment = Alignment(horizontal='left')
    ws['B{}'.format(row)].alignment = Alignment(horizontal='center')
    row += 1
    
    # Doble línea en blanco antes del resumen
    row += 2
    
    # RESUMEN
    ws['A{}'.format(row)] = 'Respuestas'
    ws['C{}'.format(row)] = 'Efectividad'
    ws['A{}'.format(row)].font = category_font
    ws['C{}'.format(row)].font = category_font
    row += 1
    
    # Positivas
    ws['A{}'.format(row)] = 'Positivas'
    ws['B{}'.format(row)] = calculos['positivas']
    ws['C{}'.format(row)] = calculos['efectividad']
    ws['C{}'.format(row)].number_format = '0.00%'
    row += 1
    
    # Negativas
    ws['A{}'.format(row)] = 'Negativo'
    ws['B{}'.format(row)] = calculos['negativas']
    ws['C{}'.format(row)] = 1 - calculos['efectividad']
    ws['C{}'.format(row)].number_format = '0.00%'
    row += 1
    
    # Total
    ws['B{}'.format(row)] = calculos['total']
    ws['B{}'.format(row)].font = result_font
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    # Guardar
    wb2.save(filepath_efectividad)
    wb2.close()
    
    # ==================== ARCHIVO 3: RESUMEN EJECUTIVO ====================
    print("    [3/3] Generando: resumen...")
    wb3 = openpyxl.Workbook()
    ws_resumen = wb3.active
    ws_resumen.title = 'Resumen'
    
    # Título
    ws_resumen['A1'] = 'FEEDBACK - EFECTIVIDAD'
    ws_resumen['A1'].font = Font(bold=True, size=14)
    ws_resumen['A2'] = 'Período: {}'.format(header_fecha)
    ws_resumen['A2'].font = Font(size=11)
    
    # Métrica principal
    ws_resumen['A4'] = '🎯 EFECTIVIDAD GENERAL'
    ws_resumen['A4'].font = Font(bold=True, size=12)
    
    ws_resumen['A5'] = 'Efectividad:'
    ws_resumen['B5'] = calculos['efectividad']
    ws_resumen['B5'].number_format = '0.00%'
    ws_resumen['B5'].font = efectividad_font
    ws_resumen['B5'].fill = efectividad_fill
    ws_resumen['B5'].alignment = Alignment(horizontal='center')
    
    # Desglose
    ws_resumen['A7'] = 'DESGLOSE'
    ws_resumen['A7'].font = Font(bold=True, size=11)
    
    ws_resumen['A8'] = 'Respuestas Positivas (Sí):'
    ws_resumen['B8'] = calculos['positivas']
    
    ws_resumen['A9'] = 'Respuestas Negativas (No):'
    ws_resumen['B9'] = calculos['negativas']
    
    ws_resumen['A10'] = 'Total de Respuestas:'
    ws_resumen['B10'] = calculos['total']
    ws_resumen['B10'].font = Font(bold=True)
    
    # Por categoría
    ws_resumen['A12'] = 'DETALLE POR CATEGORÍA'
    ws_resumen['A12'].font = Font(bold=True, size=11)
    
    ws_resumen['A13'] = 'Integraciones:'
    ws_resumen['B13'] = 'No: {:,} | Sí: {:,}'.format(valores['integraciones_no'], valores['integraciones_si'])
    
    ws_resumen['A14'] = 'Estáticos:'
    ws_resumen['B14'] = 'No: {:,} | Sí: {:,}'.format(valores['estaticos_no'], valores['estaticos_si'])
    
    ws_resumen['A15'] = 'Pushes:'
    ws_resumen['B15'] = 'No: {:,} | Sí: {:,}'.format(valores['pushes_no'], valores['pushes_si'])
    
    ws_resumen['A16'] = 'CATs:'
    ws_resumen['B16'] = 'No: {:,} | Sí: {:,}'.format(valores['cats_no'], valores['cats_si'])
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 25
    
    # Guardar
    wb3.save(filepath_resumen)
    wb3.close()
    
    print("    [OK] 3 archivos Excel creados:")
    print("        [1] {}".format(filepath_base_cruda.split('/')[-1]))
    print("        [2] {}".format(filepath_efectividad.split('/')[-1]))
    print("        [3] {}".format(filepath_resumen.split('/')[-1]))
    
    return filepath_base_cruda, filepath_efectividad, filepath_resumen



def create_or_update_dashboard_master(filepath, efectividad_valor, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea el Excel Dashboard Master con el template estándar
    Llena SOLO la celda D14 (Tasa de Efectividad)
    Usa el mismo template que CES y CSAT
    '''

    print("    [INFO] Creando Dashboard Master...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Dashboard'

    header_font = Font(bold=True)

    # Determinar el header de fecha
    if modo == 'mes':
        header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{}-{}'.format(
            fecha_inicio_obj.strftime('%d/%m'),
            fecha_fin_obj.strftime('%d/%m/%y')
        )

    # FILA 1: Encabezados
    ws['B1'] = 'Indicador'
    ws['C1'] = 'Descripción/Detalle'
    ws['D1'] = header_fecha
    ws['B1'].font = header_font
    ws['C1'].font = header_font
    ws['D1'].font = header_font

    # FILA 2-13: Otros indicadores (vacíos)
    ws['B2'] = 'Conversaciones'
    ws['C2'] = 'Q Conversaciones'

    ws['B3'] = 'Usuarios'
    ws['C3'] = 'Q Usuarios únicos'

    ws['B4'] = 'Sesiones abiertas por Pushes'
    ws['C4'] = 'Q Sesiones que se abrieron con una Push'

    ws['B5'] = 'Sesiones Alcanzadas por Pushes'
    ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'

    ws['B6'] = 'Mensajes Pushes Enviados'
    ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'

    ws['B7'] = 'Contenidos en Botmaker'
    ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'

    ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
    ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'

    ws['B9'] = 'Interacciones'
    ws['C9'] = 'Q Interacciones'

    ws['B10'] = 'Trámites, solicitudes y turnos'
    ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'

    ws['B11'] = 'contenidos mas consultados'
    ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'

    ws['B12'] = 'Derivaciones'
    ws['C12'] = 'Q Derivaciones'

    ws['B13'] = 'No entendimiento'
    ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'

    # FILA 14: Tasa de Efectividad - AQUI VA NUESTRO VALOR
    ws['B14'] = 'Tasa de Efectividad'
    ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'
    ws['D14'] = efectividad_valor
    ws['D14'].number_format = '0.00%'

    # FILA 15-17: Otros indicadores (vacíos)
    ws['B15'] = 'CES (Customer Effort Score)'
    ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'

    ws['B16'] = 'Satisfacción (CSAT)'
    ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'

    ws['B17'] = 'Uptime servidor'
    ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'

    # Ajustar anchos
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 15

    wb.save(filepath)
    print("    [OK] Dashboard Master creado (D14 = {:.2f}%)".format(efectividad_valor * 100))

def execute_query_and_save():
    '''Función principal que ejecuta la query y guarda los resultados'''
    
    print("Leyendo configuracion de fechas...")
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = read_date_config(CONFIG['config_file'])
    
    if modo is None:
        print("[ERROR] No se pudo obtener la configuracion. Abortando.")
        return None
    
    print("[OK] Configuracion leida correctamente")
    print("    Modo: {}".format(modo.upper()))
    print("    Periodo: {}".format(descripcion))
    print("    Fecha inicio: {}".format(fecha_inicio))
    print("    Fecha fin: {}".format(fecha_fin))
    
    query = build_query(fecha_inicio, fecha_fin)
    
    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))
    
    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))
    
    try:
        session = boto3.Session(region_name=CONFIG['region'])
        
        print("")
        print("Ejecutando consulta...")
        
        try:
            df = wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                workgroup=CONFIG['workgroup'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False
            )
        except Exception as e:
            if 'workgroup' in str(e).lower():
                print("[ADVERTENCIA] Intentando sin workgroup...")
                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            else:
                raise e
        
        print("[OK] Consulta ejecutada exitosamente!")
        
        # Verificar resultados
        if len(df) == 0:
            print("[ADVERTENCIA] La query no retornó resultados")
            print("    Esto puede significar que no hay reglas CXF en el período especificado")
            return None
        
        # Normalizar nombres de columnas
        df.columns = df.columns.str.lower()
        
        print("")
        print("=" * 60)
        print("RESULTADO - {}".format(descripcion.upper()))
        print("=" * 60)
        print("Total de reglas CXF encontradas: {}".format(len(df)))
        print("Total de sesiones: {:,}".format(df['cant_sesiones'].sum()))
        
        # Extraer valores para efectividad
        print("")
        print("Extrayendo reglas específicas para cálculo de Efectividad...")
        valores = extraer_valores_efectividad(df)
        
        print("")
        print("REGLAS PARA EFECTIVIDAD:")
        print("-" * 60)
        print("  Integraciones No:  {:,}".format(valores['integraciones_no']))
        print("  Integraciones Sí:  {:,}".format(valores['integraciones_si']))
        print("  Estáticos No:      {:,}".format(valores['estaticos_no']))
        print("  Estáticos Sí:      {:,}".format(valores['estaticos_si']))
        print("  Pushes No:         {:,}".format(valores['pushes_no']))
        print("  Pushes Sí:         {:,}".format(valores['pushes_si']))
        print("  CATs No:           {:,}".format(valores['cats_no']))
        print("  CATs Sí:           {:,}".format(valores['cats_si']))
        
        # Calcular efectividad
        calculos = calcular_efectividad(valores)
        
        print("")
        print("=" * 60)
        print("CÁLCULO DE EFECTIVIDAD")
        print("=" * 60)
        print("  Respuestas Positivas (Sí):  {:,}".format(calculos['positivas']))
        print("  Respuestas Negativas (No):  {:,}".format(calculos['negativas']))
        print("  Total de Respuestas:        {:,}".format(calculos['total']))
        print("")
        print("  🎯 EFECTIVIDAD:             {:.2f}%".format(calculos['efectividad'] * 100))
        print("=" * 60)
        
        filename_csv, filename_excel_detalle, filename_dashboard = generate_filename(modo, mes, anio, fecha_inicio, fecha_fin)
        output_folder = CONFIG['output_folder']
        
        os.makedirs(output_folder, exist_ok=True)
        
        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel_detalle = os.path.join(output_folder, filename_excel_detalle)
        local_path_dashboard = os.path.join(output_folder, filename_dashboard)
        
        print("")
        print("Guardando CSV...")
        df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        
        print("Generando archivos Excel...")
        filepath_base_cruda, filepath_efectividad, filepath_resumen = create_excel_with_efectividad(
            local_path_excel_detalle, df, valores, calculos, modo, mes, anio, fecha_inicio, fecha_fin
        )
        
        print("Generando Dashboard Master...")
        create_or_update_dashboard_master(local_path_dashboard, calculos['efectividad'], modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("")
        print("ARCHIVOS GENERADOS:")
        print("    [1] CSV: {}".format(filename_csv))
        print("")
        print("    [2] EXCEL BASE CRUDA: {}".format(os.path.basename(filepath_base_cruda)))
        print("        - {} reglas ordenadas por sesiones".format(len(df)))
        print("")
        print("    [3] EXCEL EFECTIVIDAD: {}".format(os.path.basename(filepath_efectividad)))
        print("        - Cálculos detallados CON SUBTOTALES")
        print("        - SUMA TOTAL DE 'NO' y 'SÍ'")
        print("        - Efectividad: {:.2f}%".format(calculos['efectividad'] * 100))
        print("")
        print("    [4] EXCEL RESUMEN: {}".format(os.path.basename(filepath_resumen)))
        print("        - Métricas principales")
        print("        - Desglose por categoría")
        print("")
        print("    [5] DASHBOARD: {}".format(filename_dashboard))
        print("        - Celda D14 = {:.2f}% (Tasa de Efectividad)".format(calculos['efectividad'] * 100))
        
        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)
        
        return df
        
    except Exception as e:
        print("")
        print("[ERROR] {}".format(str(e)))
        import traceback
        traceback.print_exc()
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("SCRIPT: FEEDBACK - EFECTIVIDAD - QUERY ATHENA")
    print("VERSIÓN CON ARCHIVOS EXCEL SEPARADOS")
    print("=" * 60)
    print("Rol requerido: PIBAConsumeBoti")
    print("Salida: CSV + 3 Excel + Dashboard Master")
    print("Query: Reglas CXF con conteo de sesiones")
    print("")
    print("ARCHIVOS GENERADOS:")
    print("  [1] CSV: feedback_efectividad_[fecha].csv")
    print("  [2] Excel BASE CRUDA: ...base_cruda.xlsx")
    print("      - Todas las reglas CXF ordenadas")
    print("  [3] Excel EFECTIVIDAD: ...efectividad.xlsx")
    print("      - Cálculos detallados con SUBTOTALES")
    print("      - SUMA TOTAL DE 'NO' y 'SÍ' ✨")
    print("  [4] Excel RESUMEN: ...resumen.xlsx")
    print("      - Métricas principales ejecutivas")
    print("  [5] Dashboard: ...dashboard.xlsx (celda D14)")
    print("")
    print("CÁLCULOS:")
    print("  - Extrae 8 reglas específicas (Integraciones, Estáticos, Pushes, CATs)")
    print("  - Calcula Positivas vs Negativas")
    print("  - Calcula Efectividad (%)")
    print("")
    print("MODOS:")
    print("  [1] MES COMPLETO: MES + AÑO")
    print("  [2] RANGO PERSONALIZADO: FECHA_INICIO + FECHA_FIN")
    print("=" * 60)
    print("")
    
    result = execute_query_and_save()
    
    if result is not None:
        print("")
        print("Para procesar otro periodo, edita config_fechas.txt")
    else:
        print("")
        print("La ejecucion fallo. Revisa los errores arriba.")
