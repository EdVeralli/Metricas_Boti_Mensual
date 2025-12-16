# -*- coding: utf-8 -*-
'''
Script para ejecutar query de No Entendimiento en Athena
Calcula el porcentaje de no entendimiento (ne + nada) del chatbot
Guarda resultado en CSV y genera Excel con análisis y cálculo de D13

ARCHIVOS DE ENTRADA (carpeta data/):
- testers.csv: Lista de usuarios de prueba a excluir
- Actualizacion_Lista_Blanca.csv: Lista de intenciones válidas

ESTRUCTURA:
No_Entendimiento/
├── data/
│   ├── testers.csv
│   └── Actualizacion_Lista_Blanca.csv
├── output/
└── No_Entendimiento.py

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti
'''
import boto3
import awswrangler as wr
import pandas as pd
import numpy as np
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'data_folder': 'data',
    'config_file': '../config_fechas.txt'
}

INTENT_NADA = 'RuleBuilder:PLBWX5XYGQ2B3GP7IN8Q-alfafc@gmail.com-1536777380652'
THRESHOLD_NE = 5.36

# ==================== FUNCIONES ====================

def read_testers(data_folder):
    '''Lee la lista de testers desde data/testers.csv'''
    testers_file = os.path.join(data_folder, 'testers.csv')
    
    if not os.path.exists(testers_file):
        print("[WARNING] No se encuentra {}".format(testers_file))
        print("          Continuando sin filtro de testers...")
        return []
    
    try:
        df = pd.read_csv(testers_file)
        if len(df.columns) > 0:
            testers = df.iloc[:, 0].astype(str).str.strip().tolist()
            print("[OK] Testers cargados: {} usuarios".format(len(testers)))
            return testers
        else:
            print("[WARNING] Archivo testers.csv vacío")
            return []
    except Exception as e:
        print("[WARNING] Error leyendo testers.csv: {}".format(str(e)))
        return []

def read_lista_blanca(data_folder):
    '''Lee la lista blanca desde data/Actualizacion_Lista_Blanca.csv'''
    lista_file = os.path.join(data_folder, 'Actualizacion_Lista_Blanca.csv')
    
    if not os.path.exists(lista_file):
        print("[WARNING] No se encuentra {}".format(lista_file))
        print("          Continuando sin filtro de lista blanca...")
        return []
    
    try:
        df = pd.read_csv(lista_file)
        if 'Nombre de la intención' in df.columns:
            lista = df['Nombre de la intención'].astype(str).str.strip().tolist()
            print("[OK] Lista blanca cargada: {} intenciones".format(len(lista)))
            return lista
        else:
            print("[WARNING] Columna 'Nombre de la intención' no encontrada")
            return []
    except Exception as e:
        print("[WARNING] Error leyendo Actualizacion_Lista_Blanca.csv: {}".format(str(e)))
        return []

def read_date_config(config_file):
    '''Lee configuración de fechas desde ../config_fechas.txt'''
    try:
        if not os.path.exists(config_file):
            return 'mes', None, None, 10, 2025, "octubre 2025"
        
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                if line.startswith('MES='):
                    mes = int(line.split('=')[1].strip())
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio = int(line.split('=')[1].strip())
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()
        
        if fecha_inicio_str and fecha_fin_str:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')
            descripcion = "{} al {}".format(
                fecha_inicio.strftime('%d/%m/%Y'),
                fecha_fin.strftime('%d/%m/%Y')
            )
            return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion
        
        if mes is not None and anio is not None:
            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)
            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion
        
        return None, None, None, None, None, None
    except Exception as e:
        print("[ERROR] Error leyendo config: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    '''Retorna la abreviatura del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_queries(fecha_inicio, fecha_fin, testers):
    '''Construye las 3 queries para Athena'''
    
    if not testers:
        testers_condition = "1=1"
    else:
        testers_str = "', '".join(testers)
        testers_condition = "substr(session_id, 1, 20) NOT IN ('{}')".format(testers_str)
    
    query_mensajes = f"""
    WITH fecha as (
        SELECT
            cast('{fecha_inicio} 00:00:00' as timestamp) as fecha_inicio,
            cast('{fecha_fin} 23:59:59' as timestamp) as fecha_fin
    ),
    insearch as (
        SELECT
            ins.message_id,
            max(results_score) as max_score
        FROM boti_intent_search ins
        LEFT JOIN fecha on 1=1
        WHERE ts between fecha_inicio and fecha_fin
        GROUP BY ins.message_id
    )
    SELECT DISTINCT 
        mm.session_id,
        mm.id, 
        mm.creation_time, 
        mm.msg_from, 
        mm.message_type, 
        mm.message, 
        mm.rule_name,
        ins.max_score
    FROM boti_message_metrics_2 mm
    LEFT JOIN insearch ins on ins.message_id = mm.id
    LEFT JOIN fecha on 1=1
    WHERE mm.session_creation_time between fecha_inicio and fecha_fin
        AND {testers_condition}
    ORDER BY session_id, creation_time
    """
    
    query_clicks = f"""
    WITH fecha as (
        SELECT
            cast('{fecha_inicio} 00:00:00' as timestamp) as fecha_inicio,
            cast('{fecha_fin} 23:59:59' as timestamp) as fecha_fin
    ),
    intents as (
        SELECT 
            ts, id, session_id, message_id, message,
            results_intent_name, results_intent_id,
            parent_intent_id, parent_intent_name,
            results_showable, results_score as score,
            CASE WHEN parent_intent_id IS NOT NULL AND parent_intent_id != '' 
                THEN parent_intent_id ELSE results_intent_id END as mostrado,
            CASE WHEN parent_intent_name IS NOT NULL AND parent_intent_name != '' 
                THEN parent_intent_name ELSE results_intent_name END as mostrado_name
        FROM boti_intent_search
        LEFT JOIN fecha on 1=1
        WHERE ts >= fecha_inicio AND ts <= fecha_fin
    )
    SELECT 
        a.*, 
        b.ts as response_ts,
        b.response_message,
        b.response_intent_id
    FROM intents a
    LEFT JOIN fecha on 1=1
    LEFT JOIN boti_intent_search_response b ON a.id = b.id
        AND b.ts >= fecha_inicio AND b.ts <= fecha_fin
    WHERE {testers_condition}
    """
    
    query_botones = f"""
    WITH fecha as (
        SELECT
            cast('{fecha_inicio} 00:00:00' as timestamp) as fecha_inicio,
            cast('{fecha_fin} 23:59:59' as timestamp) as fecha_fin
    )
    SELECT  
        ts, id, session_id, message_id, message,
        results_intent_id, results_intent_name, 
        type, one_shot, threshold_clustering, results_index
    FROM boti_intent_search_user_buttons a
    LEFT JOIN fecha on 1=1
    WHERE CAST(concat(concat(concat(concat(year, '-'), month), '-'), day) AS date) 
            between fecha_inicio and fecha_fin
        AND ts between fecha_inicio and fecha_fin
        AND message != ''
        AND {testers_condition}
    """
    
    return query_mensajes, query_clicks, query_botones

def calcular_no_entendimiento(mensajes, clicks, botones):
    '''Calcula D13 = % NE + % Nada'''
    
    print("\n[INFO] Procesando datos...")
    
    mm1 = mensajes.copy()
    mm1['creation_time'] = pd.to_datetime(mm1['creation_time'])
    mm1['usuario'] = mm1['session_id'].str[:20]
    
    search = clicks.copy()
    search['ts'] = pd.to_datetime(search['ts'])
    search['usuario'] = search['session_id'].str[:20]
    
    os = botones.copy()
    os = os[os['one_shot'] == True]
    os['ts'] = pd.to_datetime(os['ts'])
    os['usuario'] = os['session_id'].str[:20]
    
    print("      Mensajes: {:,}".format(len(mm1)))
    print("      Clicks: {:,}".format(len(search)))
    print("      Botones: {:,}".format(len(os)))
    
    search1 = search[search['session_id'].isin(mm1['session_id'].values)]
    os1 = os[os['session_id'].isin(mm1['session_id'].values)]
    
    clicks_validos = search1[
        'RuleBuilder:' + search1['mostrado'] == search1['response_intent_id']
    ]['message_id'].values
    
    os_validos = os1['message_id'].values
    
    primera_instancia1 = search[
        ~search['message_id'].isin(np.concatenate([clicks_validos, os_validos]))
    ].drop_duplicates('id')
    
    ne1_temp = primera_instancia1.groupby('id')['score'].max().reset_index()
    ne1_temp.columns = ['id', 'max_score']
    ne1_ids = ne1_temp[ne1_temp['max_score'] <= THRESHOLD_NE]['id'].values
    
    primera_instancia1 = primera_instancia1[~primera_instancia1['id'].isin(ne1_ids)]
    
    os1_count = len(os1.drop_duplicates('id'))
    click1_count = len(search1[
        'RuleBuilder:' + search1['mostrado'] == search1['response_intent_id']
    ].drop_duplicates('id'))
    abandonos1_count = len(primera_instancia1[primera_instancia1['response_message'].isna()])
    nada1_count = len(primera_instancia1[primera_instancia1['response_intent_id'] == INTENT_NADA])
    texto1_count = len(primera_instancia1[
        (primera_instancia1['response_intent_id'] != INTENT_NADA) & 
        (~primera_instancia1['response_message'].isna())
    ])
    ne1_count = len(ne1_ids)
    
    total = os1_count + click1_count + abandonos1_count + nada1_count + texto1_count + ne1_count
    
    if total > 0:
        porcentaje_ne = ne1_count / total
        porcentaje_nada = nada1_count / total
    else:
        porcentaje_ne = 0
        porcentaje_nada = 0
    
    d13_valor = porcentaje_ne + porcentaje_nada
    
    return {
        'one': os1_count,
        'click': click1_count,
        'abandonos': abandonos1_count,
        'nada': nada1_count,
        'texto': texto1_count,
        'ne': ne1_count,
        'total': total,
        'porcentaje_ne': porcentaje_ne,
        'porcentaje_nada': porcentaje_nada,
        'd13': d13_valor
    }

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    '''Genera nombres de archivos'''
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "no_entendimiento_{0}_{1}.csv".format(mes_nombre, anio)
        filename_excel_detalle = "no_entendimiento_detalle_{0}_{1}.xlsx".format(mes_nombre, anio)
        filename_dashboard = "no_entendimiento_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "no_entendimiento_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_excel_detalle = "no_entendimiento_detalle_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_dashboard = "no_entendimiento_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)
    
    return filename_csv, filename_excel_detalle, filename_dashboard

def create_excel_detalle(filepath, resultados, modo, mes, anio, fecha_inicio, fecha_fin):
    '''Crea Excel con análisis detallado'''
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'No Entendimiento'
    
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    d13_font = Font(bold=True, size=12, color="FFFFFF")
    d13_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    if modo == 'mes':
        header_fecha = '{} {}'.format(get_month_name(mes), anio)
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
        header_fecha = '{} al {}'.format(
            fecha_inicio_obj.strftime('%d/%m/%Y'),
            fecha_fin_obj.strftime('%d/%m/%Y')
        )
    
    ws['A1'] = 'NO ENTENDIMIENTO - Análisis Detallado'
    ws['A1'].font = title_font
    ws['A2'] = 'Período: {}'.format(header_fecha)
    
    row = 4
    ws['A{}'.format(row)] = 'CATEGORÍAS'
    ws['A{}'.format(row)].font = header_font
    row += 1
    
    ws['A{}'.format(row)] = 'OneShots:'
    ws['B{}'.format(row)] = resultados['one']
    row += 1
    ws['A{}'.format(row)] = 'Clicks:'
    ws['B{}'.format(row)] = resultados['click']
    row += 1
    ws['A{}'.format(row)] = 'Abandonos:'
    ws['B{}'.format(row)] = resultados['abandonos']
    row += 1
    ws['A{}'.format(row)] = 'Nada:'
    ws['B{}'.format(row)] = resultados['nada']
    row += 1
    ws['A{}'.format(row)] = 'Texto:'
    ws['B{}'.format(row)] = resultados['texto']
    row += 1
    ws['A{}'.format(row)] = 'NE:'
    ws['B{}'.format(row)] = resultados['ne']
    row += 1
    ws['A{}'.format(row)] = 'TOTAL:'
    ws['B{}'.format(row)] = resultados['total']
    ws['A{}'.format(row)].font = Font(bold=True)
    ws['B{}'.format(row)].font = Font(bold=True)
    row += 2
    
    ws['A{}'.format(row)] = 'PORCENTAJES'
    ws['A{}'.format(row)].font = header_font
    row += 1
    ws['A{}'.format(row)] = '% NE:'
    ws['B{}'.format(row)] = resultados['porcentaje_ne']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    ws['A{}'.format(row)] = '% Nada:'
    ws['B{}'.format(row)] = resultados['porcentaje_nada']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 2
    
    ws['A{}'.format(row)] = 'D13 - NO ENTENDIMIENTO:'
    ws['B{}'.format(row)] = resultados['d13']
    ws['B{}'.format(row)].number_format = '0.00%'
    ws['A{}'.format(row)].font = Font(bold=True, size=12)
    ws['B{}'.format(row)].font = d13_font
    ws['B{}'.format(row)].fill = d13_fill
    ws['B{}'.format(row)].alignment = Alignment(horizontal='center')
    
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    wb.save(filepath)

def create_or_update_dashboard_master(filepath, d13_valor, modo, mes, anio, fecha_inicio, fecha_fin):
    '''Crea o actualiza Dashboard Master (celda D13)'''
    
    if os.path.exists(filepath):
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws['D13'] = d13_valor
        ws['D13'].number_format = '0.00%'
        wb.save(filepath)
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dashboard'
        
        if modo == 'mes':
            header_fecha = '{}-{}'.format(get_month_abbr(mes), str(anio)[-2:])
        else:
            fecha_inicio_obj = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin_obj = datetime.strptime(fecha_fin, '%Y-%m-%d')
            header_fecha = '{}-{}'.format(
                fecha_inicio_obj.strftime('%d/%m'),
                fecha_fin_obj.strftime('%d/%m/%y')
            )
        
        header_font = Font(bold=True)
        ws['B1'] = 'Indicador'
        ws['C1'] = 'Descripción/Detalle'
        ws['D1'] = header_fecha
        ws['B1'].font = header_font
        ws['C1'].font = header_font
        ws['D1'].font = header_font
        
        ws['B13'] = 'No entendimiento'
        ws['C13'] = 'Performance motor de búsqueda'
        ws['D13'] = d13_valor
        ws['D13'].number_format = '0.00%'
        
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 15
        
        wb.save(filepath)

def execute_query_and_save():
    '''Función principal'''
    
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = read_date_config(CONFIG['config_file'])
    
    if modo is None:
        print("[ERROR] No se pudo leer configuracion. Abortando.")
        return None
    
    print("[OK] Periodo: {}".format(descripcion))
    
    print("\nLeyendo archivos de data/...")
    data_folder = CONFIG['data_folder']
    
    if not os.path.exists(data_folder):
        os.makedirs(data_folder, exist_ok=True)
        print("[INFO] Carpeta data/ creada")
        print("       Coloca testers.csv y Actualizacion_Lista_Blanca.csv alli")
    
    testers = read_testers(data_folder)
    lista_blanca = read_lista_blanca(data_folder)
    
    query_mensajes, query_clicks, query_botones = build_queries(fecha_inicio, fecha_fin, testers)
    
    print("\n" + "=" * 70)
    print("EJECUTANDO QUERIES EN ATHENA...")
    print("=" * 70)
    
    try:
        session = boto3.Session(region_name=CONFIG['region'])
        
        print("\n[1/3] Query Mensajes...")
        df_mensajes = wr.athena.read_sql_query(
            sql=query_mensajes,
            database=CONFIG['database'],
            workgroup=CONFIG['workgroup'],
            boto3_session=session,
            ctas_approach=False,
            unload_approach=False
        )
        print("      [OK] {:,} registros".format(len(df_mensajes)))
        
        print("\n[2/3] Query Clicks...")
        df_clicks = wr.athena.read_sql_query(
            sql=query_clicks,
            database=CONFIG['database'],
            workgroup=CONFIG['workgroup'],
            boto3_session=session,
            ctas_approach=False,
            unload_approach=False
        )
        print("      [OK] {:,} registros".format(len(df_clicks)))
        
        print("\n[3/3] Query Botones...")
        df_botones = wr.athena.read_sql_query(
            sql=query_botones,
            database=CONFIG['database'],
            workgroup=CONFIG['workgroup'],
            boto3_session=session,
            ctas_approach=False,
            unload_approach=False
        )
        print("      [OK] {:,} registros".format(len(df_botones)))
        
        resultados = calcular_no_entendimiento(df_mensajes, df_clicks, df_botones)
        
        print("\n" + "=" * 70)
        print("RESULTADOS")
        print("=" * 70)
        print("  OneShots:      {:,}".format(resultados['one']))
        print("  Clicks:        {:,}".format(resultados['click']))
        print("  Abandonos:     {:,}".format(resultados['abandonos']))
        print("  Nada:          {:,}".format(resultados['nada']))
        print("  Texto:         {:,}".format(resultados['texto']))
        print("  NE:            {:,}".format(resultados['ne']))
        print("  " + "-" * 35)
        print("  TOTAL:         {:,}".format(resultados['total']))
        print("")
        print("  % NE:          {:.2f}%".format(resultados['porcentaje_ne'] * 100))
        print("  % Nada:        {:.2f}%".format(resultados['porcentaje_nada'] * 100))
        print("")
        print("  D13:           {:.2f}%".format(resultados['d13'] * 100))
        print("=" * 70)
        
        filename_csv, filename_excel_detalle, filename_dashboard = generate_filename(
            modo, mes, anio, fecha_inicio, fecha_fin
        )
        output_folder = CONFIG['output_folder']
        os.makedirs(output_folder, exist_ok=True)
        
        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel_detalle = os.path.join(output_folder, filename_excel_detalle)
        local_path_dashboard = os.path.join(output_folder, filename_dashboard)
        
        print("\nGuardando archivos...")
        df_resultados = pd.DataFrame([resultados])
        df_resultados.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        print("  [OK] {}".format(filename_csv))
        
        create_excel_detalle(local_path_excel_detalle, resultados, modo, mes, anio, fecha_inicio, fecha_fin)
        print("  [OK] {}".format(filename_excel_detalle))
        
        create_or_update_dashboard_master(local_path_dashboard, resultados['d13'], modo, mes, anio, fecha_inicio, fecha_fin)
        print("  [OK] {} (D13 = {:.2f}%)".format(filename_dashboard, resultados['d13'] * 100))
        
        print("\n" + "=" * 70)
        print("COMPLETADO")
        print("=" * 70)
        
        return resultados
        
    except Exception as e:
        print("\n[ERROR] {}".format(str(e)))
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    print("\n" + "=" * 70)
    print("NO ENTENDIMIENTO - D13")
    print("=" * 70)
    print("\nCÁLCULO: D13 = % NE + % Nada")
    print("  - NE: score <= 5.36")
    print("  - Nada: sin resultado válido")
    print("\nARCHIVOS DE ENTRADA (carpeta data/):")
    print("  - testers.csv")
    print("  - Actualizacion_Lista_Blanca.csv")
    print("\n" + "=" * 70 + "\n")
    
    execute_query_and_save()
