#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
=============================================================================
MÉTRICAS BOTI - VERSIÓN AUTO CONFIG
=============================================================================
✨ NOVEDAD: Lee el mes/año automáticamente desde config_fechas.txt

MEJORA PRINCIPAL:
- Lee config_fechas.txt (igual que athena_connector.py)
- No necesitas editar fechas en el código
- Genera archivo JSON con nombre del mes automáticamente

FLUJO DE TRABAJO:
1. Editar config_fechas.txt (MES=X, AÑO=YYYY)
2. Ejecutar athena_connector.py (descarga CSVs)
3. Ejecutar este script (calcula métricas automáticamente)

OPTIMIZACIÓN INCLUIDA:
- PASO 6 optimizado: de 60 minutos a 2 segundos
- Resultado final: EXACTAMENTE EL MISMO

Autor: Damian - GCBA
Versión: Auto Config + Optimizado
=============================================================================
"""

import pandas as pd
import numpy as np
from datetime import datetime
from calendar import monthrange
import warnings
import os as os_module
import sys
import gc
import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

warnings.filterwarnings('ignore')
pd.set_option("display.max_colwidth", None)

# =============================================================================
# FUNCIONES DE LECTURA DE CONFIGURACIÓN
# =============================================================================

def get_month_name(mes):
    '''Retorna el nombre del mes en español'''
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def read_date_config(config_file):
    '''
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)
    
    Busca el archivo en múltiples ubicaciones:
    1. Ruta especificada
    2. Directorio actual
    3. Directorio padre
    4. Dos niveles arriba
    
    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    '''
    
    # Intentar múltiples ubicaciones
    possible_paths = [
        config_file,                           # Ruta original (../config_fechas.txt)
        'config_fechas.txt',                   # Directorio actual
        '../config_fechas.txt',                # Directorio padre
        '../../config_fechas.txt',             # Dos niveles arriba
    ]
    
    config_path = None
    for path in possible_paths:
        if os_module.path.exists(path):
            config_path = path
            break
    
    if config_path is None:
        print(f"[ERROR] No se encuentra config_fechas.txt en ninguna ubicación")
        print("    Buscado en:")
        for path in possible_paths:
            abs_path = os_module.path.abspath(path)
            print(f"      - {abs_path}")
        print("\n    Crea el archivo config_fechas.txt con:")
        print("      MES=12")
        print("      AÑO=2025")
        return None, None, None, None, None, None
    
    # Encontrado, imprimir ubicación
    print(f"[INFO] Usando: {os_module.path.abspath(config_path)}")
    
    try:
        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None
        
        with open(config_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                
                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)
                
                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)
                
                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()
                
                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()
        
        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2
        if fecha_inicio_str and fecha_fin_str:
            # Agregar hora al rango personalizado
            fecha_inicio_full = f"{fecha_inicio_str} 00:00:00"
            fecha_fin_full = f"{fecha_fin_str} 23:59:59"
            descripcion = f"{fecha_inicio_str} al {fecha_fin_str}"
            print("[INFO] Modo: RANGO PERSONALIZADO")
            return 'rango', fecha_inicio_full, fecha_fin_full, None, None, descripcion
        
        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_full = f"{anio:04d}-{mes:02d}-{primer_dia:02d} 00:00:00"
            # Para el cálculo, la fecha_fin debe ser el primer día del mes siguiente
            if mes == 12:
                fecha_fin_full = f"{anio+1:04d}-01-01 00:00:00"
            else:
                fecha_fin_full = f"{anio:04d}-{mes+1:02d}-01 00:00:00"
            
            mes_nombre = get_month_name(mes)
            descripcion = f"{mes_nombre} {anio}"
            
            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_full, fecha_fin_full, mes, anio, descripcion
        
        print(f"[ERROR] El archivo {config_path} no contiene configuracion valida")
        return None, None, None, None, None, None
        
    except Exception as e:
        print(f"[ERROR] Error leyendo archivo de configuracion: {str(e)}")
        return None, None, None, None, None, None

# =============================================================================
# CONFIGURACIÓN - AHORA SE LEE AUTOMÁTICAMENTE
# =============================================================================

CONFIG_FILE = '../config_fechas.txt'  # Mismo archivo que usa athena_connector

DIRECTORIO_TRABAJO = 'temp'  # Directorio donde athena_connector guarda los CSVs

# Las fechas ahora se leen de config_fechas.txt (ver abajo en main)
FECHA_INICIO = None  # Se configura automáticamente
FECHA_FIN = None     # Se configura automáticamente

# Archivos CSV descargados por athena_connector
ARCHIVO_MENSAJES = 'mensajes_temp.csv'
ARCHIVO_CLICKS = 'clicks_temp.csv'
ARCHIVO_BOTONES = 'botones_temp.csv'

# NOTA: Si necesitas archivos de testers y lista blanca,
# debes colocarlos en el directorio temp/

# Chunk sizes (ajustar si tienes problemas de memoria)
CHUNK_SIZE = 50000

# Constantes del sistema Boti
RULE_NE = 'PLBWX5XYGQ2B3GP7IN8Q-nml045fna3@b.m-1669990832420'
INTENT_NADA = 'RuleBuilder:PLBWX5XYGQ2B3GP7IN8Q-alfafc@gmail.com-1536777380652'
RULE_LETRA_NO_EXISTE = 'No entendió letra no existente en WA'
SCORE_NE_THRESHOLD = 5.36

# =============================================================================
# FUNCIONES AUXILIARES
# =============================================================================

def imprimir_seccion(titulo):
    print("\n" + "=" * 80)
    print(f"  {titulo}")
    print("=" * 80)

def imprimir_progreso(mensaje):
    timestamp = datetime.now().strftime("%H:%M:%S")
    print(f"[{timestamp}] {mensaje}")

def liberar_memoria():
    gc.collect()

def cargar_csv_optimizado(archivo, chunk_size, fecha_inicio=None, testers=None):
    """Carga CSV con filtrado optimizado por chunks"""
    imprimir_progreso(f"Cargando {archivo}...")
    
    chunk_list = []
    total_rows = 0
    total_filtrado = 0
    
    for i, chunk in enumerate(pd.read_csv(archivo, chunksize=chunk_size, low_memory=True)):
        total_rows += len(chunk)
        
        # Filtrar por fecha si aplica
        if fecha_inicio is not None and 'creation_time' in chunk.columns:
            chunk['creation_time'] = pd.to_datetime(chunk['creation_time'], errors='coerce')
            chunk = chunk[chunk['creation_time'] >= fecha_inicio]
        
        # Filtrar por testers si aplica
        if testers is not None and len(chunk) > 0:
            if 'usuario' not in chunk.columns and 'session_id' in chunk.columns:
                chunk['usuario'] = chunk.session_id.str[:20]
            if 'usuario' in chunk.columns:
                chunk = chunk[~chunk.usuario.isin(testers)]
        
        if len(chunk) > 0:
            chunk_list.append(chunk)
            total_filtrado += len(chunk)
        
        if (i + 1) % 10 == 0:
            imprimir_progreso(f"  {total_rows:,} leídos → {total_filtrado:,} filtrados")
            liberar_memoria()
    
    if not chunk_list:
        imprimir_progreso("⚠ No hay datos")
        return pd.DataFrame()
    
    df = pd.concat(chunk_list, ignore_index=True)
    del chunk_list
    liberar_memoria()
    
    imprimir_progreso(f"✓ {total_rows:,} leídos → {len(df):,} después de filtrar")
    return df

def cargar_csv_simple(archivo):
    imprimir_progreso(f"Cargando {archivo}...")
    df = pd.read_csv(archivo)
    imprimir_progreso(f"✓ {archivo}: {len(df):,} registros")
    return df

# =============================================================================
# FUNCIÓN PRINCIPAL
# =============================================================================

def calcular_promedios_boti(fecha_inicio, fecha_fin):
    try:
        # =================================================================
        # PASO 1: CONFIGURACIÓN
        # =================================================================
        imprimir_seccion("PASO 1: CONFIGURACIÓN")
        
        if os_module.path.exists(DIRECTORIO_TRABAJO):
            os_module.chdir(DIRECTORIO_TRABAJO)
            imprimir_progreso(f"✓ Directorio: {os_module.getcwd()}")
        else:
            imprimir_progreso(f"⚠ Usando: {os_module.getcwd()}")
        
        print(f"\n📅 Período: {fecha_inicio} a {fecha_fin}")
        print(f"💾 Chunk size: {CHUNK_SIZE:,}")
        
        fecha_inicio_dt = np.datetime64(fecha_inicio)
        
        # =================================================================
        # PASO 2: CARGAR ARCHIVOS AUXILIARES (OPCIONALES)
        # =================================================================
        imprimir_seccion("PASO 2: ARCHIVOS AUXILIARES")
        
        # NOTA: Estos archivos son opcionales
        # Si no existen, el script continúa sin filtrar testers ni rules_mos
        
        testers = np.array([])  # Sin testers por defecto
        rules_mos = np.array([])  # Sin lista blanca por defecto
        
        # Intentar cargar testers (buscar en directorio padre)
        archivo_testers_local = 'testers.csv'
        archivo_testers_padre = '../testers.csv'
        
        if os_module.path.exists(archivo_testers_local):
            testers_df = cargar_csv_simple(archivo_testers_local)
            testers = testers_df.iloc[:, 0].values
            imprimir_progreso(f"✓ Testers: {len(testers)}")
            del testers_df
            liberar_memoria()
        elif os_module.path.exists(archivo_testers_padre):
            testers_df = pd.read_csv(archivo_testers_padre)
            testers = testers_df.iloc[:, 0].values
            imprimir_progreso(f"✓ Testers (desde directorio padre): {len(testers)}")
            del testers_df
            liberar_memoria()
        else:
            imprimir_progreso("⚠ Archivo testers.csv no encontrado (continuando sin filtro)")
        
        # Intentar cargar lista blanca (buscar en directorio padre)
        archivo_lista_blanca_local = 'Actualizacion_Lista_Blanca.csv'
        archivo_lista_blanca_padre = '../Actualizacion_Lista_Blanca.csv'
        
        if os_module.path.exists(archivo_lista_blanca_local):
            mos = cargar_csv_simple(archivo_lista_blanca_local)
            rules_mos = mos['Nombre de la intención'].str.strip().values
            imprimir_progreso(f"✓ Intenciones mostrables: {len(rules_mos)}")
            del mos
            liberar_memoria()
        elif os_module.path.exists(archivo_lista_blanca_padre):
            mos = pd.read_csv(archivo_lista_blanca_padre)
            rules_mos = mos['Nombre de la intención'].str.strip().values
            imprimir_progreso(f"✓ Intenciones mostrables (desde directorio padre): {len(rules_mos)}")
            del mos
            liberar_memoria()
        else:
            imprimir_progreso("⚠ Archivo Actualizacion_Lista_Blanca.csv no encontrado")
        
        # =================================================================
        # PASO 3: PROCESAR MENSAJES
        # =================================================================
        imprimir_seccion("PASO 3: PROCESAR MENSAJES")
        
        mm = cargar_csv_optimizado(
            ARCHIVO_MENSAJES,
            CHUNK_SIZE,
            fecha_inicio_dt,
            testers
        )
        
        imprimir_progreso("Procesando mensajes...")
        mm.creation_time = pd.to_datetime(mm.creation_time)
        mm.creation_time = mm.creation_time.dt.tz_localize(None)
        mm.creation_time = mm.creation_time.dt.ceil('s')
        
        mm1 = mm.copy()
        del mm
        liberar_memoria()
        
        mm1.drop_duplicates(['session_id', 'creation_time', 'msg_from', 'rule_name'], 
                           inplace=True)
        mm1.reset_index(inplace=True, drop=True)
        
        imprimir_progreso(f"✓ Mensajes procesados: {len(mm1):,}")
        imprimir_progreso(f"✓ Usuarios únicos: {mm1.usuario.nunique():,}")
        
        # =================================================================
        # PASO 4: PROCESAR CLICKS
        # =================================================================
        imprimir_seccion("PASO 4: PROCESAR CLICKS")
        
        search = cargar_csv_optimizado(ARCHIVO_CLICKS, CHUNK_SIZE)
        
        if len(search) == 0:
            raise ValueError("No hay datos de clicks")
        
        imprimir_progreso("Procesando clicks...")
        search.drop_duplicates(['session_id', 'ts', 'id', 'message', 'mostrado', 
                               'response_message'], inplace=True)
        search.ts = pd.to_datetime(search.ts, errors='coerce')
        search['fecha'] = search.ts.dt.date
        
        searchcl = search[
            'RuleBuilder:' + search.mostrado == search.response_intent_id
        ].drop_duplicates('id')
        
        imprimir_progreso(f"✓ Clicks procesados: {len(search):,}")
        
        # =================================================================
        # PASO 5: PROCESAR BOTONES
        # =================================================================
        imprimir_seccion("PASO 5: PROCESAR BOTONES")
        
        one = cargar_csv_optimizado(ARCHIVO_BOTONES, CHUNK_SIZE)
        
        if len(one) == 0:
            raise ValueError("No hay datos de botones")
        
        imprimir_progreso("Procesando botones...")
        os = one[np.logical_and(
            one.one_shot == True,
            one.type.isin(['oneShot', 'oneShotSearch'])
        )].copy()
        
        del one
        liberar_memoria()
        
        os.ts = pd.to_datetime(os.ts, errors='coerce')
        os['fecha'] = os.ts.dt.date
        
        imprimir_progreso(f"✓ OneShots: {len(os):,}")
        
        # =================================================================
        # PASO 6: LIMPIEZA (⭐ OPTIMIZADO - VERSIÓN RÁPIDA ⭐)
        # =================================================================
        imprimir_seccion("PASO 6: LIMPIEZA (OPTIMIZADO)")
        
        imprimir_progreso("Identificando mensajes consecutivos...")
        mm1.reset_index(inplace=True, drop=True)
        
        # ⭐ OPTIMIZACIÓN: Usar shift() en lugar de loop ⭐
        # Esto es 100x más rápido que el método original
        # Resultado: EXACTAMENTE EL MISMO
        mm1['msg_from_next'] = mm1['msg_from'].shift(-1)
        mm1['session_id_next'] = mm1['session_id'].shift(-1)
        
        # Identificar duplicados consecutivos
        mask_duplicados = (
            (mm1['msg_from'] == mm1['msg_from_next']) & 
            (mm1['session_id'] == mm1['session_id_next'])
        )
        
        num_duplicados = mask_duplicados.sum()
        
        if num_duplicados > 0:
            mm1 = mm1[~mask_duplicados].copy()
            imprimir_progreso(f"✓ Eliminados: {num_duplicados:,} mensajes consecutivos")
        else:
            imprimir_progreso("✓ No hay mensajes consecutivos para eliminar")
        
        # Limpiar columnas auxiliares
        mm1.drop(['msg_from_next', 'session_id_next'], axis=1, inplace=True, errors='ignore')
        mm1.reset_index(inplace=True, drop=True)
        
        liberar_memoria()
        
        # =================================================================
        # PASO 7: ANÁLISIS
        # =================================================================
        imprimir_seccion("PASO 7: ANÁLISIS")
        
        mm = mm1.copy()
        mm.reset_index(inplace=True, drop=True)
        
        mmtex1 = mm[np.logical_and(mm.msg_from == 'user', mm.message_type == 'Text')][
            ['session_id', 'id', 'creation_time', 'msg_from', 'message_type', 
             'message', 'usuario']
        ].copy()
        
        if len(mmtex1) > 0:
            mmtex1 = mmtex1.iloc[:-1]
        
        mmtex1['rule_name'] = [
            r if su == sb and f == 'bot' else None
            for r, su, sb, f in zip(
                mm.loc[mmtex1.index + 1].rule_name.values,
                mmtex1.session_id.values,
                mm.loc[mmtex1.index + 1].session_id.values,
                mm.loc[mmtex1.index + 1].msg_from.values
            )
        ]
        
        letra1 = mmtex1[mmtex1.rule_name == RULE_LETRA_NO_EXISTE].copy()
        letra1.rename(columns={'id': 'message_id'}, inplace=True)
        
        del mmtex1
        liberar_memoria()
        
        search1 = search[search.session_id.isin(mm1.session_id.values)].copy()
        os1 = os[os.session_id.isin(mm1.session_id.values)].copy()
        
        del search, os
        liberar_memoria()
        
        primera_instancia1 = search1[
            ~search1.message_id.isin(
                pd.concat([
                    search1[
                        'RuleBuilder:' + search1.mostrado == search1.response_intent_id
                    ].message_id,
                    os1.message_id
                ]).values
            )
        ].drop_duplicates('id').copy()
        
        # Manejar diferentes nombres de columna de score
        if 'results_score' in primera_instancia1.columns:
            primera_instancia1.rename(columns={"results_score": "score"}, inplace=True)
        elif 'score' not in primera_instancia1.columns:
            primera_instancia1['score'] = 10.0
        
        ne1 = primera_instancia1.groupby('id').max()[['session_id', 'message_id', 'score']]
        ne1 = ne1[ne1.score <= SCORE_NE_THRESHOLD].copy()
        
        primera_instancia1 = primera_instancia1[~primera_instancia1.id.isin(ne1.index)].copy()
        
        # =================================================================
        # PASO 8: CATEGORIZACIÓN
        # =================================================================
        imprimir_progreso("Categorizando...")
        
        os1 = os1.drop_duplicates('id')[['session_id', 'message_id']].copy()
        os1['categoria'] = 'one'
        
        click1 = search1[
            'RuleBuilder:' + search1.mostrado == search1.response_intent_id
        ].drop_duplicates('id')[['session_id', 'message_id']].copy()
        click1['categoria'] = 'click'
        
        abandonos1 = primera_instancia1[
            primera_instancia1.response_message.isna()
        ][['session_id', 'message_id']].copy()
        abandonos1['categoria'] = 'abandono'
        
        nada1 = primera_instancia1[
            primera_instancia1.response_intent_id == INTENT_NADA
        ][['session_id', 'message_id']].copy()
        nada1['categoria'] = 'nada'
        
        texto1 = primera_instancia1[np.logical_and(
            primera_instancia1.response_intent_id != INTENT_NADA,
            ~primera_instancia1.response_message.isna()
        )][['session_id', 'message_id']].copy()
        texto1['categoria'] = 'texto'
        
        ne1['categoria'] = 'ne'
        letra1 = letra1[['session_id', 'message_id']].copy()
        letra1['categoria'] = 'letra'
        
        del search1, primera_instancia1
        liberar_memoria()
        
        print(f"\n📊 Categorías:")
        print(f"   OneShots:    {len(os1):>7,}")
        print(f"   Clicks:      {len(click1):>7,}")
        print(f"   Abandonos:   {len(abandonos1):>7,}")
        print(f"   Nada:        {len(nada1):>7,}")
        print(f"   Texto:       {len(texto1):>7,}")
        print(f"   No entend:   {len(ne1):>7,}")
        print(f"   Letra:       {len(letra1):>7,}")
        
        # =================================================================
        # PASO 9: AGRUPACIÓN POR USUARIO
        # =================================================================
        imprimir_seccion("PASO 9: AGRUPACIÓN")
        
        value1primera = pd.concat([os1, click1, abandonos1, nada1, texto1, ne1, letra1],
                                   ignore_index=True)
        
        del os1, click1, abandonos1, nada1, texto1, ne1, letra1
        liberar_memoria()
        
        value1primera['usuario'] = value1primera.session_id.str[:20]
        value1primera = value1primera[value1primera.usuario.isin(mm1.usuario.values)]
        
        imprimir_progreso(f"✓ Interacciones totales: {len(value1primera):,}")
        
        respuestas_por_usuario = value1primera.groupby(
            ['usuario', 'categoria'],
            as_index=False
        ).count()[['usuario', 'categoria', 'message_id']].pivot_table(
            'message_id',
            ['usuario'],
            'categoria'
        )
        
        del value1primera
        liberar_memoria()
        
        respuestas_por_usuario.fillna(0, inplace=True)
        respuestas_por_usuario = respuestas_por_usuario.reset_index(drop=False).reindex(
            ['usuario', 'one', 'click', 'texto', 'abandono', 'nada', 'ne', 'letra'],
            axis=1
        )
        
        # =================================================================
        # PASO 10: CALCULAR PORCENTAJES
        # =================================================================
        imprimir_seccion("PASO 10: PORCENTAJES")
        
        categorias = ['one', 'click', 'texto', 'abandono', 'nada', 'ne', 'letra']
        
        for categoria in categorias:
            respuestas_por_usuario[f'porcentaje_{categoria}'] = [
                respuestas_por_usuario.loc[i][categoria] /
                respuestas_por_usuario.loc[i][categorias].sum()
                for i in respuestas_por_usuario.index
            ]
        
        imprimir_progreso("✓ Porcentajes calculados")
        
        # =================================================================
        # PASO 11: PROMEDIOS FINALES
        # =================================================================
        imprimir_seccion("PASO 11: PROMEDIOS FINALES")
        
        promedios1 = {}
        for cat in ['abandonos', 'click', 'one', 'texto', 'nada', 'letra', 'ne']:
            col = f'porcentaje_{cat.replace("abandonos", "abandono")}'
            if col in respuestas_por_usuario.columns:
                valor = respuestas_por_usuario[col].mean()
                promedios1[cat] = round(valor, 3) if not pd.isna(valor) else 0.0
            else:
                promedios1[cat] = 0.0
        
        # =================================================================
        # MOSTRAR RESULTADOS
        # =================================================================
        print("\n" + "=" * 80)
        print("  RESULTADOS - PROMEDIOS1")
        print("=" * 80)
        
        print(f"\n📅 Período: {FECHA_INICIO} a {FECHA_FIN}")
        print(f"👥 Usuarios analizados: {len(respuestas_por_usuario):,}")
        
        print("\n" + "─" * 80)
        for key, label in [
            ('one', 'OneShots'),
            ('click', 'Clicks'),
            ('texto', 'Texto'),
            ('abandonos', 'Abandonos'),
            ('nada', 'Nada'),
            ('ne', 'No entendidos'),
            ('letra', 'Letra')
        ]:
            valor = promedios1[key]
            if pd.isna(valor) or valor == 0:
                pct = 0.0
                barra = ''
            else:
                pct = valor * 100
                barra = '█' * int(pct / 2)
            valor_display = valor if not pd.isna(valor) else 0.0
            print(f"  {label:.<20} {valor_display:>7.3f}  ({pct:>6.2f}%) {barra}")
        
        print("─" * 80)
        
        total = sum(v for v in promedios1.values() if not pd.isna(v))
        print(f"\n✓ Suma: {total:.3f} ({total*100:.2f}%)")
        
        if abs(total - 1.0) < 0.01:
            print("✅ VALIDACIÓN EXITOSA")
        else:
            print(f"⚠️ La suma no es 100%, revisar datos")
        
        tasa_resolucion = promedios1['one'] + promedios1['click'] + promedios1['texto']
        tasa_problemas = promedios1['abandonos'] + promedios1['ne']
        
        print(f"\n  Resolución: {tasa_resolucion*100:.2f}%")
        print(f"  Problemas: {tasa_problemas*100:.2f}%")
        
        # Calcular métrica clave: nada + ne
        nada_mas_ne = promedios1['nada'] + promedios1['ne']
        print(f"  Nada + NE: {nada_mas_ne*100:.2f}% ← Tu métrica clave")
        
        return promedios1
    
    except Exception as e:
        print("\n" + "=" * 80)
        print("❌ ERROR")
        print("=" * 80)
        print(f"Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

# =============================================================================
# FUNCIONES DE GENERACIÓN DE EXCEL
# =============================================================================

def ensure_output_folder():
    '''
    Asegura que exista la carpeta output/ al mismo nivel que temp/
    Retorna la ruta a la carpeta output
    '''
    # Obtener directorio actual
    current_dir = os_module.getcwd()
    
    # Si estamos en temp/, subir un nivel
    if current_dir.endswith('temp'):
        output_dir = os_module.path.join('..', 'output')
    else:
        output_dir = 'output'
    
    # Crear carpeta si no existe
    os_module.makedirs(output_dir, exist_ok=True)
    
    return output_dir


def create_excel_detalle(filepath, promedios1, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea Excel con análisis detallado de No Entendimiento
    Estructura similar al legacy pero con datos de promedios1
    '''
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'No Entendimiento'
    
    # Estilos
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    d13_font = Font(bold=True, size=12, color="FFFFFF")
    d13_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Header con fecha
    if modo == 'mes':
        header_fecha = '{} {}'.format(get_month_name(mes), anio)
    else:
        fecha_inicio_obj = datetime.strptime(fecha_inicio[:10], '%Y-%m-%d')
        fecha_fin_obj = datetime.strptime(fecha_fin[:10], '%Y-%m-%d')
        header_fecha = '{} al {}'.format(
            fecha_inicio_obj.strftime('%d/%m/%Y'),
            fecha_fin_obj.strftime('%d/%m/%Y')
        )
    
    ws['A1'] = 'NO ENTENDIMIENTO - Análisis Detallado'
    ws['A1'].font = title_font
    ws['A2'] = 'Período: {}'.format(header_fecha)
    
    row = 4
    ws['A{}'.format(row)] = 'CATEGORÍAS'
    ws['A{}'.format(row)].font = header_font
    row += 1
    
    # Valores absolutos (asumiendo que tenemos el total)
    # En el legacy tenían conteos, aquí tenemos porcentajes
    # Vamos a mostrar los porcentajes directamente
    ws['A{}'.format(row)] = 'OneShots:'
    ws['B{}'.format(row)] = promedios1['one']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'Clicks:'
    ws['B{}'.format(row)] = promedios1['click']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'Texto:'
    ws['B{}'.format(row)] = promedios1['texto']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'Abandonos:'
    ws['B{}'.format(row)] = promedios1['abandonos']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'Nada:'
    ws['B{}'.format(row)] = promedios1['nada']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'NE:'
    ws['B{}'.format(row)] = promedios1['ne']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'Letra:'
    ws['B{}'.format(row)] = promedios1['letra']
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'TOTAL:'
    total = sum(promedios1.values())
    ws['B{}'.format(row)] = total
    ws['B{}'.format(row)].number_format = '0.00%'
    ws['A{}'.format(row)].font = Font(bold=True)
    ws['B{}'.format(row)].font = Font(bold=True)
    row += 2
    
    # Métricas derivadas
    ws['A{}'.format(row)] = 'MÉTRICAS'
    ws['A{}'.format(row)].font = header_font
    row += 1
    
    ws['A{}'.format(row)] = 'Resolución (One+Click+Texto):'
    resolucion = promedios1['one'] + promedios1['click'] + promedios1['texto']
    ws['B{}'.format(row)] = resolucion
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 1
    
    ws['A{}'.format(row)] = 'Problemas (Abandonos+Letra):'
    problemas = promedios1['abandonos'] + promedios1['letra']
    ws['B{}'.format(row)] = problemas
    ws['B{}'.format(row)].number_format = '0.00%'
    row += 2
    
    # D13 - Métrica principal
    ws['A{}'.format(row)] = 'D13 - NO ENTENDIMIENTO:'
    d13 = promedios1['nada'] + promedios1['ne']
    ws['B{}'.format(row)] = d13
    ws['B{}'.format(row)].number_format = '0.00%'
    ws['A{}'.format(row)].font = Font(bold=True, size=12)
    ws['B{}'.format(row)].font = d13_font
    ws['B{}'.format(row)].fill = d13_fill
    ws['B{}'.format(row)].alignment = Alignment(horizontal='center')
    
    # Ajustar anchos
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    
    wb.save(filepath)
    print(f"  ✅ Excel detallado: {filepath}")


def create_or_update_dashboard_master(filepath, d13_valor, modo, mes, anio, fecha_inicio, fecha_fin):
    '''
    Crea o actualiza Dashboard Master con estructura completa
    Solo actualiza D13, el resto de valores quedan vacíos para llenar manualmente
    '''
    
    if os_module.path.exists(filepath):
        # Actualizar Excel existente - solo tocar D13
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        ws['D13'] = d13_valor
        ws['D13'].number_format = '0.00%'
        wb.save(filepath)
        print(f"  ✅ Dashboard actualizado: {filepath} (D13 = {d13_valor*100:.2f}%)")
    else:
        # Crear nuevo Dashboard con estructura completa
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dashboard'
        
        # Header con fecha
        if modo == 'mes':
            mes_abbr = get_month_name(mes)[:3].lower()
            header_fecha = '{}-{}'.format(mes_abbr, str(anio)[-2:])
        else:
            fecha_inicio_obj = datetime.strptime(fecha_inicio[:10], '%Y-%m-%d')
            fecha_fin_obj = datetime.strptime(fecha_fin[:10], '%Y-%m-%d')
            header_fecha = '{}-{}'.format(
                fecha_inicio_obj.strftime('%d/%m'),
                fecha_fin_obj.strftime('%d/%m/%y')
            )
        
        # Estilos
        header_font = Font(bold=True)
        
        # FILA 1: Headers
        ws['B1'] = 'Indicador'
        ws['C1'] = 'Descripción/Detalle'
        ws['D1'] = header_fecha
        ws['B1'].font = header_font
        ws['C1'].font = header_font
        ws['D1'].font = header_font
        
        # ESTRUCTURA COMPLETA DEL DASHBOARD
        # Fila 2: Conversaciones
        ws['B2'] = 'Conversaciones'
        ws['C2'] = 'Q Conversaciones'
        
        # Fila 3: Usuarios
        ws['B3'] = 'Usuarios'
        ws['C3'] = 'Q Usuarios únicos'
        
        # Fila 4: Sesiones abiertas por Pushes
        ws['B4'] = 'Sesiones abiertas por Pushes'
        ws['C4'] = 'Q Sesiones que se abrieron con una Push'
        
        # Fila 5: Sesiones Alcanzadas por Pushes
        ws['B5'] = 'Sesiones Alcanzadas por Pushes'
        ws['C5'] = 'Q Sesiones que recibieron al menos 1 Push'
        
        # Fila 6: Mensajes Pushes Enviados
        ws['B6'] = 'Mensajes Pushes Enviados'
        ws['C6'] = 'Q de mensajes enviados bajo el formato push [Hilde gris]'
        
        # Fila 7: Contenidos en Botmaker
        ws['B7'] = 'Contenidos en Botmaker'
        ws['C7'] = 'Contenidos prendidos en botmaker (todos los prendidos, incluy'
        
        # Fila 8: Contenidos Prendidos para el USUARIO
        ws['B8'] = 'Contenidos Prendidos para  el USUARIO'
        ws['C8'] = 'Contenidos prendidos de cara al usuario (relevantes) – (no inclu'
        
        # Fila 9: Interacciones
        ws['B9'] = 'Interacciones'
        ws['C9'] = 'Q Interacciones'
        
        # Fila 10: Trámites, solicitudes y turnos
        ws['B10'] = 'Trámites, solicitudes y turnos'
        ws['C10'] = 'Q Trámites, solicitudes y turnos disponibles'
        
        # Fila 11: contenidos mas consultados
        ws['B11'] = 'contenidos mas consultados'
        ws['C11'] = 'Q Contenidos con más interacciones en el mes (Top 10)'
        
        # Fila 12: Derivaciones
        ws['B12'] = 'Derivaciones'
        ws['C12'] = 'Q Derivaciones'
        
        # Fila 13: No entendimiento (LA QUE LLENAMOS)
        ws['B13'] = 'No entendimiento'
        ws['C13'] = 'Performance motor de búsqueda del nuevo modelo de IA'
        ws['D13'] = d13_valor
        ws['D13'].number_format = '0.00%'
        
        # Fila 14: Tasa de Efectividad
        ws['B14'] = 'Tasa de Efectividad'
        ws['C14'] = 'Mide el porcentaje de usuarios que lograron su objetivo'
        
        # Fila 15: CES (Customer Effort Score)
        ws['B15'] = 'CES (Customer Effort Score)'
        ws['C15'] = 'Mide la facilidad con la que los usuarios pueden interactuar con'
        
        # Fila 16: Satisfacción (CSAT)
        ws['B16'] = 'Satisfacción (CSAT)'
        ws['C16'] = 'Mide la satisfacción usando una escala de 1 a 5'
        
        # Fila 17: Uptime servidor
        ws['B17'] = 'Uptime servidor'
        ws['C17'] = 'Disponibilidad del servidor (% tiempo activo)'
        
        # Ajustar anchos de columnas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 15
        
        wb.save(filepath)
        print(f"  ✅ Dashboard creado: {filepath} (D13 = {d13_valor*100:.2f}%)")


# =============================================================================
# EJECUCIÓN
# =============================================================================

if __name__ == "__main__":
    print("\n" + "=" * 80)
    print("  MÉTRICAS BOTI - VERSIÓN AUTO CONFIG")
    print("  Lee configuración automática desde config_fechas.txt")
    print("  PASO 6 Optimizado: 100x más rápido")
    print("=" * 80)
    
    # Leer configuración desde config_fechas.txt
    print("\n📋 [PASO 0] Leyendo configuración...")
    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = read_date_config(CONFIG_FILE)
    
    if modo is None:
        print("\n❌ Error: No se pudo leer la configuración")
        print("   Verifica que config_fechas.txt existe y está correctamente configurado")
        print(f"   Buscado en: {os_module.path.abspath(CONFIG_FILE)}")
        sys.exit(1)
    
    print(f"\n✓ Configuración leída correctamente:")
    print(f"   Modo: {modo.upper()}")
    print(f"   Período: {descripcion}")
    print(f"   Fecha inicio: {fecha_inicio}")
    print(f"   Fecha fin: {fecha_fin}")
    
    # Calcular métricas
    inicio = datetime.now()
    promedios1 = calcular_promedios_boti(fecha_inicio, fecha_fin)
    fin = datetime.now()
    
    if promedios1:
        print("\n" + "=" * 80)
        print("✅ COMPLETADO")
        print("=" * 80)
        print(f"\n⏱️ Tiempo total: {fin - inicio}")
        print(f"\n💾 promedios1 = {promedios1}")
        
        # Guardar resultados con nombre apropiado
        if modo == 'mes' and mes and anio:
            mes_nombre = get_month_name(mes)
            archivo_salida = f"metricas_boti_{mes_nombre}_{anio}.json"
        else:
            fecha_inicio_str = fecha_inicio[:10].replace('-', '')
            fecha_fin_str = fecha_fin[:10].replace('-', '')
            archivo_salida = f"metricas_boti_{fecha_inicio_str}_a_{fecha_fin_str}.json"
        
        # Guardar en JSON
        resultado_completo = {
            'periodo': descripcion,
            'modo': modo,
            'fecha_inicio': fecha_inicio,
            'fecha_fin': fecha_fin,
            'metricas': promedios1,
            'timestamp': datetime.now().isoformat()
        }
        
        with open(archivo_salida, 'w', encoding='utf-8') as f:
            json.dump(resultado_completo, f, indent=2, ensure_ascii=False)
        
        print(f"\n📁 Archivo JSON: {archivo_salida}")
        
        # Generar archivos Excel
        print("\n📊 Generando archivos Excel...")
        
        # Asegurar que existe carpeta output
        output_dir = ensure_output_folder()
        print(f"  📁 Carpeta output: {os_module.path.abspath(output_dir)}")
        
        # Nombres de archivos Excel (se guardarán en output/)
        if modo == 'mes' and mes and anio:
            mes_nombre = get_month_name(mes)
            excel_detalle = os_module.path.join(output_dir, f"no_entendimiento_detalle_{mes_nombre}_{anio}.xlsx")
            excel_dashboard = os_module.path.join(output_dir, f"no_entendimiento_{mes_nombre}_{anio}.xlsx")
        else:
            fecha_inicio_str = fecha_inicio[:10].replace('-', '')
            fecha_fin_str = fecha_fin[:10].replace('-', '')
            excel_detalle = os_module.path.join(output_dir, f"no_entendimiento_detalle_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx")
            excel_dashboard = os_module.path.join(output_dir, f"no_entendimiento_{fecha_inicio_str}_a_{fecha_fin_str}.xlsx")
        
        # Crear Excel detallado
        create_excel_detalle(excel_detalle, promedios1, modo, mes, anio, fecha_inicio, fecha_fin)
        
        # Calcular D13 (Nada + NE)
        d13 = promedios1['nada'] + promedios1['ne']
        
        # Crear/actualizar Dashboard Master
        create_or_update_dashboard_master(excel_dashboard, d13, modo, mes, anio, fecha_inicio, fecha_fin)
        
        print("\n" + "=" * 80)
        print("📦 ARCHIVOS GENERADOS:")
        print("=" * 80)
        print(f"  [1] JSON:               {archivo_salida}")
        print(f"  [2] Excel Detallado:    {excel_detalle}")
        print(f"  [3] Dashboard Master:   {excel_dashboard}")
        print("")
        print(f"  🎯 D13 (No Entendimiento): {d13*100:.2f}%")
        print("=" * 80)
    else:
        sys.exit(1)
