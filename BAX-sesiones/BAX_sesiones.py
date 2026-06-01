# -*- coding: utf-8 -*-
"""
Script para ejecutar query de SESIONES BAX (canal webchat - BAX - App) en Athena.

Replica el patron estandar de los otros scripts del repo:
  - Lee configuracion de fechas desde ../config_fechas.txt
  - Filtra por: WHERE CAST(session_creation_time AS DATE) BETWEEN date '...' AND date '...'
  - Guarda CSV con datos crudos (year/month/day/channel_id/channel_name/Cant_sess)
  - Guarda Excel con total + serie diaria

MODOS SOPORTADOS:
  1. MES COMPLETO: Especificar MES y AÑO en config_fechas.txt
  2. RANGO PERSONALIZADO: Especificar FECHA_INICIO y FECHA_FIN

Workgroup: Production-caba-piba-athena-boti-group
Rol: PIBAConsumeBoti

aws-azure-login --configure --profile default
aws-azure-login --profile default --mode=gui
"""
import boto3
import awswrangler as wr
import pandas as pd
from datetime import datetime
from calendar import monthrange
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ==================== CONFIGURACION ====================
CONFIG = {
    'region': 'us-east-1',
    'workgroup': 'Production-caba-piba-athena-boti-group',
    'database': 'caba-piba-consume-zone-db',
    'output_folder': 'output',
    'config_file': '../config_fechas.txt'  # Config centralizado en raiz del proyecto
}

# ==================== FUNCIONES ====================

def read_date_config(config_file):
    """
    Lee el archivo de configuracion y determina el modo:
    - MODO 1: MES + AÑO (mes completo)
    - MODO 2: FECHA_INICIO + FECHA_FIN (rango personalizado)

    Retorna: (modo, fecha_inicio, fecha_fin, mes, anio, descripcion)
    """
    try:
        if not os.path.exists(config_file):
            print("[ERROR] No se encuentra el archivo: {}".format(config_file))
            return None, None, None, None, None, None

        mes = None
        anio = None
        fecha_inicio_str = None
        fecha_fin_str = None

        with open(config_file, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                if line.startswith('MES='):
                    mes_str = line.split('=')[1].strip()
                    mes = int(mes_str)

                if line.startswith('AÑO=') or line.startswith('ANO='):
                    anio_str = line.split('=')[1].strip()
                    anio = int(anio_str)

                if line.startswith('FECHA_INICIO='):
                    fecha_inicio_str = line.split('=')[1].strip()

                if line.startswith('FECHA_FIN='):
                    fecha_fin_str = line.split('=')[1].strip()

        # PRIORIDAD: Si hay FECHA_INICIO y FECHA_FIN, usar MODO 2 (rango personalizado)
        if fecha_inicio_str and fecha_fin_str:
            try:
                fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
                fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d')

                if fecha_inicio > fecha_fin:
                    print("[ERROR] FECHA_INICIO no puede ser posterior a FECHA_FIN")
                    return None, None, None, None, None, None

                descripcion = "{} al {}".format(
                    fecha_inicio.strftime('%d/%m/%Y'),
                    fecha_fin.strftime('%d/%m/%Y')
                )

                print("[INFO] Modo: RANGO PERSONALIZADO")
                return 'rango', fecha_inicio_str, fecha_fin_str, None, None, descripcion

            except ValueError as e:
                print("[ERROR] Formato de fecha invalido. Use YYYY-MM-DD (ej: 2026-01-01)")
                print("    Error: {}".format(str(e)))
                return None, None, None, None, None, None

        # Si no hay rango, usar MODO 1 (mes completo)
        if mes is not None and anio is not None:
            if mes < 1 or mes > 12:
                print("[ERROR] Mes invalido: {}. Debe estar entre 1 y 12".format(mes))
                return None, None, None, None, None, None

            if anio < 2020 or anio > 2030:
                print("[ADVERTENCIA] Año inusual: {}".format(anio))

            primer_dia = 1
            ultimo_dia = monthrange(anio, mes)[1]
            fecha_inicio_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, primer_dia)
            fecha_fin_str = "{:04d}-{:02d}-{:02d}".format(anio, mes, ultimo_dia)

            mes_nombre = get_month_name(mes)
            descripcion = "{} {}".format(mes_nombre, anio)

            print("[INFO] Modo: MES COMPLETO")
            return 'mes', fecha_inicio_str, fecha_fin_str, mes, anio, descripcion

        print("[ERROR] El archivo {} no contiene configuracion valida".format(config_file))
        print("    Debe tener MES+AÑO o FECHA_INICIO+FECHA_FIN")
        return None, None, None, None, None, None

    except Exception as e:
        print("[ERROR] Error leyendo archivo de configuracion: {}".format(str(e)))
        return None, None, None, None, None, None

def get_month_name(mes):
    """Retorna el nombre del mes en español"""
    if mes is None:
        return 'rango'
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }
    return meses.get(mes, 'mes_invalido')

def get_month_abbr(mes):
    """Retorna la abreviatura del mes en español"""
    if mes is None:
        return 'rango'
    meses = {
        1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr',
        5: 'may', 6: 'jun', 7: 'jul', 8: 'ago',
        9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
    }
    return meses.get(mes, 'mes')

def build_partition_filter(fecha_inicio, fecha_fin):
    """
    Construye un filtro WHERE basado en las PARTICIONES year/month/day de la tabla.
    Replica exactamente el criterio del Excel original que usa estas particiones
    (NO session_creation_time).

    Las particiones son strings SIN leading zeros (ej: month='3', day='10').
    Para rangos parciales de mes se castea day a INTEGER asi el BETWEEN funciona
    numericamente (sino '9' > '10' lexicograficamente).
    """
    inicio = datetime.strptime(fecha_inicio, '%Y-%m-%d')
    fin = datetime.strptime(fecha_fin, '%Y-%m-%d')

    # CASO 1: mismo año, mismo mes
    if inicio.year == fin.year and inicio.month == fin.month:
        # Si abarca el mes completo (1 al ultimo dia), no necesito filtro de day
        ultimo_dia_mes = monthrange(inicio.year, inicio.month)[1]
        if inicio.day == 1 and fin.day == ultimo_dia_mes:
            return "year='{0}' AND month='{1}'".format(inicio.year, inicio.month)
        return ("year='{0}' AND month='{1}' "
                "AND CAST(day AS INTEGER) BETWEEN {2} AND {3}").format(
            inicio.year, inicio.month, inicio.day, fin.day
        )

    # CASO 2: rango cruzando varios meses --> armar OR de bloques mensuales
    bloques = []
    cur_year, cur_month = inicio.year, inicio.month
    while (cur_year, cur_month) <= (fin.year, fin.month):
        es_primero = (cur_year == inicio.year and cur_month == inicio.month)
        es_ultimo = (cur_year == fin.year and cur_month == fin.month)
        ultimo_dia_mes = monthrange(cur_year, cur_month)[1]

        if es_primero and inicio.day > 1:
            bloque = ("(year='{0}' AND month='{1}' "
                      "AND CAST(day AS INTEGER) >= {2})").format(
                cur_year, cur_month, inicio.day
            )
        elif es_ultimo and fin.day < ultimo_dia_mes:
            bloque = ("(year='{0}' AND month='{1}' "
                      "AND CAST(day AS INTEGER) <= {2})").format(
                cur_year, cur_month, fin.day
            )
        else:
            # mes completo
            bloque = "(year='{0}' AND month='{1}')".format(cur_year, cur_month)

        bloques.append(bloque)

        # avanzar al mes siguiente
        if cur_month == 12:
            cur_year += 1
            cur_month = 1
        else:
            cur_month += 1

    return '(' + ' OR '.join(bloques) + ')'

def build_query(fecha_inicio, fecha_fin):
    """
    Construye la query de Sesiones BAX (canal webchat - BAX - App).

    IMPORTANTE: el filtro de fecha usa las PARTICIONES year/month/day para
    replicar exactamente el numero que da el Excel original (que tambien
    filtra por esas particiones, no por session_creation_time).
    """
    filtro_particion = build_partition_filter(fecha_inicio, fecha_fin)

    query = """SELECT year, month, day, channel_id, channel_name,
       count(distinct s.session_id) AS Cant_sess
FROM "caba-piba-consume-zone-db"."boti_session_metrics_2" s
WHERE {filtro_particion}
  AND s.channel_id LIKE '%webchat%'
  AND s.channel_name LIKE '%BAX - App%'
GROUP BY year, month, day, channel_id, channel_name
ORDER BY year, month, day ASC""".format(filtro_particion=filtro_particion)

    return query

def generate_filename(modo, mes, anio, fecha_inicio, fecha_fin):
    """Genera los nombres de archivos basados en el modo y las fechas"""
    if modo == 'mes':
        mes_nombre = get_month_name(mes)
        filename_csv = "bax_sesiones_{0}_{1}.csv".format(mes_nombre, anio)
        filename_excel = "bax_sesiones_{0}_{1}.xlsx".format(mes_nombre, anio)
    else:  # modo == 'rango'
        fecha_inicio_fmt = fecha_inicio.replace('-', '')
        fecha_fin_fmt = fecha_fin.replace('-', '')
        filename_csv = "bax_sesiones_{0}_a_{1}.csv".format(fecha_inicio_fmt, fecha_fin_fmt)
        filename_excel = "bax_sesiones_{0}_a_{1}.xlsx".format(fecha_inicio_fmt, fecha_fin_fmt)

    return filename_csv, filename_excel

def create_excel(filepath, df, total_sesiones, descripcion, modo, mes, anio, fecha_inicio, fecha_fin):
    """
    Crea un Excel con:
      - Hoja 'Resumen': total de sesiones BAX del periodo
      - Hoja 'Detalle diario': serie diaria (Fecha, Channel, Cant_sess)
      - Hoja 'Crudo': mismas filas que el CSV (year/month/day/channel_id/channel_name/Cant_sess)
    """
    print("    [INFO] Creando Excel...")

    wb = openpyxl.Workbook()

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    bold_font = Font(bold=True, size=11)
    card_font = Font(bold=True, size=12)
    card_value_font = Font(bold=True, size=14)

    # ===== HOJA 1: Resumen =====
    ws = wb.active
    ws.title = 'Resumen'

    ws['A1'] = 'Sesiones BAX (webchat - BAX - App)'
    ws['A1'].font = card_font
    ws['B1'] = int(total_sesiones)
    ws['B1'].font = card_value_font
    ws['B1'].number_format = '#,##0'

    ws['A2'] = 'Periodo: {}'.format(descripcion)
    ws['A2'].font = Font(size=10, italic=True)

    ws['A3'] = 'Fecha inicio'
    ws['B3'] = fecha_inicio
    ws['A4'] = 'Fecha fin'
    ws['B4'] = fecha_fin

    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20

    # ===== HOJA 2: Detalle diario =====
    ws2 = wb.create_sheet('Detalle diario')

    # Construir DataFrame diario (sumando por fecha por si hubiera multiples channel_id/name)
    if len(df) > 0:
        df_dia = df.copy()
        # Armar columna fecha desde year/month/day
        df_dia['Fecha'] = pd.to_datetime(
            df_dia['year'].astype(str).str.zfill(4) + '-' +
            df_dia['month'].astype(str).str.zfill(2) + '-' +
            df_dia['day'].astype(str).str.zfill(2),
            errors='coerce'
        )
        df_dia_agg = df_dia.groupby('Fecha', as_index=False)['cant_sess'].sum() \
            if 'cant_sess' in df_dia.columns else \
            df_dia.groupby('Fecha', as_index=False)['Cant_sess'].sum()
        # Normalizar nombre de columna
        df_dia_agg.columns = ['Fecha', 'Sesiones diarias']
        df_dia_agg = df_dia_agg.sort_values('Fecha').reset_index(drop=True)

        # Headers
        ws2['A1'] = 'Fecha'
        ws2['B1'] = 'Sesiones diarias'
        ws2['A1'].font = header_font
        ws2['A1'].fill = header_fill
        ws2['B1'].font = header_font
        ws2['B1'].fill = header_fill

        for idx, row in df_dia_agg.iterrows():
            fila = 2 + idx
            ws2['A{}'.format(fila)] = row['Fecha']
            ws2['A{}'.format(fila)].number_format = 'DD/MM/YYYY'
            ws2['B{}'.format(fila)] = int(row['Sesiones diarias'])
            ws2['B{}'.format(fila)].number_format = '#,##0'

        # Total
        fila_total = 2 + len(df_dia_agg)
        ws2['A{}'.format(fila_total)] = 'TOTAL'
        ws2['A{}'.format(fila_total)].font = bold_font
        ws2['B{}'.format(fila_total)] = int(df_dia_agg['Sesiones diarias'].sum())
        ws2['B{}'.format(fila_total)].font = bold_font
        ws2['B{}'.format(fila_total)].number_format = '#,##0'

        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 20

    # ===== HOJA 3: Crudo (mismas columnas que el CSV) =====
    ws3 = wb.create_sheet('Crudo')

    if len(df) > 0:
        columnas = list(df.columns)
        for col_idx, col_name in enumerate(columnas, start=1):
            cell = ws3.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            for c_idx, col_name in enumerate(columnas, start=1):
                value = row[col_name]
                cell = ws3.cell(row=r_idx, column=c_idx, value=value)
                # Formato de miles para la columna de cantidad
                if col_name.lower() == 'cant_sess':
                    try:
                        cell.value = int(value)
                        cell.number_format = '#,##0'
                    except (TypeError, ValueError):
                        pass

        # Anchos basicos
        for col_idx, col_name in enumerate(columnas, start=1):
            letra = openpyxl.utils.get_column_letter(col_idx)
            ws3.column_dimensions[letra].width = max(12, min(40, len(str(col_name)) + 2))

    wb.save(filepath)
    print("    [OK] Excel creado: {}".format(filepath))

def check_aws_credentials():
    """Verifica que las credenciales AWS esten configuradas y sean validas"""
    try:
        sts = boto3.client('sts', region_name=CONFIG['region'])
        identity = sts.get_caller_identity()

        user_arn = identity.get('Arn', '')

        print("[OK] Credenciales AWS validas")
        print("    ARN: {}".format(user_arn))

        if 'PIBAConsumeBoti' not in user_arn:
            print("")
            print("[ADVERTENCIA] No estas usando el rol correcto")
            print("    Se requiere: PIBAConsumeBoti")
            if '/' in user_arn:
                current_role = user_arn.split('/')[-2]
            else:
                current_role = 'desconocido'
            print("    Tu rol actual: {}".format(current_role))
            print("")
            print("SOLUCION:")
            print("    1. Ejecuta: aws-azure-login --profile default --mode=gui")
            print("    2. Cuando te autentiques, SELECCIONA el rol: PIBAConsumeBoti")
            print("    3. Vuelve a ejecutar este script")
            print("")
            return False

        return True

    except Exception as e:
        print("[ERROR] Error verificando credenciales: {}".format(str(e)))
        if 'ExpiredToken' in str(e):
            print("")
            print("SOLUCION:")
            print("    Tu sesión AWS expiró. Ejecuta:")
            print("    aws-azure-login --profile default --mode=gui")
            print("")
        else:
            print("")
            print("SOLUCION:")
            print("    1. Ejecuta: aws-azure-login --configure --profile default")
            print("    2. Luego: aws-azure-login --profile default --mode=gui")
            print("")
        return False

def execute_query_and_save():
    """Funcion principal: ejecuta query y guarda resultados"""

    # Verificar credenciales
    print("Verificando credenciales AWS...")
    if not check_aws_credentials():
        return None

    # Leer configuracion de fechas
    print("")
    print("Leyendo configuracion de fechas...")
    result = read_date_config(CONFIG['config_file'])

    if result[0] is None:
        print("[ERROR] No se pudo leer la configuracion de fechas")
        return None

    modo, fecha_inicio, fecha_fin, mes, anio, descripcion = result

    print("[OK] Configuracion leida:")
    print("    Modo: {}".format(modo.upper()))
    print("    Periodo: {}".format(descripcion))
    print("    Fecha inicio: {}".format(fecha_inicio))
    print("    Fecha fin: {}".format(fecha_fin))

    # Construir query
    query = build_query(fecha_inicio, fecha_fin)

    print("")
    print("Configuracion AWS:")
    print("    Region: {}".format(CONFIG['region']))
    print("    Workgroup: {}".format(CONFIG['workgroup']))
    print("    Base de datos: {}".format(CONFIG['database']))

    print("")
    print("Query a ejecutar:")
    print("    {}".format(query))

    try:
        session = boto3.Session(region_name=CONFIG['region'])

        print("")
        print("Ejecutando consulta...")

        try:
            df = wr.athena.read_sql_query(
                sql=query,
                database=CONFIG['database'],
                workgroup=CONFIG['workgroup'],
                boto3_session=session,
                ctas_approach=False,
                unload_approach=False
            )
        except Exception as e:
            if 'workgroup' in str(e).lower() or 'GetWorkGroup' in str(e):
                print("")
                print("[ADVERTENCIA] Error con workgroup '{}'".format(CONFIG['workgroup']))
                print("    Intentando sin especificar workgroup...")

                df = wr.athena.read_sql_query(
                    sql=query,
                    database=CONFIG['database'],
                    boto3_session=session,
                    ctas_approach=False,
                    unload_approach=False
                )
            else:
                raise e

        print("")
        print("[OK] Consulta ejecutada exitosamente!")

        if len(df) == 0:
            print("[ADVERTENCIA] La query no retorno resultados para el periodo solicitado.")
            return None

        # Normalizar nombre de la columna de cantidad
        # (Athena/Presto puede devolver lowercase: 'cant_sess')
        col_count = None
        for candidate in ['Cant_sess', 'cant_sess', 'CANT_SESS']:
            if candidate in df.columns:
                col_count = candidate
                break
        if col_count is None:
            print("[ERROR] No se encontro la columna de cantidad de sesiones en el resultado")
            print("    Columnas: {}".format(df.columns.tolist()))
            return None

        # Renombrar a un nombre canonico
        if col_count != 'Cant_sess':
            df = df.rename(columns={col_count: 'Cant_sess'})

        total_sesiones = int(df['Cant_sess'].sum())

        print("")
        print("=" * 60)
        print("RESULTADO - {}".format(descripcion.upper()))
        print("=" * 60)
        print("Filas devueltas: {:,}".format(len(df)))
        print("TOTAL SESIONES BAX (webchat - BAX - App): {:,}".format(total_sesiones))
        print("=" * 60)

        # Generar nombres de archivo
        filename_csv, filename_excel = generate_filename(modo, mes, anio, fecha_inicio, fecha_fin)
        output_folder = CONFIG['output_folder']

        os.makedirs(output_folder, exist_ok=True)

        local_path_csv = os.path.join(output_folder, filename_csv)
        local_path_excel = os.path.join(output_folder, filename_excel)

        # Guardar CSV
        print("")
        print("Guardando CSV...")
        df.to_csv(local_path_csv, index=False, encoding='utf-8-sig')
        print("    [OK] CSV guardado: {}".format(filename_csv))

        # Crear Excel
        print("Generando Excel...")
        create_excel(local_path_excel, df, total_sesiones, descripcion, modo, mes, anio, fecha_inicio, fecha_fin)

        print("")
        print("ARCHIVOS GENERADOS:")
        print("    Carpeta: {}/".format(output_folder))
        print("")
        print("    [CSV] Nombre: {}".format(filename_csv))
        print("          Ruta: {}".format(os.path.abspath(local_path_csv)))
        print("          Tamaño: {:,} bytes".format(os.path.getsize(local_path_csv)))
        print("")
        print("    [EXCEL] Nombre: {}".format(filename_excel))
        print("            Ruta: {}".format(os.path.abspath(local_path_excel)))
        print("            Tamaño: {:,} bytes".format(os.path.getsize(local_path_excel)))
        print("            Hojas: Resumen / Detalle diario / Crudo")
        print("            Total BAX: {:,}".format(total_sesiones))

        print("")
        print("=" * 60)
        print("PROCESO COMPLETADO EXITOSAMENTE")
        print("=" * 60)

        return df

    except Exception as e:
        print("")
        print("[ERROR] ERROR DURANTE LA EJECUCION")
        print("    Tipo: {}".format(type(e).__name__))
        print("    Mensaje: {}".format(str(e)))

        error_str = str(e).lower()

        print("")
        print("DIAGNOSTICO:")
        if 'table' in error_str and 'not' in error_str:
            print("    [!] La tabla no existe o no tienes permisos para accederla")
            print("    Verifica acceso a: boti_session_metrics_2")
        elif 'workgroup' in error_str:
            print("    [!] Problema con el workgroup")
        elif 'permission' in error_str or 'denied' in error_str:
            print("    [!] Problema de permisos")
        elif 'openpyxl' in error_str:
            print("    [!] Falta libreria openpyxl para generar Excel")
            print("    Ejecuta: pip install openpyxl")
        elif 'timeout' in error_str or 'timed out' in error_str:
            print("    [!] La query tomo demasiado tiempo")
        else:
            print("    [!] Error inesperado")

        import traceback
        traceback.print_exc()
        return None

# ==================== EJECUCION PRINCIPAL ====================

if __name__ == "__main__":
    print("")
    print("=" * 60)
    print("  IMPORTANTE: LOGIN AWS REQUERIDO")
    print("=" * 60)
    print("  Antes de continuar, asegurate de estar logueado en AWS.")
    print("  Si tu token expiro, ejecuta en otra terminal:")
    print("")
    print("      aws-azure-login --profile default --mode=gui")
    print("")
    print("=" * 60)
    try:
        input("  Presiona ENTER para continuar (o Ctrl+C para cancelar)...")
    except KeyboardInterrupt:
        print("")
        print("  Cancelado por el usuario.")
        raise SystemExit(0)
    print("")

    print("=" * 60)
    print("SCRIPT: SESIONES BAX (webchat - BAX - App) - QUERY ATHENA")
    print("=" * 60)
    print("Lee configuracion desde: {}".format(CONFIG['config_file']))
    print("Rol requerido: PIBAConsumeBoti")
    print("Tabla: boti_session_metrics_2")
    print("Filtro: channel_id LIKE '%webchat%' AND channel_name LIKE '%BAX - App%'")
    print("Filtro fecha: PARTICIONES year/month/day (igual que Excel original)")
    print("Salida: CSV crudo + Excel (Resumen / Detalle diario / Crudo)")
    print("")
    print("MODOS:")
    print("  [1] MES COMPLETO: MES + AÑO")
    print("  [2] RANGO PERSONALIZADO: FECHA_INICIO + FECHA_FIN")
    print("=" * 60)
    print("")

    result = execute_query_and_save()

    if result is not None:
        print("")
        print("Para procesar otro periodo, edita ../config_fechas.txt")
    else:
        print("")
        print("[ERROR] El proceso no se completo correctamente. Revisa los mensajes anteriores.")
